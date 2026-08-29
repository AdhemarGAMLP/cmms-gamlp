# excel_utils.py
import os
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Font, Alignment
import pythoncom
import win32com.client

def obtener_ruta_plantilla(nombre_archivo):
    """
    Busca la plantilla Excel por nombre en todas las ubicaciones posibles.
    Orden de prioridad:
      1. Junto al .exe instalado                 (Inno Setup / Carpeta de instalación)
      2. Subcarpeta _internal junto al .exe      (PyInstaller onedir)
      3. Carpeta temporal _MEIPASS              (PyInstaller onefile con --add-data)
      4. Carpeta del proyecto en desarrollo      (C:\\Users\\...\\Desktop\\HEAS_CMMS\\plantillas)
      5. Carpeta de Datos de Gestión             (C:\\Users\\...\\Desktop\\Datos_De_Gestion_HEAS\\plantillas)
      6. Variable de entorno HEAS_PLANTILLAS_DIR (ruta personalizada)
    """
    import sys
    from config import BASE_DIR

    candidatos = []

    if hasattr(sys, 'frozen'):
        dir_exe = os.path.dirname(sys.executable)
        # 1. Junto al .exe instalado: C:\Program Files\HEAS GESTION\plantillas\
        candidatos.append(os.path.join(dir_exe, "plantillas", nombre_archivo))
        # 2. Subcarpeta _internal: C:\Program Files\HEAS GESTION\_internal\plantillas\
        candidatos.append(os.path.join(dir_exe, "_internal", "plantillas", nombre_archivo))
        # 3. Un nivel arriba por si el exe está en una subcarpeta
        candidatos.append(os.path.join(os.path.dirname(dir_exe), "plantillas", nombre_archivo))

    # 4. Carpeta temporal PyInstaller onefile (_MEIPASS)
    if hasattr(sys, '_MEIPASS'):
        candidatos.append(os.path.join(sys._MEIPASS, "plantillas", nombre_archivo))
        candidatos.append(os.path.join(sys._MEIPASS, nombre_archivo))

    # 5. Entorno de desarrollo / Código fuente
    ruta_script = os.path.dirname(os.path.abspath(__file__))
    candidatos.append(os.path.join(ruta_script, "plantillas", nombre_archivo))
    candidatos.append(os.path.join(os.path.expanduser("~"), "Desktop", "HEAS_CMMS", "plantillas", nombre_archivo))

    # 6. Carpeta en Datos_De_Gestion_HEAS
    candidatos.append(os.path.join(BASE_DIR, "plantillas", nombre_archivo))

    # 7. Variable de entorno personalizada
    env_dir = os.environ.get("HEAS_PLANTILLAS_DIR", "")
    if env_dir:
        candidatos.append(os.path.join(env_dir, nombre_archivo))

    for ruta in candidatos:
        if os.path.exists(ruta):
            return ruta

    # Si no se encontró, devolver el candidato principal para mostrar el mensaje de error
    return candidatos[0] if candidatos else os.path.join(ruta_script, "plantillas", nombre_archivo)



def _celda_real(ws, coordenada):
    """Resuelve rangos combinados devolviendo la celda física superior izquierda correspondiente."""
    col_letter, row_str = coordinate_from_string(coordenada)
    row = int(row_str)
    col = column_index_from_string(col_letter)
    for merged_range in ws.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1) and \
           col in range(merged_range.min_col, merged_range.max_col + 1):
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return ws[coordenada]

def escribir_en_celda_segura(ws, coordenada, valor, centrar=False, negrita=False):
    """Escribe un valor forzando alineación o peso de fuente si la plantilla carece de estilos."""
    celda = _celda_real(ws, coordenada)
    celda.value = valor
    if centrar:
        current_wrap = celda.alignment.wrap_text if celda.alignment else False
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=current_wrap)
    if negrita:
        f = celda.font
        celda.font = Font(name=f.name if f else 'Calibri', size=f.size if f else 11, bold=True, color=f.color if f else '000000')

def marcar_x(ws, coordenada):
    """Inyecta una X de verificación centrada y en negrita para simular un checkbox."""
    celda = _celda_real(ws, coordenada)
    celda.value = 'X'
    celda.font = Font(name='Calibri', bold=True, color='000000', size=celda.font.size if celda.font else 12)
    celda.alignment = Alignment(horizontal='center', vertical='center')

def escribir_texto_largo(ws, coordenada, valor):
    """Formatea la escritura de bloques de texto largos asegurando el ajuste de líneas superior-izquierdo."""
    celda = _celda_real(ws, coordenada)
    celda.value = valor
    celda.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

def exportar_excel_a_pdf(ruta_excel, ruta_pdf, rango_impresion=None, landscape=False):
    """
    Renderiza hojas de Excel a PDF mediante Windows COM API.
    Si rango_impresion es None, respeta estrictamente el área configurada en la plantilla de origen.
    """
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(ruta_excel, UpdateLinks=False, ReadOnly=True)
        ws = wb.ActiveSheet

        try:
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = 1
            if rango_impresion:
                ws.PageSetup.PrintArea = rango_impresion
            ws.PageSetup.CenterHorizontally = True
            
            # Forzar márgenes pequeños y orientación vertical para tamaño completo de hoja
            ws.PageSetup.LeftMargin = 10
            ws.PageSetup.RightMargin = 10
            ws.PageSetup.TopMargin = 10
            ws.PageSetup.BottomMargin = 10
            if landscape:
                ws.PageSetup.Orientation = 2 # xlLandscape
            else:
                ws.PageSetup.Orientation = 1 # xlPortrait
        except Exception:
            pass

        ws.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        wb.Close(False)
        excel.Quit()
        return True
    except Exception as e:
        print(f"❌ Error al exportar a PDF: {e}")
        return False
    finally:
        pythoncom.CoUninitialize()