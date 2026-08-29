# main.py
import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import psycopg2.extras
import socket
import os
import qrcode
import json
from PIL import Image
import fitz  # PyMuPDF
import pythoncom
import win32com.client
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import threading

# Importar constantes de diseño y backend centralizado
from estilos import *
from config import CONFIG, CARPETAS
from database import inicializar_bd, obtener_conexion
from auth import inicializar_usuarios, login
from web_server import arrancar_hilo_web
from excel_utils import obtener_ruta_plantilla, escribir_en_celda_segura, exportar_excel_a_pdf

# Importar las vistas modulares del subpaquete vistas
from vistas.inventario import VistaInventario
from vistas.catalogo import VistaCatalogo
from vistas.repuestos import VistaRepuestos
from vistas.cronograma import VistaCronograma
from vistas.historial import VistaHistorial
from vistas.protocolos import VistaProtocolos

# ========================================================
# INTERFAZ DE LOGIN
# ========================================================
class VentanaLogin(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("HEAS CMMS - Iniciar Sesión")
        self.geometry("400x350")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        self.usuario_autenticado = None

        ctk.CTkLabel(self, text="HEAS Biomédica", font=ctk.CTkFont(size=26, weight="bold"), text_color=C_BLUE).pack(pady=(35, 5))
        ctk.CTkLabel(self, text="Sistema de Gestión Hospitalaria", font=ctk.CTkFont(size=12), text_color=C_SUBTEXT).pack(pady=(0, 20))
        
        marco = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=14)
        marco.pack(padx=30, pady=5, fill="both", expand=True)

        self.e_user = ctk.CTkEntry(marco, placeholder_text="Usuario (admin)", width=280, corner_radius=8)
        self.e_user.pack(pady=(25, 10))
        
        self.e_pass = ctk.CTkEntry(marco, placeholder_text="Contraseña", show="*", width=280, corner_radius=8)
        self.e_pass.pack(pady=10)
        self.e_pass.bind("<Return>", lambda e: self.intentar_login())
        
        ctk.CTkButton(marco, text="Ingresar al Sistema", command=self.intentar_login, height=42, corner_radius=8, font=ctk.CTkFont(weight="bold", size=13), fg_color=C_BLUE, hover_color=C_BLUE_HOVER).pack(pady=20)

    def intentar_login(self):
        u = login(self.e_user.get(), self.e_pass.get())
        if u:
            self.usuario_autenticado = u
            self.destroy()
        else:
            messagebox.showerror("Acceso Denegado", "Usuario o contraseña incorrectos.")

# ========================================================
# NÚCLEO PRINCIPAL DEL SOFTWARE
# ========================================================
class SistemaMantenimiento(ctk.CTk):
    def __init__(self, usuario):
        super().__init__()
        self.usuario_actual = usuario
        self.es_jefe = usuario.get("rol") == "jefe"
        
        self.title(f"HEAS CMMS - Hospital El Alto Sur | Electromedicina (Rol: {usuario['rol'].upper()})")
        self.geometry("1300x800")
        self.after(0, lambda: self.state('zoomed'))
        ctk.set_appearance_mode("light")
        self.configure(fg_color=C_BG)

        self.ip_local = self.obtener_ip_local()
        self.alertas_activas = []
        self._calendario_sucio = True
        
        arrancar_hilo_web(self)
        self.cargar_datos_memoria()
        self.configurar_estilo_ttk()
        self.crear_interfaz_base()
        
        self.vistas = {}
        self.crear_vistas_modulares()
        self.datos_sucios = False
        self.chequear_datos_sucios()
        self.mostrar_vista("Inventario")
        self.actualizar_boton_alertas()
        self.verificar_y_ejecutar_backup_auto()
        self.protocol("WM_DELETE_WINDOW", self.al_cerrar_aplicacion)
        self.sidebar_mode = None
        self.bind("<Configure>", self.al_redimensionar)

    def obtener_ip_local(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "127.0.0.1"

    def configurar_estilo_ttk(self):
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background=C_CARD, foreground=C_TEXT, rowheight=40, fieldbackground=C_CARD, borderwidth=0, font=('Segoe UI', 10))
        style.map('Treeview', background=[('selected', C_BLUE)], foreground=[('selected', 'white')])
        style.configure("Treeview.Heading", background=C_BG, foreground=C_SUBTEXT, font=('Segoe UI', 11, 'bold'), borderwidth=0)
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        # Diseño y comportamiento de Scrollbar minimalista y moderno estilo CustomTkinter
        style.layout("Vertical.TScrollbar", [
            ('Vertical.Scrollbar.trough', {
                'children': [
                    ('Vertical.Scrollbar.thumb', {'expand': '1', 'sticky': 'nswe'})
                ],
                'sticky': 'ns'
            })
        ])
        style.configure("Vertical.TScrollbar", 
                        background="#8F8F93", 
                        troughcolor=C_CARD, 
                        bordercolor=C_CARD, 
                        thickness=9, 
                        relief="flat")
        style.map("Vertical.TScrollbar", 
                  background=[('pressed', '#5A5A5C'), ('active', '#707073')])

        style.layout("Horizontal.TScrollbar", [
            ('Horizontal.Scrollbar.trough', {
                'children': [
                    ('Horizontal.Scrollbar.thumb', {'expand': '1', 'sticky': 'nswe'})
                ],
                'sticky': 'ew'
            })
        ])
        style.configure("Horizontal.TScrollbar", 
                        background="#8F8F93", 
                        troughcolor=C_CARD, 
                        bordercolor=C_CARD, 
                        thickness=9, 
                        relief="flat")
        style.map("Horizontal.TScrollbar", 
                  background=[('pressed', '#5A5A5C'), ('active', '#707073')])

    def chequear_datos_sucios(self):
        if getattr(self, "datos_sucios", False):
            self.datos_sucios = False
            print("[INFO] Recibida actualización desde Web Server. Recargando datos en GUI...")
            self.cargar_datos_memoria()
            for nombre_v, vista_v in self.vistas.items():
                if hasattr(vista_v, 'refrescar_datos'):
                    vista_v.refrescar_datos()
            if "Cronograma" in self.vistas:
                try:
                    self.vistas["Cronograma"].dibujar_mes(self.vistas["Cronograma"].anio_actual, self.vistas["Cronograma"].mes_actual)
                    self.vistas["Cronograma"].dibujar_anio(self.vistas["Cronograma"].anio_vista)
                except Exception as e:
                    print("Error al refrescar vistas del cronograma:", e)
        self.after(2000, self.chequear_datos_sucios)

    def cargar_datos_memoria(self):
        self.datos = {"catalogo": [], "repuestos": [], "equipos": [], "protocolos": []}
        self.eventos_calendario = {}
        self.hoy = datetime.now().date()
        self.hora_actual = datetime.now().hour
        self.alertas_activas = []

        conn = obtener_conexion()
        if not conn: 
            return
            
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute("SELECT * FROM catalogo ORDER BY nombre ASC")
        self.datos["catalogo"] = [dict(r) for r in cur.fetchall()]
        
        cur.execute("SELECT * FROM repuestos")
        self.datos["repuestos"] = [dict(r) for r in cur.fetchall()]

        try:
            cur.execute("SELECT * FROM protocolos ORDER BY fecha DESC, turno ASC")
            self.datos["protocolos"] = [dict(r) for r in cur.fetchall()]
        except:
            conn.rollback()

        try:
            cur.execute("SELECT * FROM historial_intervenciones ORDER BY COALESCE(fecha_entrega, fecha) DESC, COALESCE(hora_entrega, '00:00') DESC, id DESC")
            todas_inter = [dict(h) for h in cur.fetchall()]
            hist_por_equipo = {}
            for h in todas_inter:
                hist_por_equipo.setdefault(h['equipo_id'], []).append(h)

            cur.execute("SELECT * FROM equipos")
            eqs_db = [dict(r) for r in cur.fetchall()]

            for eq in eqs_db:
                eq['historial_intervenciones'] = hist_por_equipo.get(eq['id'], [])
                self.datos["equipos"].append(eq)
                
                crit = eq.get("criticidad", "Riesgo Medio")
                meses = 3 if "Alto" in crit else (4 if "Medio" in crit else 6)
                f_reg = eq.get("fecha_registro", self.hoy)
                
                if isinstance(f_reg, datetime): 
                    f_reg = f_reg.date()
                elif isinstance(f_reg, str):
                    try: 
                        f_reg = datetime.strptime(f_reg, "%Y-%m-%d").date()
                    except: 
                        f_reg = self.hoy
                        
                prevs = [m for m in eq.get("historial_intervenciones", []) if m["tipo"] == "Preventivo"]
                if prevs:
                    uf = prevs[0]["fecha"]
                    if isinstance(uf, datetime): 
                        f_reg = uf.date()
                    elif isinstance(uf, str):
                        try: 
                            f_reg = datetime.strptime(uf, "%Y-%m-%d").date()
                        except: 
                            pass
                            
                f_prox = f_reg + relativedelta(months=+meses)
                dias_restantes = (f_prox - self.hoy).days
                estado = "Vencido" if dias_restantes < 0 else ("Por Vencer" if dias_restantes <= 30 else "Al Día")

                if estado == "Vencido":
                    self.alertas_activas.append(f"⚠️ Mantenimiento VENCIDO: {eq['id']} - {eq['nombre']}")
                eq['f_prox'] = f_prox
                if estado == "Por Vencer":
                    self.alertas_activas.append(f"⏳ Próximo a vencer: {eq['id']} - {eq['nombre']} ({dias_restantes} días)")

                self.eventos_calendario.setdefault(f_prox, []).append({'eq': eq['nombre'], 'estado': estado, 'id': eq['id'], 'f_prox': f_prox})
        except:
            pass

        prot_hoy = [p for p in self.datos["protocolos"] if p['fecha'] == self.hoy]
        tipos = ['Gases Medicinales', 'Resonador Magnético']
        turnos_req = []
        if self.hora_actual >= 8: 
            turnos_req.append('Mañana')
        if self.hora_actual >= 14: 
            turnos_req.append('Tarde')
        if self.hora_actual >= 23: 
            turnos_req.append('Noche')
            
        for t_req in turnos_req:
            for tipo in tipos:
                if not any(p['tipo_protocolo'] == tipo and p['turno'] == t_req for p in prot_hoy):
                    self.alertas_activas.append(f"🚨 FALTA PROTOCOLO: {tipo} (Turno: {t_req})")

        cur.close()
        conn.close()

    def crear_interfaz_base(self):
        self.sidebar = ctk.CTkFrame(self, width=250, corner_radius=0, fg_color=C_CARD, border_width=1, border_color=C_BORDER)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        
        ctk.CTkLabel(self.sidebar, text="HEAS\nGestión Clínica", font=ctk.CTkFont(size=22, weight="bold"), text_color=C_TEXT).pack(pady=(30, 5), padx=20)
        ctk.CTkLabel(self.sidebar, text=f"Red QR: {self.ip_local}", font=ctk.CTkFont(size=11), text_color=C_SUBTEXT).pack(pady=(0, 20))

        self.btn_alertas = ctk.CTkButton(self.sidebar, text="🔔 Alertas (0)", height=40, font=ctk.CTkFont(weight="bold"), fg_color=C_BG, text_color=C_TEXT, command=self.mostrar_ventana_alertas)
        self.btn_alertas.pack(pady=(0, 20), padx=15, fill="x")

        btn_estilo = {"fg_color": "transparent", "text_color": C_TEXT, "hover_color": C_BG, "anchor": "center", "height": 45, "font": ctk.CTkFont(size=14, weight="bold")}
        
        self.btn_nav_inv = ctk.CTkButton(self.sidebar, text="📦 Inventario", command=lambda: self.mostrar_vista("Inventario"), **btn_estilo)
        self.btn_nav_inv.pack(pady=2, padx=15, fill="x")
        
        self.btn_nav_cat = ctk.CTkButton(self.sidebar, text="🏷️ Catálogo", command=lambda: self.mostrar_vista("Catalogo"), **btn_estilo)
        self.btn_nav_cat.pack(pady=2, padx=15, fill="x")
        
        self.btn_nav_rep = ctk.CTkButton(self.sidebar, text="🔧 Repuestos", command=lambda: self.mostrar_vista("Repuestos"), **btn_estilo)
        self.btn_nav_rep.pack(pady=2, padx=15, fill="x")
        
        self.btn_nav_cro = ctk.CTkButton(self.sidebar, text="📅 Cronograma", command=lambda: self.mostrar_vista("Cronograma"), **btn_estilo)
        self.btn_nav_cro.pack(pady=2, padx=15, fill="x")
        
        self.btn_nav_hist = ctk.CTkButton(self.sidebar, text="📋 Historial", command=lambda: self.mostrar_vista("Historial"), **btn_estilo)
        self.btn_nav_hist.pack(pady=2, padx=15, fill="x")
        
        self.btn_nav_prot = ctk.CTkButton(self.sidebar, text="📝 Protocolos", command=lambda: self.mostrar_vista("Protocolos"), **btn_estilo)
        self.btn_nav_prot.pack(pady=2, padx=15, fill="x")

        ctk.CTkButton(self.sidebar, text="✚ Mantenimiento", height=45, corner_radius=10, font=ctk.CTkFont(size=13, weight="bold"), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, command=self.modulo_mantenimiento).pack(side="bottom", pady=30, padx=20, fill="x")
        
        self.contenedor_principal = ctk.CTkFrame(self, fg_color=C_BG)
        self.contenedor_principal.pack(side="right", fill="both", expand=True)

    def crear_vistas_modulares(self):
        self.vistas["Inventario"] = VistaInventario(self.contenedor_principal, self)
        self.vistas["Catalogo"] = VistaCatalogo(self.contenedor_principal, self)
        self.vistas["Repuestos"] = VistaRepuestos(self.contenedor_principal, self)
        self.vistas["Cronograma"] = VistaCronograma(self.contenedor_principal, self)
        self.vistas["Historial"] = VistaHistorial(self.contenedor_principal, self)
        self.vistas["Protocolos"] = VistaProtocolos(self.contenedor_principal, self)

    def mostrar_vista(self, nombre):
        for btn in [self.btn_nav_inv, self.btn_nav_cat, self.btn_nav_rep, self.btn_nav_cro, self.btn_nav_hist, self.btn_nav_prot]:
            btn.configure(fg_color="transparent")
            
        for vista in self.vistas.values():
            vista.pack_forget()
            
        self.vistas[nombre].pack(fill="both", expand=True)
        
        if hasattr(self.vistas[nombre], 'refrescar_datos'):
            self.vistas[nombre].refrescar_datos()
            
        if nombre == "Inventario": 
            self.btn_nav_inv.configure(fg_color=C_BORDER)
        elif nombre == "Catalogo": 
            self.btn_nav_cat.configure(fg_color=C_BORDER)
        elif nombre == "Repuestos": 
            self.btn_nav_rep.configure(fg_color=C_BORDER)
        elif nombre == "Historial": 
            self.btn_nav_hist.configure(fg_color=C_BORDER)
        elif nombre == "Protocolos": 
            self.btn_nav_prot.configure(fg_color=C_BORDER)
        elif nombre == "Cronograma":
            self.btn_nav_cro.configure(fg_color=C_BORDER)
            if self._calendario_sucio: 
                self.vistas["Cronograma"].dibujar_mes(self.vistas["Cronograma"].anio_actual, self.vistas["Cronograma"].mes_actual)
                self.vistas["Cronograma"].dibujar_anio(self.vistas["Cronograma"].anio_vista)
                self.vistas["Cronograma"].refrescar_datos()
                self._calendario_sucio = False

    def actualizar_boton_alertas(self):
        c = len(self.alertas_activas)
        self.btn_alertas.configure(text=f"🔔 Alertas ({c})")
        if c > 0: 
            self.btn_alertas.configure(fg_color=C_RED, hover_color=C_RED_HOVER, text_color="white")
        else: 
            self.btn_alertas.configure(fg_color=C_BG, text_color=C_TEXT)

    def mostrar_ventana_alertas(self):
        v = ctk.CTkToplevel(self)
        v.title("Alertas")
        v.geometry("500x500")
        v.transient(self)
        v.grab_set()
        v.configure(fg_color=C_CARD)
        
        txt = ctk.CTkTextbox(v, fg_color=C_BG, font=ctk.CTkFont(size=14))
        txt.pack(fill="both", expand=True, padx=20, pady=20)
        
        if not self.alertas_activas: 
            txt.insert("end", "✅ Todo está al día.")
        else:
            for a in self.alertas_activas: 
                txt.insert("end", f"{a}\n\n")
                
        txt.configure(state="disabled")

    # ========================================================
    # FORMULARIO MAESTRO DE REGISTRO DE EQUIPOS
    # ========================================================
    def abrir_formulario_equipo(self, eq_edit=None):
        vent = ctk.CTkToplevel(self)
        vent.title("Ficha de Equipo")
        vent.geometry("800x680")
        vent.transient(self)
        vent.grab_set()
        vent.configure(fg_color=C_BG)
        
        ctk.CTkLabel(vent, text="Ficha Técnica Institucional", font=ctk.CTkFont(size=22, weight="bold"), text_color=C_TEXT).pack(pady=15)
        
        sf = ctk.CTkScrollableFrame(vent, fg_color=C_CARD, corner_radius=12)
        sf.pack(pady=5, padx=20, fill="both", expand=True)

        ctk.CTkLabel(sf, text="1. Identificación y Ubicación", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(10, 5))
        
        val_cat = [f"{c['nombre']} - {c.get('marca', '')} - {c.get('modelo', '')}" for c in self.datos["catalogo"]]
        combo_tipo = ctk.CTkComboBox(sf, width=500, values=val_cat if val_cat else ["No hay modelos"])
        combo_tipo.pack(pady=5)
        
        e_id = ctk.CTkEntry(sf, placeholder_text="Número de Serie / ID", width=500)
        e_id.pack(pady=5)
        
        e_servicio = ctk.CTkEntry(sf, placeholder_text="Servicio (Ej. Rayos X)", width=500)
        e_servicio.pack(pady=5)
        
        e_area = ctk.CTkEntry(sf, placeholder_text="Ubicación Física", width=500)
        e_area.pack(pady=5)

        ctk.CTkLabel(sf, text="2. Adquisición", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(15, 5))
        
        e_procedencia = ctk.CTkEntry(sf, placeholder_text="Procedencia", width=500)
        e_procedencia.pack(pady=5)
        
        e_fabricante = ctk.CTkEntry(sf, placeholder_text="Fabricante Original", width=500)
        e_fabricante.pack(pady=5)
        
        e_proveedor = ctk.CTkEntry(sf, placeholder_text="Proveedor Local", width=500)
        e_proveedor.pack(pady=5)
        
        e_anio = ctk.CTkEntry(sf, placeholder_text="Año de Fabricación", width=500)
        e_anio.pack(pady=5)
        
        f_cal = ctk.CTkFrame(sf, fg_color="transparent")
        f_cal.pack(pady=5, anchor="w")
        
        ctk.CTkLabel(f_cal, text="Fecha de Instalación:", text_color=C_SUBTEXT).pack(side="left", padx=5)
        cal_adq = DateEntry(f_cal, width=15, font=('Segoe UI', 11), background=C_BLUE, foreground='white', borderwidth=0, date_pattern='y-mm-dd')
        cal_adq.pack(side="left")

        ctk.CTkLabel(sf, text="3. Especificaciones", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(15, 5))
        
        frame_checks = ctk.CTkFrame(sf, fg_color=C_BG, corner_radius=10)
        frame_checks.pack(fill="x", pady=5)

        self.var_electrico = ctk.StringVar()
        self.var_electronico = ctk.StringVar()
        self.var_mecanico = ctk.StringVar()
        self.var_hidraulico = ctk.StringVar()
        self.var_neumatico = ctk.StringVar()
        self.var_vapor = ctk.StringVar()
        
        ft = ctk.CTkFrame(frame_checks, fg_color="transparent")
        ft.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(ft, text="Tecnología:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkCheckBox(ft, text="Eléctrico", variable=self.var_electrico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Electrónico", variable=self.var_electronico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Mecánico", variable=self.var_mecanico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Hidráulico", variable=self.var_hidraulico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Neumático", variable=self.var_neumatico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Vapor", variable=self.var_vapor, onvalue="X", offvalue="").pack(anchor="w", pady=2)

        self.var_compra = ctk.StringVar()
        self.var_comodato = ctk.StringVar()
        self.var_donacion = ctk.StringVar()
        
        fa = ctk.CTkFrame(frame_checks, fg_color="transparent")
        fa.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(fa, text="Adquisición:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkCheckBox(fa, text="Compra", variable=self.var_compra, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fa, text="Comodato", variable=self.var_comodato, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fa, text="Donación", variable=self.var_donacion, onvalue="X", offvalue="").pack(anchor="w", pady=2)

        self.var_fijo = ctk.StringVar()
        self.var_movil = ctk.StringVar()
        self.var_portatil = ctk.StringVar()
        
        fti = ctk.CTkFrame(frame_checks, fg_color="transparent")
        fti.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(fti, text="Tipo:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkCheckBox(fti, text="Fijo", variable=self.var_fijo, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fti, text="Móvil", variable=self.var_movil, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fti, text="Portátil", variable=self.var_portatil, onvalue="X", offvalue="").pack(anchor="w", pady=2)

        self.var_garantia = ctk.StringVar()
        
        fg = ctk.CTkFrame(frame_checks, fg_color="transparent")
        fg.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(fg, text="Garantía:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkRadioButton(fg, text="Sí", variable=self.var_garantia, value="Con Garantía").pack(anchor="w", pady=2)
        ctk.CTkRadioButton(fg, text="No", variable=self.var_garantia, value="Sin Garantía").pack(anchor="w", pady=2)

        ctk.CTkLabel(sf, text="4. Criticidad", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(15, 5))
        
        p_cat = [
            "Intercambiabilidad", 
            "Régimen operación", 
            "Nivel de utilización", 
            "Parámetro principal", 
            "Mantenibilidad", 
            "Conservabilidad", 
            "Grado de automatización", 
            "Valor de compra", 
            "Factibilidad aprovto", 
            "Seguridad operacional", 
            "Condiciones de explot.", 
            "Protección de MA", 
            "Edad del equipo"
        ]
        f_cat = ctk.CTkFrame(sf, fg_color=C_BG, corner_radius=10)
        f_cat.pack(fill="x", pady=5, padx=20)
        
        # Configurar columnas con pesos y anchos mínimos
        f_cat.grid_columnconfigure(0, weight=3, minsize=200)
        f_cat.grid_columnconfigure(1, weight=1, minsize=80)
        f_cat.grid_columnconfigure(2, weight=1, minsize=80)
        f_cat.grid_columnconfigure(3, weight=1, minsize=80)
        
        # Cabecera de columnas para las opciones
        ctk.CTkLabel(f_cat, text="", anchor="w").grid(row=0, column=0, padx=(20, 10), pady=6, sticky="w")
        for col_idx, text in [(1, "I"), (2, "II"), (3, "III")]:
            ctk.CTkLabel(f_cat, text=text, font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).grid(row=0, column=col_idx, sticky="nsew", pady=6)
        
        self.variables_cat = []
        for i, param in enumerate(p_cat):
            ctk.CTkLabel(f_cat, text=param, anchor="w", font=ctk.CTkFont(size=12)).grid(row=i+1, column=0, padx=(20, 10), pady=4, sticky="w")
            var = ctk.StringVar(value="")
            self.variables_cat.append(var)
            
            for col_idx, val in [(1, "1"), (2, "2"), (3, "3")]:
                f_btn = ctk.CTkFrame(f_cat, fg_color="transparent")
                f_btn.grid(row=i+1, column=col_idx, sticky="nsew", pady=4)
                rb = ctk.CTkRadioButton(f_btn, text="", variable=var, value=val, width=20, height=20)
                rb.pack(expand=True)

        combo_estado = ctk.CTkComboBox(sf, values=["Operativo", "Baja"], width=500)
        combo_estado.pack(pady=(20,5))
        
        # FUNCIONALIDAD DE FOTO RESTAURADA CORRECTAMENTE
        ruta_foto = ctk.StringVar(value="")
        
        def seleccionar_foto():
            r = filedialog.askopenfilename(filetypes=[("Imágenes", "*.jpg *.jpeg *.png")])
            if r:
                ruta_foto.set(r)
                btn_foto.configure(text="✅ Fotografía Adjuntada Correctamente", fg_color=C_BLUE, hover_color=C_BLUE_HOVER)
                
        btn_foto = ctk.CTkButton(sf, text="📸 Adjuntar Fotografía del Equipo", command=seleccionar_foto, fg_color=C_BLUE, hover_color=C_BLUE_HOVER, width=500, font=ctk.CTkFont(weight="bold", size=14), height=35)
        btn_foto.pack(pady=10)

        if eq_edit:
            e_id.insert(0, eq_edit["id"])
            e_id.configure(state="disabled")
            
            e_servicio.insert(0, eq_edit.get("servicio",""))
            e_area.insert(0, eq_edit.get("area",""))
            e_procedencia.insert(0, eq_edit.get("procedencia",""))
            e_fabricante.insert(0, eq_edit.get("fabricante",""))
            e_proveedor.insert(0, eq_edit.get("proveedor",""))
            e_anio.insert(0, eq_edit.get("anio_fab",""))
            
            combo_estado.set(eq_edit.get("estado","Operativo"))
            combo_tipo.set(f"{eq_edit['nombre']} - {eq_edit.get('marca','')} - {eq_edit.get('modelo','')}")
            
            self.var_electrico.set(eq_edit.get("t_elec", ""))
            self.var_electronico.set(eq_edit.get("t_elco", ""))
            self.var_mecanico.set(eq_edit.get("t_mec", ""))
            self.var_hidraulico.set(eq_edit.get("t_hid", ""))
            self.var_neumatico.set(eq_edit.get("t_neu", ""))
            self.var_vapor.set(eq_edit.get("t_vap", ""))
            
            self.var_compra.set(eq_edit.get("a_comp", ""))
            self.var_comodato.set(eq_edit.get("a_como", ""))
            self.var_donacion.set(eq_edit.get("a_don", ""))
            
            self.var_fijo.set(eq_edit.get("te_fijo", ""))
            self.var_movil.set(eq_edit.get("te_mov", ""))
            self.var_portatil.set(eq_edit.get("te_por", ""))
            
            self.var_garantia.set(eq_edit.get("garantia", ""))
            
            foto_guardada = eq_edit.get("foto", "")
            if foto_guardada and os.path.exists(foto_guardada):
                ruta_foto.set(foto_guardada)
                btn_foto.configure(text="✅ Fotografía Existente en Sistema", fg_color=C_BLUE, hover_color=C_BLUE_HOVER)
            
            c_det = eq_edit.get("categorizacion_detalle") or []
            if isinstance(c_det, str):
                c_det = json.loads(c_det)
                
            for i, val in enumerate(c_det):
                if i < len(self.variables_cat):
                    if val in ("I","II","III"):
                        newval = "3" if val == "I" else ("2" if val == "II" else "1")
                        self.variables_cat[i].set(newval)
                    else:
                        self.variables_cat[i].set(str(val))

        def guardar():
            if not e_id.get(): 
                return
                
            ts = combo_tipo.get().split(" - ")
            n_nom = ts[0]
            n_mar = ts[1] if len(ts)>1 else ""
            n_mod = ts[2] if len(ts)>2 else ""
            
            puntajes = []
            for v in self.variables_cat:
                if v.get():
                    puntajes.append(int(v.get()))
                else:
                    puntajes.append(0)
                    
            puntaje_total = sum(puntajes)
            
            if puntaje_total >= 30:
                criticidad_final = "Riesgo Alto"
            elif puntaje_total >= 20:
                criticidad_final = "Riesgo Medio"
            else:
                criticidad_final = "Riesgo Bajo"
                
            detalles_cat = [v.get() for v in self.variables_cat]
            
            try:
                conn = obtener_conexion()
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO equipos (id, nombre, marca, modelo, servicio, area, procedencia, fabricante, proveedor, anio_fab,
                    t_elec, t_elco, t_mec, t_hid, t_neu, t_vap, a_comp, a_como, a_don, te_fijo, te_mov, te_por, garantia, criticidad, categorizacion_detalle, estado, fecha_adquisicion, fecha_registro, foto)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                    nombre=EXCLUDED.nombre, marca=EXCLUDED.marca, modelo=EXCLUDED.modelo, servicio=EXCLUDED.servicio, area=EXCLUDED.area, procedencia=EXCLUDED.procedencia, fabricante=EXCLUDED.fabricante, proveedor=EXCLUDED.proveedor, anio_fab=EXCLUDED.anio_fab,
                    t_elec=EXCLUDED.t_elec, t_elco=EXCLUDED.t_elco, t_mec=EXCLUDED.t_mec, t_hid=EXCLUDED.t_hid, t_neu=EXCLUDED.t_neu, t_vap=EXCLUDED.t_vap, a_comp=EXCLUDED.a_comp, a_como=EXCLUDED.a_como, a_don=EXCLUDED.a_don,
                    te_fijo=EXCLUDED.te_fijo, te_mov=EXCLUDED.te_mov, te_por=EXCLUDED.te_por, garantia=EXCLUDED.garantia, criticidad=EXCLUDED.criticidad, categorizacion_detalle=EXCLUDED.categorizacion_detalle, estado=EXCLUDED.estado, fecha_adquisicion=EXCLUDED.fecha_adquisicion, foto=EXCLUDED.foto;
                """, (e_id.get(), n_nom, n_mar, n_mod, e_servicio.get(), e_area.get(), e_procedencia.get(), e_fabricante.get(), e_proveedor.get(), e_anio.get(),
                      self.var_electrico.get(), self.var_electronico.get(), self.var_mecanico.get(), self.var_hidraulico.get(), self.var_neumatico.get(), self.var_vapor.get(),
                      self.var_compra.get(), self.var_comodato.get(), self.var_donacion.get(), self.var_fijo.get(), self.var_movil.get(), self.var_portatil.get(),
                      self.var_garantia.get(), criticidad_final, json.dumps(detalles_cat), combo_estado.get(), cal_adq.get_date(), datetime.now().strftime("%Y-%m-%d"), ruta_foto.get()))
                
                conn.commit()
                cur.close()
                conn.close()
                
                self.cargar_datos_memoria()
                self.vistas["Inventario"].refrescar_datos()
                self._calendario_sucio = True
                vent.destroy()
            except Exception as e:
                messagebox.showerror("Error SQL", str(e))
                
        ctk.CTkButton(sf, text="Guardar Cambios y Cerrar", font=ctk.CTkFont(weight="bold", size=14), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, height=45, command=guardar).pack(pady=20, padx=20, fill="x")

    # ========================================================
    # PREVISUALIZACIÓN DE HOJA DE VIDA Y EXCEL
    # ========================================================
    def abrir_hoja_vida_click(self, event):
        eq_id = self.vistas["Inventario"].obtener_id_seleccionado()
        if not eq_id: 
            return
        
        eq_act = next((eq for eq in self.datos["equipos"] if str(eq["id"]) == str(eq_id)), None)
        
        if not eq_act: 
            return
        
        v_hv = ctk.CTkToplevel(self)
        v_hv.title(f"Hoja de Vida y Previsualización PDF")
        v_hv.geometry("1200x700")
        v_hv.after(0, lambda: v_hv.state('zoomed'))
        v_hv.transient(self)
        v_hv.configure(fg_color=C_BG)
        
        frame_izq = ctk.CTkFrame(v_hv, fg_color="transparent")
        frame_izq.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        frame_der = ctk.CTkScrollableFrame(v_hv, width=680, fg_color=C_CARD, corner_radius=10)
        frame_der.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        ctk.CTkLabel(frame_der, text="Previsualización del Documento PDF", font=ctk.CTkFont(weight="bold", size=16), text_color=C_TEXT).pack(pady=10)
        
        lbl_preview = ctk.CTkLabel(frame_der, text="⚙️ Generando documento...", text_color=C_SUBTEXT)
        lbl_preview.pack(expand=True, pady=20)
        
        # Sanitizar el ID para evitar caracteres inválidos en rutas de archivos de Windows (ej. '|')
        id_sanitizado = "".join(c if c not in '<>:"/\\|?*' else "_" for c in str(eq_act['id']))
        area_name = eq_act.get("area", "General")
        area_folder = "".join([c for c in area_name if c.isalnum() or c==' ']).strip()
        dir_ficha = os.path.join(CARPETAS["areas"], area_folder, "equipos")
        os.makedirs(dir_ficha, exist_ok=True)
        ruta_excel_guardado = os.path.join(dir_ficha, f"Ficha_Tecnica_{id_sanitizado}.xlsx")
        ruta_excel_absoluta = os.path.abspath(ruta_excel_guardado)
        
        def actualizar_excel_datos_silencioso():
            # CORRECCIÓN DE FOTOGRAFÍA DOBLE Y DESCENTRADA:
            # 1. Cargamos siempre la plantilla limpia original
            ruta_plantilla = obtener_ruta_plantilla("plantilla_ficha.xlsx")
            if not os.path.exists(ruta_plantilla):
                raise FileNotFoundError(f"No se encontró la plantilla en:\n{ruta_plantilla}")
                
            wb = openpyxl.load_workbook(ruta_plantilla)
            hoja = wb.active
            
            
            def escribir(celda, valor):
                try:
                    hoja[celda] = valor
                    if valor == 'X':
                        hoja[celda].font = Font(name='Calibri', bold=True, color='000000', size=11)
                        hoja[celda].alignment = Alignment(horizontal='center', vertical='center')
                except: 
                    pass
                
            escribir('K8', eq_act.get('nombre', ''))
            escribir('H11', eq_act.get('servicio', ''))
            escribir('H12', eq_act.get('area', ''))
            escribir('H13', eq_act.get('marca', ''))
            escribir('H14', eq_act.get('modelo', ''))
            escribir('H15', eq_act['id'])
            escribir('H16', eq_act.get('procedencia', ''))
            escribir('H17', eq_act.get('fabricante', ''))
            escribir('H18', eq_act.get('garantia', ''))
            escribir('H19', eq_act.get('proveedor', ''))
            escribir('H20', 1)
            escribir('H21', eq_act.get('anio_fab', ''))
            escribir('H22', eq_act.get('fecha_adquisicion', ''))
            
            escribir('S25', eq_act.get('t_elec', ''))
            escribir('S27', eq_act.get('t_elco', ''))
            escribir('S29', eq_act.get('t_mec', ''))
            escribir('Z25', eq_act.get('t_hid', ''))
            escribir('Z27', eq_act.get('t_neu', ''))
            escribir('Z29', eq_act.get('t_vap', ''))
            
            escribir('S33', eq_act.get('a_comp', ''))
            escribir('S35', eq_act.get('a_como', ''))
            escribir('S37', eq_act.get('a_don', ''))
            escribir('Y33', eq_act.get('te_fijo', ''))
            escribir('Y35', eq_act.get('te_mov', ''))
            escribir('Y37', eq_act.get('te_por', ''))
            
            cat_data = eq_act.get("categorizacion_detalle") or []
            if isinstance(cat_data, str):
                cat_data = json.loads(cat_data)
                
            for i in range(13):
                valor = str(cat_data[i]) if i < len(cat_data) else ""
                if valor in ("1", "III"): 
                    escribir(f'AO{24+i}', 'X')
                elif valor in ("2", "II"): 
                    escribir(f'AM{24+i}', 'X')
                elif valor in ("3", "I"): 
                    escribir(f'AK{24+i}', 'X')
                
            # 3. Insertamos la foto del equipo médico en AA11 con tamaño 220x220 (centrado perfecto del código anterior)
            foto_path = eq_act.get('foto')
            if foto_path and os.path.exists(foto_path):
                try:
                    img = ExcelImage(foto_path)
                    img.width = 220
                    img.height = 220
                    hoja.add_image(img, 'AA11') 
                except Exception as e:
                    print(f"Aviso: No se pudo inyectar la imagen en el Excel: {e}")

            try:
                wb.save(ruta_excel_guardado)
            except PermissionError:
                raise PermissionError("El archivo Excel está abierto o bloqueado por otra aplicación (como Microsoft Excel). Por favor, cierra Excel antes de continuar.")

        def hilo_generar_preview():
            try:
                pythoncom.CoInitialize()
                actualizar_excel_datos_silencioso()
                
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.ScreenUpdating = False
                
                wb_pdf = excel.Workbooks.Open(ruta_excel_absoluta, UpdateLinks=False, ReadOnly=True)
                ws = wb_pdf.ActiveSheet
                
                # CORRECCIÓN DE PDF CORTADO: Aplicamos configuración de márgenes y área de impresión exactas del código anterior
                try:
                    ws.PageSetup.Zoom = False
                    ws.PageSetup.FitToPagesWide = 1
                    ws.PageSetup.FitToPagesTall = 1
                    ws.PageSetup.PrintArea = "$A$1:$AQ$67"
                    ws.PageSetup.CenterHorizontally = True
                    ws.PageSetup.LeftMargin = 10
                    ws.PageSetup.RightMargin = 10
                    ws.PageSetup.TopMargin = 10
                    ws.PageSetup.BottomMargin = 10
                except: 
                    pass
                
                ruta_pdf_temp = os.path.abspath(os.path.join(dir_ficha, f"~preview_{id_sanitizado}.pdf"))
                if os.path.exists(ruta_pdf_temp):
                    try: 
                        os.remove(ruta_pdf_temp)
                    except: 
                        pass
                    
                ws.ExportAsFixedFormat(0, ruta_pdf_temp)
                wb_pdf.Close(False)
                excel.Quit()
                
                doc = fitz.open(ruta_pdf_temp)
                page = doc.load_page(0)
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                img_ctk = ctk.CTkImage(light_image=img, size=(640, int(img.height * (640 / img.width))))
                
                v_hv.after(0, lambda: lbl_preview.configure(text="", image=img_ctk))
                lbl_preview.image = img_ctk
                doc.close()
            except Exception as e:
                import traceback
                traceback.print_exc()
                err_msg = str(e)
                v_hv.after(0, lambda msg=err_msg: lbl_preview.configure(text=f"Error al generar vista previa: {msg}"))
            finally:
                pythoncom.CoUninitialize()

        threading.Thread(target=hilo_generar_preview, daemon=True).start()
        
        enl = f"http://{self.ip_local}:5000/equipo/{eq_act['id']}"
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(enl)
        qr.make(fit=True)
        img_qr = qr.make_image(fill_color="black", back_color="white").get_image()
        
        m_info = ctk.CTkFrame(frame_izq, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        m_info.pack(pady=20, fill="x")
        
        c_img = ctk.CTkFrame(m_info, fg_color="transparent")
        c_img.pack(side="left", padx=10, pady=10)
        ctk.CTkLabel(c_img, image=ctk.CTkImage(light_image=img_qr, size=(120,120)), text="").pack(pady=5)

        def btn_ver_excel():
            try:
                actualizar_excel_datos_silencioso()
                os.startfile(ruta_excel_absoluta)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                
        def btn_descargar_pdf():
            import shutil
            ruta_pdf = filedialog.asksaveasfilename(initialdir=dir_ficha, initialfile=f"Ficha_{eq_act['id']}.pdf", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
            if ruta_pdf:
                ruta_temp = os.path.abspath(os.path.join(dir_ficha, f"~preview_{id_sanitizado}.pdf"))
                if os.path.exists(ruta_temp):
                    try:
                        shutil.copy2(ruta_temp, ruta_pdf)
                        os.startfile(ruta_pdf)
                    except Exception as e:
                        messagebox.showerror("Error", f"No se pudo copiar el archivo temporal: {e}")
                else:
                    messagebox.showinfo("Espera", "Generando el documento PDF...")
                    try:
                        actualizar_excel_datos_silencioso()
                        excel = win32com.client.DispatchEx("Excel.Application")
                        excel.Visible = False
                        excel.DisplayAlerts = False
                        
                        wb_pdf = excel.Workbooks.Open(ruta_excel_absoluta, UpdateLinks=False, ReadOnly=True)
                        ws = wb_pdf.ActiveSheet
                        
                        try:
                            ws.PageSetup.Zoom = False
                            ws.PageSetup.FitToPagesWide = 1
                            ws.PageSetup.FitToPagesTall = 1
                            ws.PageSetup.PrintArea = "$A$1:$AQ$67"
                            ws.PageSetup.CenterHorizontally = True
                            ws.PageSetup.LeftMargin = 10
                            ws.PageSetup.RightMargin = 10
                            ws.PageSetup.TopMargin = 10
                            ws.PageSetup.BottomMargin = 10
                        except:
                            pass
                        
                        ws.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
                        wb_pdf.Close(False)
                        excel.Quit()
                        os.startfile(ruta_pdf)
                    except Exception as e:
                        messagebox.showerror("Error", f"Fallo al generar PDF: {e}")
                    finally:
                        pythoncom.CoUninitialize()

        ctk.CTkButton(c_img, text="📄 Abrir Excel", fg_color=C_GREEN, hover_color=C_GREEN_HOVER, command=btn_ver_excel).pack(pady=5, fill="x")
        ctk.CTkButton(c_img, text="⬇ Exportar PDF", fg_color=C_BLUE, hover_color=C_BLUE_HOVER, command=btn_descargar_pdf).pack(pady=5, fill="x")

        # Restaurando textos que faltaban de tu código original
        i_txt = ctk.CTkFrame(m_info, fg_color="transparent")
        i_txt.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        ctk.CTkLabel(i_txt, text=eq_act['nombre'], font=ctk.CTkFont(size=20, weight="bold"), text_color=C_BLUE, anchor="w").pack(fill="x")
        ctk.CTkLabel(i_txt, text=f"Marca/Modelo: {eq_act.get('marca','')} / {eq_act.get('modelo','')}", font=ctk.CTkFont(size=14, weight="bold"), anchor="w").pack(fill="x")
        ctk.CTkLabel(i_txt, text=f"Servicio: {eq_act.get('servicio','')} - {eq_act['area']}", font=ctk.CTkFont(size=14), text_color=C_SUBTEXT, anchor="w").pack(fill="x")
        
        ctk.CTkLabel(frame_izq, text="Historial de Intervenciones", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10, anchor="w", padx=20)
        
        tab_h = ttk.Treeview(frame_izq, columns=("Fecha","Tipo","Detalle"), show="headings", height=8)
        for c in ("Fecha","Tipo","Detalle"):
            tab_h.heading(c, text=c)
            tab_h.column(c, anchor="w" if c=="Detalle" else "center")
            
        tab_h.pack(pady=10, padx=20, fill="both", expand=True)
        
        for inter in eq_act.get("historial_intervenciones", []):
            tab_h.insert("", "end", values=(inter["fecha"], inter["tipo"], inter.get("trabajo", inter.get("detalle", ""))))

    def verificar_autorizacion_jefe(self, password_plano):
        from auth import verificar_password
        conn = obtener_conexion()
        if not conn: 
            return False
        try:
            cur = conn.cursor()
            cur.execute("SELECT password_hash FROM usuarios WHERE rol = 'jefe' AND activo = TRUE")
            rows = cur.fetchall()
            cur.close()
            conn.close()
            for r in rows:
                if verificar_password(password_plano, r[0]):
                    return True
        except:
            pass
        return False

    def verificar_y_ejecutar_backup_auto(self):
        import os
        import json
        import threading
        from datetime import datetime, timedelta
        from database import crear_backup_json
        
        def tarea_backup():
            try:
                base_dir = os.path.dirname(__file__)
                dir_respaldos = os.path.join(base_dir, "respaldos")
                if not os.path.exists(dir_respaldos):
                    os.makedirs(dir_respaldos, exist_ok=True)
                
                meta_path = os.path.join(dir_respaldos, "metadata.json")
                necesita_backup = False
                ultimo_fecha = None
                
                if os.path.exists(meta_path):
                    try:
                        with open(meta_path, "r", encoding="utf-8") as f:
                            meta = json.load(f)
                            fecha_str = meta.get("ultimo_respaldo_auto")
                            if fecha_str:
                                ultimo_fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    except Exception as e:
                        print("Error al leer metadata.json:", e)
                        necesita_backup = True
                else:
                    necesita_backup = True
                
                if not necesita_backup and ultimo_fecha:
                    if self.hoy - ultimo_fecha >= timedelta(days=7):
                        necesita_backup = True
                
                if necesita_backup:
                    hoy_str = self.hoy.strftime("%Y-%m-%d")
                    destino = os.path.join(dir_respaldos, f"respaldo_automatico_{hoy_str}.json")
                    if crear_backup_json(destino):
                        with open(meta_path, "w", encoding="utf-8") as f:
                            json.dump({"ultimo_respaldo_auto": hoy_str}, f, indent=2)
                        print("[OK] Respaldo automatico semanal creado correctamente.")
            except Exception as ex:
                print("Error en hilo de backup automático:", ex)
                
        threading.Thread(target=tarea_backup, daemon=True).start()

    def al_cerrar_aplicacion(self):
        import os
        from database import crear_backup_json
        try:
            base_dir = os.path.dirname(__file__)
            dir_respaldos = os.path.join(base_dir, "respaldos")
            if not os.path.exists(dir_respaldos):
                os.makedirs(dir_respaldos, exist_ok=True)
            
            hoy_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            destino = os.path.join(dir_respaldos, f"respaldo_cierre_{hoy_str}.json")
            if crear_backup_json(destino):
                print(f"[OK] Respaldo de cierre creado: {destino}")
        except Exception as e:
            print("Error al crear backup al cerrar:", e)
        finally:
            try:
                self.destroy()
            except:
                pass
            os._exit(0)

    def al_redimensionar(self, event):
        if event.widget != self:
            return
        
        h = self.winfo_height()
        nuevo_modo = "compact" if h < 780 else "spacious"
        
        if self.sidebar_mode != nuevo_modo:
            self.sidebar_mode = nuevo_modo
            self.aplicar_modo_sidebar(nuevo_modo)

    def aplicar_modo_sidebar(self, modo):
        if modo == "compact":
            self.lbl_logo.configure(font=ctk.CTkFont(size=16, weight="bold"))
            self.top_sidebar.pack_configure(pady=(10, 2))
            self.lbl_qr.pack_configure(pady=(0, 2))
            self.lbl_name.pack_configure(pady=(0, 6))
            self.btn_alertas.configure(height=34)
            self.btn_alertas.pack_configure(pady=(0, 8))
            self.btn_bottom_mantenimiento.configure(height=36, font=ctk.CTkFont(size=12, weight="bold"))
            self.bottom_sidebar.pack_configure(pady=8)
            
            for btn in self.botones_nav:
                btn.configure(height=34, font=ctk.CTkFont(size=12, weight="bold"))
                btn.pack_configure(pady=1)
        else:
            self.lbl_logo.configure(font=ctk.CTkFont(size=20, weight="bold"))
            self.top_sidebar.pack_configure(pady=(20, 5))
            self.lbl_qr.pack_configure(pady=(0, 2))
            self.lbl_name.pack_configure(pady=(0, 10))
            self.btn_alertas.configure(height=38)
            self.btn_alertas.pack_configure(pady=(0, 10))
            self.btn_bottom_mantenimiento.configure(height=42, font=ctk.CTkFont(size=13, weight="bold"))
            self.bottom_sidebar.pack_configure(pady=15)
            
            for btn in self.botones_nav:
                btn.configure(height=40, font=ctk.CTkFont(size=14, weight="bold"))
                btn.pack_configure(pady=2)

    # ========================================================
    # FORMULARIO: MANTENIMIENTO MAESTRO 
    # ========================================================
    def modulo_mantenimiento(self):
        v = ctk.CTkToplevel(self)
        v.title("Mantenimiento y Hoja de Trabajo")
        v.geometry("750x700")
        v.transient(self)
        v.grab_set()
        v.configure(fg_color=C_BG)
        
        ctk.CTkLabel(v, text="Registrar Intervención y Generar Hoja", font=ctk.CTkFont(size=20, weight="bold"), text_color=C_TEXT).pack(pady=15)
        
        sf = ctk.CTkScrollableFrame(v, fg_color=C_CARD, corner_radius=12)
        sf.pack(fill="both", expand=True, padx=20, pady=10)
        
        ctk.CTkLabel(sf, text="Seleccione el Equipo Médico:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10,0))
        
        noms = [f"{eq['id']} - {eq['nombre']}" for eq in self.datos["equipos"]]
        c_eq = ctk.CTkComboBox(sf, values=noms if noms else ["Vacío"], width=600)
        c_eq.pack(pady=5, padx=10)
        
        f_extra = ctk.CTkFrame(sf, fg_color="transparent")
        f_extra.pack(fill="x", pady=5)
        
        ctk.CTkLabel(f_extra, text="Servicio:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_servicio_excel = ctk.CTkEntry(f_extra, width=200)
        e_servicio_excel.pack(side="left", padx=5)
        
        ctk.CTkLabel(f_extra, text="Tipo:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_tipo_eq = ctk.CTkEntry(f_extra, width=150, placeholder_text="Ej. Móvil, Fijo")
        e_tipo_eq.pack(side="left", padx=5)

        f_fechas = ctk.CTkFrame(sf, fg_color="transparent")
        f_fechas.pack(fill="x", pady=5)
        
        ctk.CTkLabel(f_fechas, text="Fecha Recepción:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_fecha_rec = ctk.CTkEntry(f_fechas, width=120)
        e_fecha_rec.insert(0, datetime.now().strftime("%d / %m / %Y"))
        e_fecha_rec.pack(side="left", padx=5)
        ctk.CTkButton(f_fechas, text="Usar Actual", width=80, fg_color=C_BLUE, command=lambda: [e_fecha_rec.delete(0, "end"), e_fecha_rec.insert(0, datetime.now().strftime("%d / %m / %Y"))]).pack(side="left", padx=5)
        
        ctk.CTkLabel(f_fechas, text="Fecha Entrega:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_fecha_ent = ctk.CTkEntry(f_fechas, width=120)
        e_fecha_ent.insert(0, datetime.now().strftime("%d / %m / %Y"))
        e_fecha_ent.pack(side="left", padx=5)
        ctk.CTkButton(f_fechas, text="Usar Actual", width=80, fg_color=C_BLUE, command=lambda: [e_fecha_ent.delete(0, "end"), e_fecha_ent.insert(0, datetime.now().strftime("%d / %m / %Y"))]).pack(side="left", padx=5)

        f_tiempo_rep = ctk.CTkFrame(sf, fg_color="transparent")
        f_tiempo_rep.pack(fill="x", pady=5)
        ctk.CTkLabel(f_tiempo_rep, text="Tiempo de Reparación (horas):", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_tiempo_reparacion = ctk.CTkEntry(f_tiempo_rep, width=100)
        e_tiempo_reparacion.insert(0, "0")
        e_tiempo_reparacion.pack(side="left", padx=5)

        f_opc = ctk.CTkFrame(sf, fg_color="transparent")
        f_opc.pack(fill="x", pady=10)
        
        f_tipo = ctk.CTkFrame(f_opc, fg_color="transparent")
        f_tipo.pack(side="left", expand=True, fill="both", padx=10)
        ctk.CTkLabel(f_tipo, text="Tipo Mantenimiento:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        var_tipo = ctk.StringVar(value="")
        ctk.CTkRadioButton(f_tipo, text="Preventivo", variable=var_tipo, value="Preventivo").pack(anchor="w", pady=5)
        ctk.CTkRadioButton(f_tipo, text="Correctivo", variable=var_tipo, value="Correctivo").pack(anchor="w", pady=5)

        f_cond = ctk.CTkFrame(f_opc, fg_color="transparent")
        f_cond.pack(side="left", expand=True, fill="both", padx=10)
        ctk.CTkLabel(f_cond, text="Condición Encontrada:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        var_cond = ctk.StringVar(value="")
        for op in ["Óptimo", "Aceptable", "Crítica", "Inoperante", "F/Servicio"]:
            ctk.CTkRadioButton(f_cond, text=op, variable=var_cond, value=op).pack(anchor="w", pady=2)

        f_est = ctk.CTkFrame(f_opc, fg_color="transparent")
        f_est.pack(side="left", expand=True, fill="both", padx=10)
        ctk.CTkLabel(f_est, text="Estado del Equipo:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        var_est = ctk.StringVar(value="")
        for op in ["Óptimo", "Bueno", "Regular", "Malo", "Obsoleto"]:
            ctk.CTkRadioButton(f_est, text=op, variable=var_est, value=op).pack(anchor="w", pady=2)

        ctk.CTkLabel(sf, text="Deficiencia Encontrada:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(15,0))
        e_def = ctk.CTkTextbox(sf, width=600, height=60, fg_color=C_BG, corner_radius=8)
        e_def.pack(padx=10, pady=5)
        
        ctk.CTkLabel(sf, text="D. Trabajo Realizado:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10,0))
        e_trab = ctk.CTkTextbox(sf, width=600, height=80, fg_color=C_BG, corner_radius=8)
        e_trab.pack(padx=10, pady=5)
        
        ctk.CTkLabel(sf, text="E. Observaciones y Recomendaciones:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10,0))
        e_obs = ctk.CTkTextbox(sf, width=600, height=80, fg_color=C_BG, corner_radius=8)
        e_obs.pack(padx=10, pady=5)

        f_acciones = ctk.CTkFrame(sf, fg_color="transparent")
        ruta_ht_excel_act = ctk.StringVar()

        def guardar():
            sel = c_eq.get()
            if not sel or "Vacío" in sel:
                messagebox.showwarning("Aviso", "Seleccione un equipo.")
                return
            if not var_tipo.get() or not var_cond.get() or not var_est.get():
                messagebox.showwarning("Dato Obligatorio", "Debe marcar todos los botones de opciones.")
                return
            
            id_e = sel.split(" - ")[0]
            eq_data = next((eq for eq in self.datos["equipos"] if eq["id"] == id_e), None)
            
            if not eq_data: 
                return

            # Validar si el mantenimiento es preventivo y está vencido (más del 5 del mes siguiente al programado)
            if var_tipo.get() == "Preventivo":
                crit_eq = eq_data.get("criticidad", "Riesgo Medio")
                meses_eq = 3 if "Alto" in crit_eq else (4 if "Medio" in crit_eq else 6)
                f_reg_eq = eq_data.get("fecha_adquisicion") or eq_data.get("fecha_registro", self.hoy)
                if isinstance(f_reg_eq, datetime):
                    f_reg_eq = f_reg_eq.date()
                elif isinstance(f_reg_eq, str):
                    try:
                        f_reg_eq = datetime.strptime(f_reg_eq, "%Y-%m-%d").date()
                    except:
                        f_reg_eq = eq_data.get("fecha_registro", self.hoy)
                        if isinstance(f_reg_eq, str):
                            try:
                                f_reg_eq = datetime.strptime(f_reg_eq, "%Y-%m-%d").date()
                            except:
                                f_reg_eq = self.hoy
                        elif isinstance(f_reg_eq, datetime):
                            f_reg_eq = f_reg_eq.date()
                
                if eq_data.get("garantia") == "Con Garantía" and eq_data.get("fecha_vencimiento_garantia"):
                    f_venc_g = eq_data.get("fecha_vencimiento_garantia")
                    if isinstance(f_venc_g, str):
                        try: f_venc_g = datetime.strptime(f_venc_g, "%Y-%m-%d").date()
                        except: f_venc_g = None
                    elif isinstance(f_venc_g, datetime):
                        f_venc_g = f_venc_g.date()
                    if f_venc_g:
                        f_reg_eq = f_venc_g + relativedelta(days=+1)

                # Buscar el primer slot no completado como f_prox_eq
                f_prox_eq = None
                f_check_eq = f_reg_eq
                iter_count_eq = 0
                while iter_count_eq < 50:
                    iter_count_eq += 1
                    f_check_eq = f_check_eq + relativedelta(months=+meses_eq)
                    
                    slot_is_completed = False
                    for m in eq_data.get("historial_intervenciones", []):
                        if m["tipo"] == "Preventivo":
                            m_prog = m.get("fecha_programada")
                            if m_prog:
                                if isinstance(m_prog, str):
                                    try: m_prog_d = datetime.strptime(m_prog, "%Y-%m-%d").date()
                                    except: m_prog_d = None
                                else:
                                    m_prog_d = m_prog
                                if m_prog_d == f_check_eq:
                                    slot_is_completed = True
                                    break
                            else:
                                m_f = m.get("fecha")
                                if isinstance(m_f, str):
                                    try: m_f_d = datetime.strptime(m_f, "%Y-%m-%d").date()
                                    except: m_f_d = None
                                else:
                                    m_f_d = m_f
                                if m_f_d and m_f_d.year == f_check_eq.year and m_f_d.month == f_check_eq.month:
                                    slot_is_completed = True
                                    break
                    if not slot_is_completed:
                        f_prox_eq = f_check_eq
                        break
                
                if not f_prox_eq:
                    f_prox_eq = f_reg_eq + relativedelta(months=+meses_eq)
                
                limit_date_eq = date(f_prox_eq.year, f_prox_eq.month, 1) + relativedelta(months=+1, day=5)
                if self.hoy > limit_date_eq:
                    # Mantenimiento vencido detectado
                    if not self.es_jefe:
                        from customtkinter import CTkInputDialog
                        dialog = CTkInputDialog(text=f"El mantenimiento preventivo de este equipo venció el {limit_date_eq.strftime('%d/%m/%Y')}.\nSe requiere autorización del Jefe para registrarlo.\nIngrese la contraseña del Jefe/Administrador:", title="Autorización Requerida")
                        pwd_val = dialog.get_input()
                        if not pwd_val:
                            messagebox.showwarning("Acceso Denegado", "Operación cancelada. El registro de mantenimientos vencidos requiere autorización.")
                            return
                        
                        if not self.verificar_autorizacion_jefe(pwd_val):
                            messagebox.showerror("Error de Autorización", "Contraseña del Jefe incorrecta. No se autorizó el registro.")
                            return
                        else:
                            messagebox.showinfo("Autorizado", "Registro de mantenimiento vencido autorizado correctamente por el Jefe.")

            txt_def = e_def.get("1.0", "end-1c").strip()
            txt_trab = e_trab.get("1.0", "end-1c").strip()
            txt_obs = e_obs.get("1.0", "end-1c").strip()
            
            try: 
                f_entrega_iso = datetime.strptime(e_fecha_ent.get().strip(), "%d / %m / %Y").strftime("%Y-%m-%d")
            except: 
                f_entrega_iso = datetime.now().strftime("%Y-%m-%d")

            try:
                conn = obtener_conexion()
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO historial_intervenciones 
                    (equipo_id, fecha, tipo, detalle, condicion, estado_equipo, deficiencia, trabajo, observaciones, fecha_entrega, servicio_ht, tipo_ht) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (id_e, datetime.now().strftime("%Y-%m-%d"), var_tipo.get(), txt_trab, var_cond.get(), var_est.get(), txt_def, txt_trab, txt_obs, f_entrega_iso, e_servicio_excel.get(), e_tipo_eq.get()))
                conn.commit()
                cur.close()
                conn.close()
                
                self.cargar_datos_memoria()
                self.vistas["Historial"].refrescar_datos()
                self._calendario_sucio = True
                
                messagebox.showinfo("Éxito", "Intervención guardada en PostgreSQL.")
                btn_guardar.configure(state="disabled")
                generar_excel_ht(eq_data)
                
            except Exception as e:
                messagebox.showerror("Error SQL", str(e))

        def generar_excel_ht(eq_data):
            ruta_plantilla_ht = obtener_ruta_plantilla("plantilla_trabajo.xlsx")
            
            if not os.path.exists(ruta_plantilla_ht):
                messagebox.showerror("Error", f"No se encontró la plantilla en:\n{ruta_plantilla_ht}")
                return
                
            try:
                wb = openpyxl.load_workbook(ruta_plantilla_ht)
                ws = wb.active
                
                def w(celda, valor, is_check=False):
                    try:
                        ws[celda] = valor
                        if is_check and valor == 'X':
                            ws[celda].font = Font(name='Calibri', bold=True, color='000000', size=14)
                            ws[celda].alignment = Alignment(horizontal='center', vertical='center')
                    except: 
                        pass
                        
                def w_texto(celda, valor):
                    try:
                        ws[celda] = valor
                        ws[celda].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    except: 
                        pass
                        
                w('F11', eq_data.get('servicio', ''))
                w('AA11', e_servicio_excel.get())
                w('S21', e_tipo_eq.get())
                w('J15', eq_data.get('nombre', ''))
                w('AE15', eq_data.get('id', ''))
                w('E17', eq_data.get('procedencia', ''))
                w('AB17', eq_data.get('anio_fab', ''))
                w('E19', eq_data.get('marca', ''))
                w('AB19', eq_data.get('fabricante', ''))
                w('F21', eq_data.get('modelo', ''))
                w('AG21', eq_data.get('id', ''))
                w('M23', e_fecha_rec.get())
                w('I62', f"{e_fecha_ent.get()}  {datetime.now().strftime('%H:%M')}")
                
                # CORRECCIÓN DE "CONDICIÓN ENCONTRADA" VACÍA (Restaurado desde tu código original)
                if var_cond.get() == "Óptimo":
                    w('P26', 'X', True)
                elif var_cond.get() == "Aceptable":
                    w('W26', 'X', True)
                elif var_cond.get() == "Crítica":
                    w('AC26', 'X', True)
                elif var_cond.get() == "Inoperante":
                    w('AJ26', 'X', True)
                elif var_cond.get() == "F/Servicio":
                    w('AP26', 'X', True)

                # Restauración de Estados del Equipo
                if var_est.get() == "Óptimo":
                    w('O29', 'X', True)
                elif var_est.get() == "Bueno":
                    w('U29', 'X', True)
                elif var_est.get() == "Regular":
                    w('AB29', 'X', True)
                elif var_est.get() == "Malo":
                    w('AH29', 'X', True)
                elif var_est.get() == "Obsoleto":
                    w('AO29', 'X', True)

                # Restauración Tipo Mantenimiento
                if var_tipo.get() == "Preventivo":
                    w('Q43', 'X', True)
                else:
                    w('AL43', 'X', True)

                w_texto('B33', e_def.get("1.0", "end-1c").strip())
                w_texto('B47', e_trab.get("1.0", "end-1c").strip())
                w_texto('B53', e_obs.get("1.0", "end-1c").strip())
                
                timestamp_seguro = datetime.now().strftime('%H%M%S')
                nombre_salida = f"HT_{eq_data['id']}_{datetime.now().strftime('%Y%m%d')}_{timestamp_seguro}.xlsx"
                area_name = eq_data.get("area", "General")
                area_folder = "".join([c for c in area_name if c.isalnum() or c==' ']).strip()
                dir_mantenimiento = os.path.join(CARPETAS["areas"], area_folder, "mantenimientos")
                os.makedirs(dir_mantenimiento, exist_ok=True)
                ruta_salida = os.path.join(dir_mantenimiento, nombre_salida)
                
                wb.save(ruta_salida)
                ruta_ht_excel_act.set(os.path.abspath(ruta_salida))
                f_acciones.pack(fill="x", pady=10, padx=20)
                
            except Exception as e:
                messagebox.showerror("Error Excel", f"Fallo al generar hoja:\n{e}")

        def abrir_excel_ht():
            try: 
                os.startfile(ruta_ht_excel_act.get())
            except Exception as e: 
                messagebox.showerror("Error", str(e))

        def exportar_pdf_ht():
            ruta_pdf = filedialog.asksaveasfilename(initialdir=dir_mantenimiento, initialfile=f"Hoja_Trabajo.pdf", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
            if not ruta_pdf: 
                return
            
            messagebox.showinfo("Exportando", "Por favor espera, generando PDF...")
            # Eliminado PrintArea para que se respete el de la plantilla original (no más recortes)
            if exportar_excel_a_pdf(ruta_ht_excel_act.get(), ruta_pdf):
                os.startfile(os.path.abspath(ruta_pdf))
            else:
                messagebox.showerror("Error", "Fallo al generar el PDF.")

        btn_guardar = ctk.CTkButton(sf, text="Guardar Registro y Generar Hoja", height=45, fg_color=C_BLUE, hover_color=C_BLUE_HOVER, font=ctk.CTkFont(weight="bold"), command=guardar)
        btn_guardar.pack(pady=20, padx=20, fill="x")
        
        ctk.CTkButton(f_acciones, text="📄 Abrir Hoja de Trabajo (Excel)", font=ctk.CTkFont(weight="bold"), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, command=abrir_excel_ht).pack(side="left", expand=True, padx=5)
        ctk.CTkButton(f_acciones, text="⬇ Exportar a PDF", font=ctk.CTkFont(weight="bold"), fg_color=C_PURPLE, hover_color="#963ECA", command=exportar_pdf_ht).pack(side="left", expand=True, padx=5)

# ========================================================
# ARRANQUE OFICIAL DE LA APLICACIÓN
# ========================================================
if __name__ == "__main__":
    inicializar_bd()
    inicializar_usuarios()
    
    login_win = VentanaLogin()
    login_win.mainloop()
    
    if login_win.usuario_autenticado:
        app = SistemaMantenimiento(usuario=login_win.usuario_autenticado)
        app.mainloop()