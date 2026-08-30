# vistas/cronograma.py
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import calendar
from datetime import date, datetime
from estilos import *

class VistaCronograma(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=C_BG)
        self.app = app
        self.hoy = self.app.hoy
        self.mes_actual = self.hoy.month
        self.anio_actual = self.hoy.year
        self.anio_vista = self.hoy.year
        self.construir_ui()

    def construir_ui(self):
        # Cabecera con Botón de Descarga
        f_cab = ctk.CTkFrame(self, fg_color="transparent")
        f_cab.pack(pady=(30, 10), padx=30, fill="x")
        ctk.CTkLabel(f_cab, text="Cronograma de Mantenimiento", font=ctk.CTkFont(size=28, weight="bold"), text_color=C_TEXT).pack(side="left")
        
        self.btn_descargar = ctk.CTkButton(f_cab, text="⬇ Descargar Cronograma Anual", font=ctk.CTkFont(weight="bold"), 
                                            fg_color=C_GREEN, hover_color=C_GREEN_HOVER, corner_radius=8, 
                                            command=self.abrir_selector_gestion)
        self.btn_descargar.pack(side="right")
        
        # Tabview Estilo Moderno
        self.tabview = ctk.CTkTabview(self, fg_color=C_CARD, border_width=1, border_color=C_BORDER, text_color=C_TEXT, corner_radius=16, 
                                      segmented_button_selected_color=C_BLUE, segmented_button_selected_hover_color=C_BLUE_HOVER,
                                      segmented_button_unselected_color=C_BG,
                                      segmented_button_unselected_hover_color=C_CARD_HOVER)
        self.tabview.pack(pady=10, padx=30, fill="both", expand=True)
        
        tab_lista = self.tabview.add("📋 Resumen")
        tab_mes = self.tabview.add("📅 Mensual")
        tab_anio = self.tabview.add("🗓️ Mapa Anual")

        # ------------------- TAB 1: RESUMEN (LISTA) -------------------
        f_filtros_crono = ctk.CTkFrame(tab_lista, fg_color="transparent")
        f_filtros_crono.pack(fill="x", pady=10, padx=15)
        
        self.busqueda_var = ctk.StringVar()
        self.busqueda_var.trace_add("write", lambda *args: self.refrescar_datos())
        ctk.CTkLabel(f_filtros_crono, text="🔍 Buscar:", font=ctk.CTkFont(weight="bold"), text_color=C_TEXT).pack(side="left", padx=5)
        e_buscar = ctk.CTkEntry(f_filtros_crono, textvariable=self.busqueda_var, placeholder_text="Buscar ID o Equipo...", width=250, fg_color=C_BG, border_color=C_BORDER, corner_radius=10)
        e_buscar.pack(side="left", padx=5)
        
        ctk.CTkLabel(f_filtros_crono, text="Ordenar por:", font=ctk.CTkFont(weight="bold", size=12), text_color=C_TEXT).pack(side="left", padx=(15, 5))
        self.combo_ordenar = ctk.CTkComboBox(f_filtros_crono, values=["ID", "Equipo (A-Z)", "Equipo (Z-A)", "Criticidad", "Fecha Próx", "Estado"], command=lambda e: self.refrescar_datos(), width=160, fg_color=C_BG, border_color=C_BORDER)
        self.combo_ordenar.pack(side="left", padx=5)
        self.combo_ordenar.set("ID")
        
        f_tree_crono = ctk.CTkFrame(tab_lista, fg_color="transparent")
        f_tree_crono.pack(pady=(5, 15), padx=15, fill="both", expand=True)
        self.t_cro_lista = ttk.Treeview(f_tree_crono, columns=("ID", "Equipo", "Criticidad", "Prox", "Estado"), show="headings")
        scrollbar_crono = ttk.Scrollbar(f_tree_crono, orient="vertical", command=self.t_cro_lista.yview, style="Vertical.TScrollbar")
        self.t_cro_lista.configure(yscrollcommand=scrollbar_crono.set)
        for c in ("ID", "Equipo", "Criticidad", "Prox", "Estado"):
            self.t_cro_lista.heading(c, text=c)
            self.t_cro_lista.column(c, anchor="center")
        self.t_cro_lista.pack(side="left", fill="both", expand=True)
        scrollbar_crono.pack(side="right", fill="y", padx=(5, 0))
        
        # Colores semánticos suaves para las filas
        self.t_cro_lista.tag_configure("Vencido", background="#FEE2E2", foreground="#B91C1C")
        self.t_cro_lista.tag_configure("Por Vencer", background="#FEF3C7", foreground="#B45309")
        self.t_cro_lista.tag_configure("Al Día", background="#D1FAE5", foreground="#047857")
        self.t_cro_lista.tag_configure("Dado de Baja", background="#F1F5F9", foreground="#64748B")


        def abrir_ficha_desde_crono(event):
            sel = self.t_cro_lista.focus()
            if sel:
                vals = self.t_cro_lista.item(sel, "values")
                if vals and len(vals) > 0:
                    eq_id = vals[0]
                    self.app.abrir_hoja_vida_click(equipo_id=eq_id)

        self.t_cro_lista.bind("<Double-1>", abrir_ficha_desde_crono)

        # ------------------- TAB 2: MENSUAL -------------------
        f_mes_izq = ctk.CTkFrame(tab_mes, fg_color="transparent")
        f_mes_izq.pack(side="left", fill="both", expand=True, padx=10)
        
        f_mes_der = ctk.CTkFrame(tab_mes, width=320, fg_color=C_BG, corner_radius=12)
        f_mes_der.pack(side="right", fill="y", padx=10, pady=10)
        f_mes_der.pack_propagate(False)
        
        ctk.CTkLabel(f_mes_der, text="Agenda Diaria", font=ctk.CTkFont(size=20, weight="bold"), text_color=C_TEXT).pack(pady=(20, 5))
        self.lbl_det_fecha = ctk.CTkLabel(f_mes_der, text="Toca un día con mantenimientos", font=ctk.CTkFont(size=13), text_color=C_SUBTEXT)
        self.lbl_det_fecha.pack()
        
        self.txt_det = ctk.CTkTextbox(f_mes_der, fg_color="transparent", text_color=C_TEXT, font=ctk.CTkFont(size=14))
        self.txt_det.pack(fill="both", expand=True, padx=15, pady=15)
        
        h_mes = ctk.CTkFrame(f_mes_izq, fg_color="transparent")
        h_mes.pack(fill="x", pady=10)
        self.lbl_mes_tit = ctk.CTkLabel(h_mes, text="", font=ctk.CTkFont(size=24, weight="bold"), text_color=C_TEXT)
        self.lbl_mes_tit.pack(side="left", padx=10)

        ctk.CTkButton(h_mes, text="Siguiente >", width=100, fg_color=C_BG, text_color=C_BLUE, hover_color=C_BORDER, corner_radius=8, command=lambda: self.nav_mes(1)).pack(side="right", padx=5)
        ctk.CTkButton(h_mes, text="< Anterior", width=100, fg_color=C_BG, text_color=C_BLUE, hover_color=C_BORDER, corner_radius=8, command=lambda: self.nav_mes(-1)).pack(side="right", padx=5)
        
        self.grid_mes = ctk.CTkFrame(f_mes_izq, fg_color="transparent")
        self.grid_mes.pack(fill="both", expand=True)
        for i, d in enumerate(["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]):
            self.grid_mes.grid_columnconfigure(i, weight=1)
            ctk.CTkLabel(self.grid_mes, text=d, font=ctk.CTkFont(weight="bold"), text_color=C_SUBTEXT).grid(row=0, column=i, pady=5)
            
        self.botones_mes = []
        for f in range(6):
            fila_b = []
            for c in range(7):
                b = ctk.CTkButton(self.grid_mes, text="", width=45, height=45, corner_radius=22, font=ctk.CTkFont(size=14, weight="bold"))
                b.grid(row=f + 1, column=c, padx=4, pady=4)
                fila_b.append(b)
            self.botones_mes.append(fila_b)

        # ------------------- TAB 3: MAPA ANUAL -------------------
        h_anio = ctk.CTkFrame(tab_anio, fg_color="transparent")
        h_anio.pack(fill="x", pady=10)
        self.lbl_anio_tit = ctk.CTkLabel(h_anio, text="", font=ctk.CTkFont(size=26, weight="bold"), text_color=C_TEXT)
        self.lbl_anio_tit.pack(side="left", padx=10)

        ctk.CTkButton(h_anio, text="Año Siguiente >", width=120, fg_color=C_BG, text_color=C_BLUE, hover_color=C_BORDER, corner_radius=8, command=lambda: self.nav_anio(1)).pack(side="right", padx=5)
        ctk.CTkButton(h_anio, text="< Año Anterior", width=120, fg_color=C_BG, text_color=C_BLUE, hover_color=C_BORDER, corner_radius=8, command=lambda: self.nav_anio(-1)).pack(side="right", padx=5)
        
        grid_anio = ctk.CTkFrame(tab_anio, fg_color="transparent")
        grid_anio.pack(fill="both", expand=True)
        for i in range(4): grid_anio.grid_columnconfigure(i, weight=1)
        for i in range(3): grid_anio.grid_rowconfigure(i, weight=1)
        
        self.meses_labels_anio = []
        self.botones_anio = {m: [] for m in range(1, 13)}
        
        for mes in range(1, 13):
            marco = ctk.CTkFrame(grid_anio, fg_color=C_BG, corner_radius=10)
            marco.grid(row=(mes - 1) // 4, column=(mes - 1) % 4, padx=6, pady=6, sticky="nsew")
            lbl = ctk.CTkLabel(marco, text="", font=ctk.CTkFont(weight="bold", size=13), text_color=C_TEXT)
            lbl.pack(pady=(5, 0))
            self.meses_labels_anio.append(lbl)
            
            g_dias = ctk.CTkFrame(marco, fg_color="transparent")
            g_dias.pack(expand=True, padx=4, pady=4)
            for i in range(7): g_dias.grid_columnconfigure(i, weight=1)
            
            for f in range(6):
                fila_b = []
                for c in range(7):
                    # Usar tk.Label nativo para obtener un renderizado 100x más rápido
                    b = tk.Label(g_dias, text="", width=3, height=1, font=("Segoe UI", 11, "bold"), bg=C_BG, fg=C_TEXT, relief="flat")
                    b.grid(row=f, column=c, padx=1, pady=1)
                    fila_b.append(b)
                self.botones_anio[mes].append(fila_b)

    def obtener_eventos_sede(self, f_dia):
        """Retorna los eventos del calendario que pertenecen a la sede activa."""
        evs = self.app.eventos_calendario.get(f_dia, [])
        if not hasattr(self.app, "equipo_cumple_sede_activa"):
            return evs
        return [e for e in evs if self.app.equipo_cumple_sede_activa(e)]

    def refrescar_datos(self):
        """Actualiza la lista rápida showing ONLY the next scheduled maintenance for each equipment."""
        for i in self.t_cro_lista.get_children():
            self.t_cro_lista.delete(i)
            
        from dateutil.relativedelta import relativedelta
        from datetime import datetime, date
        
        items = []
        for eq in self.app.datos.get("equipos", []):
            if hasattr(self.app, "equipo_cumple_sede_activa") and not self.app.equipo_cumple_sede_activa(eq):
                continue
            crit = str(eq.get("criticidad") or "Riesgo Medio")
            meses = 3 if "Alto" in crit else (4 if "Medio" in crit else 6)
            
            f_reg = eq.get("fecha_adquisicion") or eq.get("fecha_registro", self.app.hoy)
            if isinstance(f_reg, datetime):
                f_reg = f_reg.date()
            elif isinstance(f_reg, str):
                try:
                    f_reg = datetime.strptime(f_reg, "%Y-%m-%d").date()
                except:
                    # Fallback a fecha_registro
                    f_reg = eq.get("fecha_registro", self.app.hoy)
                    if isinstance(f_reg, str):
                        try:
                            f_reg = datetime.strptime(f_reg, "%Y-%m-%d").date()
                        except:
                            f_reg = self.app.hoy
                    elif isinstance(f_reg, datetime):
                        f_reg = f_reg.date()
                    
            if eq.get("garantia") == "Con Garantía" and eq.get("fecha_vencimiento_garantia"):
                f_venc_g = eq.get("fecha_vencimiento_garantia")
                if isinstance(f_venc_g, str):
                    try:
                        f_venc_g = datetime.strptime(f_venc_g, "%Y-%m-%d").date()
                    except:
                        f_venc_g = None
                elif isinstance(f_venc_g, datetime):
                    f_venc_g = f_venc_g.date()
                if f_venc_g:
                    f_reg = f_venc_g + relativedelta(days=+1)

            # Buscar el primer slot no completado como f_prox
            f_prox = None
            estado = "Al Día"
            f_check = f_reg
            iter_count = 0
            while iter_count < 50:
                iter_count += 1
                f_check = f_check + relativedelta(months=+meses)
                
                slot_is_completed = False
                for m in eq.get("historial_intervenciones", []):
                    if m["tipo"] == "Preventivo":
                        m_prog = m.get("fecha_programada")
                        if m_prog:
                            if isinstance(m_prog, str):
                                try: m_prog_d = datetime.strptime(m_prog, "%Y-%m-%d").date()
                                except: m_prog_d = None
                            else:
                                m_prog_d = m_prog
                            if m_prog_d == f_check:
                                slot_is_completed = True
                                break
                        else:
                            m_f = m.get("fecha")
                            if isinstance(m_f, str):
                                try: m_f_d = datetime.strptime(m_f, "%Y-%m-%d").date()
                                except: m_f_d = None
                            else:
                                m_f_d = m_f
                            if m_f_d and m_f_d.year == f_check.year and m_f_d.month == f_check.month:
                                slot_is_completed = True
                                break
                if not slot_is_completed:
                    f_prox = f_check
                    dias_restantes = (f_prox - self.app.hoy).days
                    limit_date = date(f_prox.year, f_prox.month, 1) + relativedelta(months=+1, day=5)
                    if eq.get("estado") == "Baja":
                        estado = "Dado de Baja"
                    else:
                        estado = "Vencido" if self.app.hoy > limit_date else ("Por Vencer" if dias_restantes <= 30 else "Al Día")
                    break
            
            if not f_prox:
                f_prox = f_reg + relativedelta(months=+meses)
                dias_restantes = (f_prox - self.app.hoy).days
                limit_date = date(f_prox.year, f_prox.month, 1) + relativedelta(months=+1, day=5)
                if eq.get("estado") == "Baja":
                    estado = "Dado de Baja"
                else:
                    estado = "Vencido" if self.app.hoy > limit_date else ("Por Vencer" if dias_restantes <= 30 else "Al Día")
            
            eq_label = f"{eq['nombre']} ({eq.get('marca', '')} {eq.get('modelo', '')})"
            
            items.append({
                "id": eq["id"],
                "eq_label": eq_label,
                "crit": crit,
                "f_prox": f_prox,
                "estado": estado
            })
            
        # Filtro por búsqueda
        t = self.busqueda_var.get().lower().strip() if hasattr(self, "busqueda_var") else ""
        if t:
            items = [it for it in items if (
                t in str(it["id"]).lower() or
                t in str(it["eq_label"]).lower() or
                t in str(it["crit"]).lower() or
                t in str(it["estado"]).lower()
            )]
            
        # Ordenación
        criterio = self.combo_ordenar.get() if hasattr(self, "combo_ordenar") else "ID"
        if criterio == "Equipo (A-Z)":
            items.sort(key=lambda x: str(x["eq_label"]).lower())
        elif criterio == "Equipo (Z-A)":
            items.sort(key=lambda x: str(x["eq_label"]).lower(), reverse=True)
        elif criterio == "Criticidad":
            prioridad = {"Riesgo Alto": 1, "Riesgo Medio": 2, "Riesgo Bajo": 3}
            items.sort(key=lambda x: prioridad.get(x["crit"], 4))
        elif criterio == "Fecha Próx":
            items.sort(key=lambda x: x["f_prox"])
        elif criterio == "Estado":
            prioridad_est = {"Vencido": 1, "Por Vencer": 2, "Al Día": 3}
            items.sort(key=lambda x: prioridad_est.get(x["estado"], 4))
        else: # ID
            items.sort(key=lambda x: str(x["id"]).lower())
            
        for it in items:
            f_prox_str = it['f_prox'].strftime("%Y-%m-%d") if it['estado'] != "Dado de Baja" else "N/A"
            self.t_cro_lista.insert("", "end", values=(it['id'], it['eq_label'], it['crit'], f_prox_str, it['estado']), tags=(it['estado'],))

    def dibujar_mes(self, y, m):
        nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        self.lbl_mes_tit.configure(text=f"{nombres[m - 1]} {y}")
        cal = calendar.monthcalendar(y, m)
        
        for f in range(6):
            for c in range(7):
                btn = self.botones_mes[f][c]
                if f < len(cal) and cal[f][c] != 0:
                    dia = cal[f][c]
                    f_iter = date(y, m, dia)
                    bg, text_c = "transparent", C_TEXT
                    
                    evs_sede = self.obtener_eventos_sede(f_iter)
                    if evs_sede:
                        est = [e['estado'] for e in evs_sede]
                        if "Vencido" in est: bg, text_c = C_RED, "white"
                        elif "Pendiente Este Mes" in est: bg, text_c = C_YELLOW, "black"
                        elif "Realizado a Tiempo" in est: bg, text_c = C_GREEN, "white"
                        elif "Realizado Fuera de Fecha" in est: bg, text_c = C_PURPLE, "white"
                        elif "Futuro" in est: bg, text_c = C_BLUE, "white"
                        
                    btn.configure(text=str(dia), fg_color=bg, text_color=text_c, state="normal", command=lambda fdt=f_iter: self.click_dia(fdt))
                else:
                    btn.configure(text="", fg_color="transparent", state="disabled", command=lambda: None)

    def dibujar_anio(self, y):
        self.lbl_anio_tit.configure(text=str(y))
        nombres = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        
        for mes in range(1, 13):
            self.meses_labels_anio[mes - 1].configure(text=nombres[mes - 1])
            cal = calendar.monthcalendar(y, mes)
            for f in range(6):
                for c in range(7):
                    btn = self.botones_anio[mes][f][c]
                    if f < len(cal) and cal[f][c] != 0:
                        dia = cal[f][c]
                        f_iter = date(y, mes, dia)
                        bg, text_c = C_BG, C_SUBTEXT
                        
                        if f_iter == self.hoy: text_c = C_BLUE
                        evs_sede = self.obtener_eventos_sede(f_iter)
                        if evs_sede:
                            est = [e['estado'] for e in evs_sede]
                            if "Vencido" in est: bg, text_c = C_RED, "white"
                            elif "Pendiente Este Mes" in est: bg, text_c = C_YELLOW, "black"
                            elif "Realizado a Tiempo" in est: bg, text_c = C_GREEN, "white"
                            elif "Realizado Fuera de Fecha" in est: bg, text_c = C_PURPLE, "white"
                            elif "Futuro" in est: bg, text_c = C_BLUE, "white"
                            
                        btn.configure(text=str(dia), bg=bg, fg=text_c, cursor="hand2")
                        btn.bind("<Button-1>", lambda event, fdt=f_iter: self.click_dia_anio(fdt))
                    else:
                        btn.configure(text="", bg=C_BG, fg=C_SUBTEXT, cursor="")
                        btn.unbind("<Button-1>")

    def click_dia(self, f_click):
        self.lbl_det_fecha.configure(text=f_click.strftime("%d %B, %Y"))
        self.txt_det.configure(state="normal")
        self.txt_det.delete("1.0", "end")
        evs_sede = self.obtener_eventos_sede(f_click)
        if evs_sede:
            for ev in evs_sede:
                if ev['estado'] == "Vencido": sim = "🔴"
                elif ev['estado'] == "Pendiente Este Mes": sim = "🟡"
                elif ev['estado'] == "Realizado a Tiempo": sim = "🟢"
                elif ev['estado'] == "Realizado Fuera de Fecha": sim = "🟣"
                else: sim = "🔵"
                self.txt_det.insert("end", f"{sim} ID: {ev['id']} ({ev['estado']})\n{ev['eq']}\n\n")
        else:
            self.txt_det.insert("end", "Día libre de mantenimientos.")
        self.txt_det.configure(state="disabled")

    def click_dia_anio(self, f_click):
        pop = ctk.CTkToplevel(self.app)
        pop.title("Detalles del Día")
        pop.geometry("350x450")
        pop.transient(self.app)
        pop.grab_set()
        pop.configure(fg_color=C_BG)
        
        ctk.CTkLabel(pop, text=f_click.strftime("%d de %B, %Y"), font=ctk.CTkFont(size=18, weight="bold"), text_color=C_BLUE).pack(pady=(20, 10))
        txt = ctk.CTkTextbox(pop, fg_color=C_CARD, corner_radius=10, text_color=C_TEXT, font=ctk.CTkFont(size=14))
        txt.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        evs_sede = self.obtener_eventos_sede(f_click)
        if evs_sede:
            for ev in evs_sede:
                sim = "🔴" if ev['estado'] == "Vencido" else ("🟡" if ev['estado'] == "Por Vencer" else "🟢")
                txt.insert("end", f"{sim} ID: {ev['id']}\n{ev['eq']}\n\n")
        else:
            txt.insert("end", "\nDía libre.")
        txt.configure(state="disabled")

    def nav_mes(self, d):
        self.mes_actual += d
        if self.mes_actual > 12:
            self.mes_actual = 1
            self.anio_actual += 1
        if self.mes_actual < 1:
            self.mes_actual = 12
            self.anio_actual -= 1
        self.dibujar_mes(self.anio_actual, self.mes_actual)

    def nav_anio(self, d):
        self.anio_vista += d
        self.dibujar_anio(self.anio_vista)

    def abrir_selector_gestion(self):
        vent = ctk.CTkToplevel(self)
        vent.title("Selección de Gestión")
        vent.geometry("380x250")
        vent.transient(self.app)
        vent.grab_set()
        vent.configure(fg_color=C_CARD)
        
        ctk.CTkLabel(vent, text="Descarga de Cronograma Anual", font=ctk.CTkFont(size=16, weight="bold"), text_color=C_TEXT).pack(pady=15)
        ctk.CTkLabel(vent, text="Seleccione la gestión de inicio (Año):", font=ctk.CTkFont(size=13)).pack(anchor="w", padx=40, pady=(10, 0))
        
        anio_h = datetime.now().year
        opciones_anios = [str(anio_h - 1), str(anio_h), str(anio_h + 1), str(anio_h + 2)]
        combo_anio = ctk.CTkComboBox(vent, values=opciones_anios, width=300)
        combo_anio.set(str(anio_h))
        combo_anio.pack(pady=10)
        
        def confirmar():
            anio_sel = int(combo_anio.get())
            vent.destroy()
            self.ejecutar_descarga_excel(anio_sel)
            
        ctk.CTkButton(vent, text="Exportar a Excel", fg_color=C_GREEN, hover_color=C_GREEN_HOVER, font=ctk.CTkFont(weight="bold"), command=confirmar).pack(pady=20)

    def ejecutar_descarga_excel(self, anio_inicio):
        from excel_utils import obtener_ruta_plantilla
        from dateutil.relativedelta import relativedelta
        import openpyxl
        from copy import copy
        
        sede_activa = getattr(self.app, "contexto_sede", {}) or {}
        cen_nom = sede_activa.get("centro_salud")
        red_nom = sede_activa.get("red_salud")
        
        if cen_nom and not str(cen_nom).startswith("[ Todos"):
            sufijo_sede = str(cen_nom).replace(" ", "_").replace("/", "_")
        elif red_nom and not str(red_nom).startswith("[ Todas"):
            sufijo_sede = str(red_nom).split("(")[0].strip().replace(" ", "_")
        else:
            sufijo_sede = "GAMLP_General"
            
        ruta_guardar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Cronograma Anual",
            initialfile=f"Cronograma_Anual_{sufijo_sede}_{anio_inicio}.xlsx"
        )
        if not ruta_guardar:
            return
            
        try:
            ruta_plat = obtener_ruta_plantilla("plantilla_cronograma.xlsx")
            wb = openpyxl.load_workbook(ruta_plat)
            ws = wb.active
            
            Y = anio_inicio
            
            # Actualizar celda E4 (Título de la gestión)
            ws["E4"] = f"Cronograma de Mantenimientos Preventivos Gestion {Y}"
            
            # Actualizar trimestres en fila 6
            ws["K6"] = f"1ro TRIM {Y}"
            ws["N6"] = f"2do TRIM {Y}"
            ws["Q6"] = f"3ro TRIM {Y}"
            ws["T6"] = f"4to TRIM {Y}"
            
            # Obtener equipos filtrados según la sede activa (o todos si es acceso general)
            if hasattr(self.app, "equipo_cumple_sede_activa"):
                eqs_crono = [eq for eq in self.app.datos.get("equipos", []) if self.app.equipo_cumple_sede_activa(eq)]
            else:
                eqs_crono = list(self.app.datos.get("equipos", []))
            
            if not eqs_crono:
                messagebox.showwarning("Sin Equipos", "No se encontraron equipos registrados para la sede o centro de salud seleccionado.")
                return
                    
            # Eliminar duplicados por ID de equipo
            unicos_eqs = {}
            for eq in eqs_crono:
                eq_id = str(eq.get("id") or "").strip()
                if eq_id and eq_id not in unicos_eqs:
                    unicos_eqs[eq_id] = eq
            eqs_lista = list(unicos_eqs.values())
            
            # Ordenar: 1° Red de Salud -> 2° Centro de Salud -> 3° Área -> 4° Equipo -> 5° ID
            def sort_key_crono(e):
                return (
                    str(e.get("red_salud_nombre") or "ZZZ").lower(),
                    str(e.get("centro_salud_nombre") or "ZZZ").lower(),
                    str(e.get("area") or "ZZZ").lower(),
                    str(e.get("nombre") or "ZZZ").lower(),
                    str(e.get("id") or "ZZZ").lower()
                )
            eqs_lista.sort(key=sort_key_crono)
                
            start_row = 8
            thin_side = openpyxl.styles.borders.Side(style="thin", color="CBD5E1")
            grid_border = openpyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            font_body = openpyxl.styles.Font(name="Segoe UI", size=10)
            font_marca_x = openpyxl.styles.Font(name="Segoe UI", size=11, bold=True, color="1E3A8A")
            fill_x = openpyxl.styles.PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

            for idx, eq in enumerate(eqs_lista):
                r = start_row + idx
                
                # Columnas del nuevo formato:
                # B (2): N°
                # C (3): RED
                # D (4): CENTRO DE SALUD
                # E (5): AREA
                # F (6): EQUIPO
                # G (7): CODIGO DE ACTIVOS FIJOS
                # H (8): MARCA
                # I (9): MODELO
                # J (10): ESTADO
                
                ws.cell(row=r, column=2, value=idx + 1)
                ws.cell(row=r, column=3, value=eq.get("red_salud_nombre") or "-")
                ws.cell(row=r, column=4, value=eq.get("centro_salud_nombre") or "-")
                ws.cell(row=r, column=5, value=eq.get("area") or "-")
                ws.cell(row=r, column=6, value=eq.get("nombre") or "-")
                ws.cell(row=r, column=7, value=str(eq.get("id") or "-"))
                ws.cell(row=r, column=8, value=eq.get("marca") or "-")
                ws.cell(row=r, column=9, value=eq.get("modelo") or "-")
                
                estado_str = str(eq.get("estado") or "Operativo")
                ws.cell(row=r, column=10, value=estado_str)
                
                # Formato y estilo base de la fila
                for c in range(2, 23): # Columnas B (2) a V (22)
                    c_cell = ws.cell(row=r, column=c)
                    c_cell.font = font_body
                    c_cell.border = grid_border
                    
                    if c == 2 or c == 7 or c == 10: # N°, Código AF, Estado
                        c_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    elif c in range(11, 23): # Meses
                        c_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    else:
                        c_cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

                # Proyección para las marcas 'X' según criticidad
                if estado_str == "Baja":
                    # No colocar marcas 'X' para equipos de baja
                    pass
                else:
                    crit_eq = str(eq.get("criticidad") or "Riesgo Medio")
                    meses_eq = 3 if ("Alto" in crit_eq or crit_eq == "I") else (4 if ("Medio" in crit_eq or crit_eq == "II") else 6)
                    
                    f_reg = eq.get("fecha_adquisicion") or eq.get("fecha_registro", self.app.hoy)
                    if isinstance(f_reg, datetime):
                        f_reg = f_reg.date()
                    elif isinstance(f_reg, str):
                        try:
                            f_reg = datetime.strptime(f_reg, "%Y-%m-%d").date()
                        except:
                            f_reg = self.app.hoy
                    elif not isinstance(f_reg, date):
                        f_reg = self.app.hoy
                            
                    if eq.get("garantia") == "Con Garantía" and eq.get("fecha_vencimiento_garantia"):
                        f_venc_g = eq.get("fecha_vencimiento_garantia")
                        if isinstance(f_venc_g, str):
                            try:
                                f_venc_g = datetime.strptime(f_venc_g, "%Y-%m-%d").date()
                            except:
                                f_venc_g = None
                        elif isinstance(f_venc_g, datetime):
                            f_venc_g = f_venc_g.date()
                        if f_venc_g:
                            f_reg = f_venc_g + relativedelta(days=+1)

                    f_iter = f_reg
                    # Proyectar intervenciones
                    while f_iter.year < Y:
                        f_iter = f_iter + relativedelta(months=+meses_eq)
                        
                    while f_iter.year == Y:
                        c_idx = 11 + (f_iter.month - 1)
                        if 11 <= c_idx <= 22:
                            c_cell = ws.cell(row=r, column=c_idx)
                            c_cell.value = "X"
                            c_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                            c_cell.font = font_marca_x
                            c_cell.fill = fill_x
                        f_iter = f_iter + relativedelta(months=+meses_eq)
                
            # Borrar filas sobrantes si la plantilla tenía más filas previas
            total_rows = ws.max_row
            needed_rows = start_row + len(eqs_lista) - 1
            if total_rows > needed_rows and needed_rows >= start_row:
                ws.delete_rows(needed_rows + 1, total_rows - needed_rows)
                
            wb.save(ruta_guardar)
            messagebox.showinfo("Éxito", f"Cronograma Anual {Y} exportado correctamente con {len(eqs_lista)} equipos.")
        except Exception as e:
            messagebox.showerror("Error al Exportar", f"Hubo un error al generar el archivo:\n{e}")