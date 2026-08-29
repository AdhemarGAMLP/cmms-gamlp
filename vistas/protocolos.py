# vistas/protocolos.py
import os
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import PatternFill
import psycopg2.extras

from database import obtener_conexion
from estilos import *
from config import CARPETAS
from excel_utils import obtener_ruta_plantilla, escribir_en_celda_segura, exportar_excel_a_pdf

class VistaProtocolos(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=C_BG)
        self.app = app
        self.construir_ui()

    def construir_ui(self):
        f_top = ctk.CTkFrame(self, fg_color="transparent")
        f_top.pack(pady=(30, 15), padx=30, fill="x")
        ctk.CTkLabel(f_top, text="Protocolos Diarios de Electromedicina", font=ctk.CTkFont(size=28, weight="bold"), text_color=C_TEXT).pack(side="left")

        # Barra de Filtros
        f_filtros = ctk.CTkFrame(self, fg_color="transparent")
        f_filtros.pack(pady=(5, 10), padx=30, fill="x")
        
        self.filtro_fecha_var = ctk.StringVar()
        self.filtro_fecha_var.trace_add("write", lambda *args: self.refrescar_datos())
        ctk.CTkLabel(f_filtros, text="📅 Filtrar Fecha:", font=ctk.CTkFont(weight="bold"), text_color=C_TEXT).pack(side="left", padx=5)
        e_fecha = ctk.CTkEntry(f_filtros, textvariable=self.filtro_fecha_var, placeholder_text="AAAA-MM-DD", width=120, fg_color=C_CARD, border_color=C_BORDER, corner_radius=10)
        e_fecha.pack(side="left", padx=5)
        
        self.filtro_resp_var = ctk.StringVar()
        self.filtro_resp_var.trace_add("write", lambda *args: self.refrescar_datos())
        ctk.CTkLabel(f_filtros, text="👤 Responsable:", font=ctk.CTkFont(weight="bold"), text_color=C_TEXT).pack(side="left", padx=(15, 5))
        e_resp = ctk.CTkEntry(f_filtros, textvariable=self.filtro_resp_var, placeholder_text="Nombre de responsable...", width=200, fg_color=C_CARD, border_color=C_BORDER, corner_radius=10)
        e_resp.pack(side="left", padx=5)
        
        def limpiar_filtros():
            self.filtro_fecha_var.set("")
            self.filtro_resp_var.set("")
        ctk.CTkButton(f_filtros, text="Limpiar Filtros", width=100, fg_color=C_BG, text_color=C_BLUE, hover_color=C_BORDER, corner_radius=8, command=limpiar_filtros).pack(side="left", padx=15)

        marco = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        marco.pack(padx=30, pady=10, fill="both", expand=True)
        
        cols = ("Fecha", "Tipo Protocolo", "Turnos (M | T | N)", "Responsable(s)")
        f_tree_prot = ctk.CTkFrame(marco, fg_color="transparent")
        f_tree_prot.pack(pady=12, padx=12, fill="both", expand=True)
        self.tabla_prot = ttk.Treeview(f_tree_prot, columns=cols, show="headings")
        self.tabla_prot.tag_configure("completo", background="#D1FAE5", foreground="#047857")
        self.tabla_prot.tag_configure("incompleto", background="#FEE2E2", foreground="#B91C1C")
        self.tabla_prot.tag_configure("futuro", background="#FFFFFF", foreground=C_TEXT)
        scrollbar_prot = ttk.Scrollbar(f_tree_prot, orient="vertical", command=self.tabla_prot.yview, style="Vertical.TScrollbar")
        self.tabla_prot.configure(yscrollcommand=scrollbar_prot.set)
        for c in cols:
            self.tabla_prot.heading(c, text=c)
            self.tabla_prot.column(c, anchor="center")
        self.tabla_prot.pack(side="left", fill="both", expand=True)
        scrollbar_prot.pack(side="right", fill="y", padx=(5, 0))

        f_bot = ctk.CTkFrame(self, fg_color="transparent")
        f_bot.pack(pady=(10, 25), padx=30, fill="x")
        
        ctk.CTkButton(f_bot, text="✚ Registrar/Modificar Protocolo", font=ctk.CTkFont(weight="bold", size=13), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, corner_radius=10, height=42, command=self.abrir_formulario_protocolo).pack(side="left", expand=True, padx=8)
        ctk.CTkButton(f_bot, text="📄 Abrir Excel Completo", font=ctk.CTkFont(weight="bold", size=13), fg_color=C_GREEN, hover_color=C_GREEN_HOVER, corner_radius=10, height=42, command=self.abrir_excel_protocolo).pack(side="left", expand=True, padx=8)
        ctk.CTkButton(f_bot, text="⬇ Exportar a PDF", font=ctk.CTkFont(weight="bold", size=13), fg_color=C_ORANGE, hover_color=C_ORANGE_LIGHT, corner_radius=10, height=42, command=self.exportar_pdf_protocolo).pack(side="left", expand=True, padx=8)


    def refrescar_datos(self):
        for i in self.tabla_prot.get_children():
            self.tabla_prot.delete(i)
            
        from collections import defaultdict
        grouped = defaultdict(lambda: {"Mañana": False, "Tarde": False, "Noche": False, "responsables": set()})
        
        filtro_f = self.filtro_fecha_var.get().strip().lower() if hasattr(self, "filtro_fecha_var") else ""
        filtro_r = self.filtro_resp_var.get().strip().lower() if hasattr(self, "filtro_resp_var") else ""
        
        for p in self.app.datos.get("protocolos", []):
            fecha = p['fecha']
            tipo = p['tipo_protocolo']
            turno = p['turno']
            resp = p.get('responsable', '')
            
            if filtro_f and filtro_f not in str(fecha).lower():
                continue
            if filtro_r and filtro_r not in str(resp).lower():
                continue
            # Normalizar tipo y turno por problemas de codificación de la BD
            if tipo:
                if "Gases" in tipo:
                    tipo = "Gases Medicinales"
                elif "Resonador" in tipo:
                    tipo = "Resonador Magnético"
            if turno:
                if turno.startswith("Ma"):
                    turno = "Mañana"
                elif turno.startswith("Ta"):
                    turno = "Tarde"
                elif turno.startswith("No"):
                    turno = "Noche"
            
            key = (fecha, tipo)
            if turno in grouped[key]:
                grouped[key][turno] = True
            if resp:
                grouped[key]["responsables"].add(resp)
                
        # Ordenar por fecha descendente
        sorted_keys = sorted(grouped.keys(), key=lambda x: x[0], reverse=True)
        
        from datetime import datetime
        now = datetime.now()
        hoy_str = now.strftime("%Y-%m-%d")
        curr_hour = now.hour
        
        def get_shift_status(fecha_str, turno_name, is_filled):
            if is_filled:
                return "[✔]"
            # Si no está lleno
            if fecha_str < hoy_str:
                # Fechas pasadas siempre son salteadas (rojo)
                return "[✘]"
            elif fecha_str > hoy_str:
                # Fechas futuras no se tienen que llenar aún (blanco/gris)
                return "[ ]"
            else:
                # Hoy
                if turno_name == "Mañana":
                    # Mañana se considera salteada si ya es tarde (después de las 14:00)
                    return "[✘]" if curr_hour >= 14 else "[ ]"
                elif turno_name == "Tarde":
                    # Tarde se considera salteada si ya es de noche/mañana (después de las 20:00)
                    return "[✘]" if curr_hour >= 20 else "[ ]"
                else: # Noche
                    # Noche no se considera salteada durante el día actual
                    return "[ ]"

        for key in sorted_keys:
            fecha, tipo = key
            info = grouped[key]
            
            fecha_str = str(fecha)
            m_status = get_shift_status(fecha_str, "Mañana", info["Mañana"])
            t_status = get_shift_status(fecha_str, "Tarde", info["Tarde"])
            n_status = get_shift_status(fecha_str, "Noche", info["Noche"])
            turnos_str = f"  {m_status}     {t_status}     {n_status}  "
            
            resps_str = ", ".join(sorted(list(info["responsables"])))
            
            # Determinar tag
            if info["Mañana"] and info["Tarde"] and info["Noche"]:
                tag_row = "completo"
            elif fecha_str > hoy_str:
                tag_row = "futuro"
            else:
                tag_row = "incompleto"
                
            self.tabla_prot.insert("", "end", values=(fecha_str, tipo, turnos_str, resps_str), tags=(tag_row,))

    def abrir_formulario_protocolo(self, fecha_preset=None):
        if not fecha_preset:
            sel = self.tabla_prot.focus()
            if sel:
                val = self.tabla_prot.item(sel, "values")
                fecha_preset = val[0]
        v = ctk.CTkToplevel(self.app)
        v.title("Registro de Protocolo Diario")
        v.geometry("800x800")
        v.transient(self.app)
        v.grab_set()
        v.configure(fg_color=C_BG)

        ctk.CTkLabel(v, text="Registrar Turno", font=ctk.CTkFont(size=20, weight="bold"), text_color=C_TEXT).pack(pady=15)
        sf = ctk.CTkScrollableFrame(v, fg_color=C_CARD, corner_radius=12)
        sf.pack(fill="both", expand=True, padx=20, pady=10)

        # Contenedor de metadata compacto y elegante
        f_meta = ctk.CTkFrame(sf, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_meta.pack(fill="x", padx=5, pady=5)
        
        f_meta.columnconfigure(0, weight=1)
        f_meta.columnconfigure(1, weight=1)
        f_meta.columnconfigure(2, weight=1)
        f_meta.columnconfigure(3, weight=2)
        
        # Fecha
        ctk.CTkLabel(f_meta, text="Fecha", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).grid(row=0, column=0, padx=10, pady=(10, 2), sticky="w")
        cal_fecha = DateEntry(f_meta, width=12, font=('Segoe UI', 11), background=C_BLUE, foreground='white', borderwidth=0, date_pattern='y-mm-dd')
        if fecha_preset:
            cal_fecha.set_date(fecha_preset)
        cal_fecha.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="ew")
        
        # Tipo de Protocolo
        ctk.CTkLabel(f_meta, text="Tipo de Protocolo", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).grid(row=0, column=1, padx=10, pady=(10, 2), sticky="w")
        var_tipo = ctk.StringVar(value="Gases Medicinales")
        c_tipo = ctk.CTkComboBox(f_meta, variable=var_tipo, values=["Gases Medicinales", "Resonador Magnético"], width=160, fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
        c_tipo.grid(row=1, column=1, padx=10, pady=(0, 10), sticky="ew")
        
        # Turno del Día
        ctk.CTkLabel(f_meta, text="Turno", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).grid(row=0, column=2, padx=10, pady=(10, 2), sticky="w")
        var_turno = ctk.StringVar(value="Mañana")
        c_turno_combo = ctk.CTkComboBox(f_meta, variable=var_turno, values=["Mañana", "Tarde", "Noche"], width=120, fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
        c_turno_combo.grid(row=1, column=2, padx=10, pady=(0, 10), sticky="ew")
        
        # Responsable
        ctk.CTkLabel(f_meta, text="Responsable / Técnico", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).grid(row=0, column=3, padx=10, pady=(10, 2), sticky="w")
        e_resp = ctk.CTkEntry(f_meta, placeholder_text="Nombre del encargado", fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
        e_resp.grid(row=1, column=3, padx=10, pady=(0, 10), sticky="ew")

        lbl_estado_existente = ctk.CTkLabel(sf, text="", font=ctk.CTkFont(size=12), text_color=C_ORANGE)
        lbl_estado_existente.pack(anchor="w", padx=10, pady=5)



        # Contenedor para Gases Medicinales
        f_gases = ctk.CTkFrame(sf, fg_color="transparent")
        self.gases_fields = {}

        def crear_entrada(parent, etiqueta, row, col, width=130):
            ctk.CTkLabel(parent, text=etiqueta, font=ctk.CTkFont(size=10, weight="bold"), text_color=C_TEXT).grid(row=row*2, column=col, padx=4, pady=(4,0), sticky="w")
            entry = ctk.CTkEntry(parent, width=width, fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
            entry.grid(row=row*2+1, column=col, padx=4, pady=(0,4), sticky="ew")
            return entry

        def crear_combo(parent, etiqueta, valores, row, col, width=130):
            ctk.CTkLabel(parent, text=etiqueta, font=ctk.CTkFont(size=10, weight="bold"), text_color=C_TEXT).grid(row=row*2, column=col, padx=4, pady=(4,0), sticky="w")
            combo = ctk.CTkComboBox(parent, values=valores, width=width, fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
            combo.grid(row=row*2+1, column=col, padx=4, pady=(0,4), sticky="ew")
            return combo

        def crear_tiempo_casillas(parent, etiqueta, row, col):
            ctk.CTkLabel(parent, text=etiqueta, font=ctk.CTkFont(size=10, weight="bold"), text_color=C_TEXT).grid(row=row*2, column=col, padx=4, pady=(4,0), sticky="w")
            f_time = ctk.CTkFrame(parent, fg_color="transparent")
            f_time.grid(row=row*2+1, column=col, padx=4, pady=(0,4), sticky="ew")
            e1 = ctk.CTkEntry(f_time, width=45, fg_color=C_BG, border_color=C_BORDER, corner_radius=5)
            e1.pack(side="left", padx=1)
            e2 = ctk.CTkEntry(f_time, width=45, fg_color=C_BG, border_color=C_BORDER, corner_radius=5)
            e2.pack(side="left", padx=1)
            e3 = ctk.CTkEntry(f_time, width=45, fg_color=C_BG, border_color=C_BORDER, corner_radius=5)
            e3.pack(side="left", padx=1)
            return (e1, e2, e3)

        # Contenedor para Resonador Magnético
        f_resonador = ctk.CTkFrame(sf, fg_color="transparent")
        self.resonador_fields = {}

        # 1. Compresor de A ICC (Módulo de Enfriamiento)
        ctk.CTkLabel(f_resonador, text="1. Compresor de A ICC (Módulo de Enfriamiento)", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(10,5), padx=5)
        f_comp = ctk.CTkFrame(f_resonador, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_comp.pack(fill="x", padx=5, pady=5)
        
        self.resonador_fields['res_comp_estado'] = ctk.CTkCheckBox(f_comp, text="Compresor Encendido / Verificado", font=ctk.CTkFont(size=11, weight="bold"))
        self.resonador_fields['res_comp_estado'].grid(row=1, column=0, padx=15, pady=15, sticky="w")
        self.resonador_fields['res_comp_obs'] = crear_entrada(f_comp, "Observaciones del Compresor", 0, 1, width=420)

        # 2. Flujograma de Agua
        ctk.CTkLabel(f_resonador, text="2. Flujograma de Agua (Flujómetro mín 50 a 60 LPM)", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(15,5), padx=5)
        f_flujo = ctk.CTkFrame(f_resonador, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_flujo.pack(fill="x", padx=5, pady=5)
        self.resonador_fields['res_flujo_estado'] = crear_entrada(f_flujo, "Flujograma de Agua (LPM - Mínimo 50 a 60)", 0, 0, width=250)
        self.resonador_fields['res_flujo_obs'] = crear_entrada(f_flujo, "Observaciones de Flujograma", 0, 1, width=420)

        # 3. Temperatura de Agua
        ctk.CTkLabel(f_resonador, text="3. Temperatura de Agua (Water Temp - Rango: 8 °C a 11 °C)", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(15,5), padx=5)
        f_temp = ctk.CTkFrame(f_resonador, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_temp.pack(fill="x", padx=5, pady=5)
        self.resonador_fields['res_temp_valor'] = crear_entrada(f_temp, "Temperatura de Agua (°C)", 0, 0, width=250)
        self.resonador_fields['res_temp_obs'] = crear_entrada(f_temp, "Observaciones de Temperatura", 0, 1, width=420)

        # 4. Flujo de Agua
        ctk.CTkLabel(f_resonador, text="4. Flujo de Agua (Water Flow - Rango: 5 LPM a 6 LPM)", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(15,5), padx=5)
        f_flow = ctk.CTkFrame(f_resonador, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_flow.pack(fill="x", padx=5, pady=5)
        self.resonador_fields['res_flow_valor'] = crear_entrada(f_flow, "Flujo de Agua (LPM)", 0, 0, width=250)
        self.resonador_fields['res_flow_obs'] = crear_entrada(f_flow, "Observaciones de Flujo", 0, 1, width=420)

        # 1. Oxígeno Medicinal
        ctk.CTkLabel(f_gases, text="1. Central de Oxígeno Medicinal", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(10,5), padx=5)
        f_ox = ctk.CTkFrame(f_gases, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_ox.pack(fill="x", padx=5, pady=5)
        
        self.gases_fields['ox_tanque_mano'] = crear_entrada(f_ox, "Mano. Tanque", 0, 0)
        self.gases_fields['ox_tanque_nivel'] = crear_entrada(f_ox, "Nivel Tanque (in)", 0, 1)
        self.gases_fields['ox_red_p_ent'] = crear_entrada(f_ox, "Reduc. P. Ent.", 0, 2)
        self.gases_fields['ox_red_p_sal'] = crear_entrada(f_ox, "Reduc. P. Sal.", 0, 3)
        
        self.gases_fields['ox_res_rd_ent'] = crear_entrada(f_ox, "Rampa Der. P. Ent.", 1, 0)
        self.gases_fields['ox_res_rd_sal'] = crear_entrada(f_ox, "Rampa Der. P. Sal.", 1, 1)
        self.gases_fields['ox_res_ri_ent'] = crear_entrada(f_ox, "Rampa Izq. P. Ent.", 1, 2)
        self.gases_fields['ox_res_ri_sal'] = crear_entrada(f_ox, "Rampa Izq. P. Sal.", 1, 3)
        
        self.gases_fields['ox_res_emerg_ent'] = crear_entrada(f_ox, "Central Emerg. P. Ent.", 2, 0)
        self.gases_fields['ox_res_emerg_sal'] = crear_entrada(f_ox, "Central Emerg. P. Sal.", 2, 1)
        self.gases_fields['ox_res_sw_ent'] = crear_entrada(f_ox, "Switch Princ. P. Ent.", 2, 2)
        self.gases_fields['ox_res_sw_sal'] = crear_entrada(f_ox, "Switch Princ. P. Sal.", 2, 3)
        
        self.gases_fields['ox_res_sw_emerg_ent'] = crear_entrada(f_ox, "Switch Emerg. P. Ent.", 3, 0)
        self.gases_fields['ox_res_sw_emerg_sal'] = crear_entrada(f_ox, "Switch Emerg. P. Sal.", 3, 1)

        # 2. Vacío Medicinal (Diseño Cuadrícula Mejorado con Checkboxes)
        ctk.CTkLabel(f_gases, text="2. Central de Vacío Medicinal", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(15,5), padx=5)
        f_vac = ctk.CTkFrame(f_gases, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_vac.pack(fill="x", padx=5, pady=5)
        
        # Entradas de tanque y red
        self.gases_fields['vac_tanque'] = crear_entrada(f_vac, "Vacío Tanque (mmHg)", 0, 0)
        self.gases_fields['vac_red'] = crear_entrada(f_vac, "Vacío Red (mmHg)", 0, 1)

        # Separador / Encabezados de bomba
        ctk.CTkLabel(f_vac, text="Bomba", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT).grid(row=2, column=0, padx=10, pady=(15,5), sticky="w")
        ctk.CTkLabel(f_vac, text="Principal", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT).grid(row=2, column=1, padx=10, pady=(15,5), sticky="w")
        ctk.CTkLabel(f_vac, text="Marcha (Check)", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT).grid(row=2, column=2, padx=10, pady=(15,5), sticky="w")
        ctk.CTkLabel(f_vac, text="Nivel de Aceite", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT).grid(row=2, column=3, padx=10, pady=(15,5), sticky="w")

        # Bomba 1
        ctk.CTkLabel(f_vac, text="Bomba de Vacío 1", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b1_principal'] = ctk.CTkCheckBox(f_vac, text="")
        self.gases_fields['vac_b1_principal'].grid(row=3, column=1, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b1_marca'] = ctk.CTkCheckBox(f_vac, text="")
        self.gases_fields['vac_b1_marca'].grid(row=3, column=2, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b1_aceite'] = ctk.CTkComboBox(f_vac, values=["0% (0/5)", "20% (1/5)", "40% (2/5)", "60% (3/5)", "80% (4/5)", "100% (5/5)"], width=130, fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
        self.gases_fields['vac_b1_aceite'].grid(row=3, column=3, padx=10, pady=5, sticky="ew")

        # Bomba 2
        ctk.CTkLabel(f_vac, text="Bomba de Vacío 2", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).grid(row=4, column=0, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b2_principal'] = ctk.CTkCheckBox(f_vac, text="")
        self.gases_fields['vac_b2_principal'].grid(row=4, column=1, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b2_marca'] = ctk.CTkCheckBox(f_vac, text="")
        self.gases_fields['vac_b2_marca'].grid(row=4, column=2, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b2_aceite'] = ctk.CTkComboBox(f_vac, values=["0% (0/5)", "20% (1/5)", "40% (2/5)", "60% (3/5)", "80% (4/5)", "100% (5/5)"], width=130, fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
        self.gases_fields['vac_b2_aceite'].grid(row=4, column=3, padx=10, pady=5, sticky="ew")

        # Bomba 3
        ctk.CTkLabel(f_vac, text="Bomba de Vacío 3", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).grid(row=5, column=0, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b3_principal'] = ctk.CTkCheckBox(f_vac, text="")
        self.gases_fields['vac_b3_principal'].grid(row=5, column=1, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b3_marca'] = ctk.CTkCheckBox(f_vac, text="")
        self.gases_fields['vac_b3_marca'].grid(row=5, column=2, padx=10, pady=5, sticky="w")
        self.gases_fields['vac_b3_aceite'] = ctk.CTkComboBox(f_vac, values=["0% (0/5)", "20% (1/5)", "40% (2/5)", "60% (3/5)", "80% (4/5)", "100% (5/5)"], width=130, fg_color=C_BG, border_color=C_BORDER, corner_radius=8)
        self.gases_fields['vac_b3_aceite'].grid(row=5, column=3, padx=10, pady=5, sticky="ew")

        # 3. Aire Comprimido
        ctk.CTkLabel(f_gases, text="3. Central de Aire Comprimido (Compresores / Secadores)", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(15,5), padx=5)
        f_aire = ctk.CTkFrame(f_gases, fg_color=C_CARD, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_aire.pack(fill="x", padx=5, pady=5)
        
        # Compresor 1
        self.gases_fields['aire_c1_p'] = crear_entrada(f_aire, "C1 P. (bar)", 0, 0)
        self.gases_fields['aire_c1_t'] = crear_entrada(f_aire, "C1 T. (°C)", 0, 1)
        self.gases_fields['aire_c1_m1'] = crear_tiempo_casillas(f_aire, "C1 Tiempo M1", 0, 2)
        self.gases_fields['aire_c1_m2'] = crear_tiempo_casillas(f_aire, "C1 Tiempo M2", 0, 3)
        self.gases_fields['aire_c1_m3'] = crear_tiempo_casillas(f_aire, "C1 Tiempo M3", 1, 0)
        self.gases_fields['aire_c1_m4'] = crear_tiempo_casillas(f_aire, "C1 Tiempo M4", 1, 1)
        
        # Compresor 2
        self.gases_fields['aire_c2_p'] = crear_entrada(f_aire, "C2 P. (bar)", 2, 0)
        self.gases_fields['aire_c2_t'] = crear_entrada(f_aire, "C2 T. (°C)", 2, 1)
        self.gases_fields['aire_c2_m1'] = crear_tiempo_casillas(f_aire, "C2 Tiempo M1", 2, 2)
        self.gases_fields['aire_c2_m2'] = crear_tiempo_casillas(f_aire, "C2 Tiempo M2", 2, 3)
        self.gases_fields['aire_c2_m3'] = crear_tiempo_casillas(f_aire, "C2 Tiempo M3", 3, 0)
        self.gases_fields['aire_c2_m4'] = crear_tiempo_casillas(f_aire, "C2 Tiempo M4", 3, 1)
        
        # Secadores
        self.gases_fields['aire_s1_p'] = crear_entrada(f_aire, "S1 P. Aire (psi)", 4, 0)
        self.gases_fields['aire_s1_t'] = crear_entrada(f_aire, "S1 T. Aire (°C)", 4, 1)
        self.gases_fields['aire_s1_p_evap'] = crear_entrada(f_aire, "S1 P. Evap (psi)", 4, 2)
        self.gases_fields['aire_s1_t_evap'] = crear_entrada(f_aire, "S1 T. Evap (°C)", 4, 3)
        
        self.gases_fields['aire_s2_p'] = crear_entrada(f_aire, "S2 P. Aire (psi)", 5, 0)
        self.gases_fields['aire_s2_t'] = crear_entrada(f_aire, "S2 T. Aire (°C)", 5, 1)
        self.gases_fields['aire_s2_p_evap'] = crear_entrada(f_aire, "S2 P. Evap (psi)", 5, 2)
        self.gases_fields['aire_s2_t_evap'] = crear_entrada(f_aire, "S2 T. Evap (°C)", 5, 3)

        # 4. Observaciones
        ctk.CTkLabel(f_gases, text="Observaciones Generales:", font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).pack(anchor="w", pady=(15,5), padx=5)
        self.txt_obs_gases = ctk.CTkTextbox(f_gases, height=80, fg_color=C_BG, border_color=C_BORDER, border_width=1, corner_radius=10)
        self.txt_obs_gases.pack(fill="x", padx=10, pady=5)

        def get_field_val(field):
            if isinstance(field, tuple):  # triple of CTkEntry
                return [e.get().strip() for e in field]
            elif isinstance(field, ctk.CTkCheckBox):
                return 'X' if field.get() == 1 else ''
            else:
                return field.get().strip()

        def set_field_val(field, val):
            if isinstance(field, tuple):
                for e, v in zip(field, val if val and isinstance(val, list) else ["","",""]):
                    e.delete(0, "end")
                    e.insert(0, str(v))
            elif isinstance(field, ctk.CTkCheckBox):
                if val == 'X' or val == 1 or val is True:
                    field.select()
                else:
                    field.deselect()
            else:
                if isinstance(field, ctk.CTkComboBox):
                    field.set(str(val) if val else "")
                else:
                    field.delete(0, "end")
                    field.insert(0, str(val) if val else "")

        def verificar_existente(*args):
            try:
                f = cal_fecha.get_date().strftime("%Y-%m-%d")
            except: 
                return
            conn = obtener_conexion()
            if not conn: 
                return
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            
            # Consultar todos los registros del día para normalizarlos en Python
            cur.execute("SELECT tipo_protocolo, turno, responsable, datos FROM protocolos WHERE fecha=%s", (f,))
            rows = cur.fetchall()
            cur.close(); conn.close()
            
            tipo_sel = var_tipo.get()
            turno_sel = var_turno.get()
            
            # Obtener responsable por defecto según el usuario logueado
            usuario_actual = self.app.usuario_actual
            default_resp = ""
            if usuario_actual:
                if usuario_actual.get("nombre_usuario") == "admin":
                    default_resp = "Rudel Adhemar Santos Medina"
                else:
                    default_resp = usuario_actual.get("nombre_completo", "")

            # Filtrar y normalizar
            r = None
            todos_turnos = []
            for row in rows:
                t_tipo = row['tipo_protocolo']
                t_turno = row['turno']
                
                # Normalizar tipo
                if t_tipo:
                    if "Gases" in t_tipo:
                        t_tipo_norm = "Gases Medicinales"
                    elif "Resonador" in t_tipo:
                        t_tipo_norm = "Resonador Magnético"
                    else:
                        t_tipo_norm = t_tipo
                else:
                    t_tipo_norm = ""
                    
                # Normalizar turno
                if t_turno:
                    if t_turno.startswith("Ma"):
                        t_turno_norm = "Mañana"
                    elif t_turno.startswith("Ta"):
                        t_turno_norm = "Tarde"
                    elif t_turno.startswith("No"):
                        t_turno_norm = "Noche"
                    else:
                        t_turno_norm = t_turno
                else:
                    t_turno_norm = ""
                    
                if t_tipo_norm == tipo_sel:
                    todos_turnos.append(row)
                    if t_turno_norm == turno_sel:
                        r = row
            
            if tipo_sel == "Gases Medicinales":
                obs_del_dia = ""
                for t in todos_turnos:
                    t_datos = t['datos']
                    if t_datos:
                        if isinstance(t_datos, str):
                            import json
                            t_datos = json.loads(t_datos)
                        t_obs = t_datos.get('observaciones_gases', '')
                        if t_obs and len(t_obs) > len(obs_del_dia):
                            obs_del_dia = t_obs

                if r:
                    lbl_estado_existente.configure(text=f"⚠️ Ya registrado por: {r['responsable']} (Se sobreescribirá al guardar)")
                    e_resp.delete(0, "end")
                    e_resp.insert(0, r['responsable'])
                    if r['datos']:
                        import json
                        datos_cargados = r['datos']
                        if isinstance(datos_cargados, str):
                            datos_cargados = json.loads(datos_cargados)
                        for key, field in self.gases_fields.items():
                            set_field_val(field, datos_cargados.get(key))
                        self.txt_obs_gases.delete("1.0", "end")
                        self.txt_obs_gases.insert("1.0", datos_cargados.get('observaciones_gases', ''))
                else:
                    lbl_estado_existente.configure(text="✓ Turno libre, sin registro previo.")
                    e_resp.delete(0, "end")
                    e_resp.insert(0, default_resp)
                    for key, field in self.gases_fields.items():
                        set_field_val(field, None)
                    self.txt_obs_gases.delete("1.0", "end")
                    if obs_del_dia:
                        self.txt_obs_gases.insert("1.0", obs_del_dia)
            
            else: # Resonador Magnético
                obs_dia = {'res_comp_obs': '', 'res_flujo_obs': '', 'res_temp_obs': '', 'res_flow_obs': ''}
                for t in todos_turnos:
                    t_datos = t['datos']
                    if t_datos:
                        if isinstance(t_datos, str):
                            import json
                            t_datos = json.loads(t_datos)
                        for k in obs_dia.keys():
                            val = t_datos.get(k, '')
                            if val and len(val) > len(obs_dia[k]):
                                obs_dia[k] = val

                if r:
                    lbl_estado_existente.configure(text=f"⚠️ Ya registrado por: {r['responsable']} (Se sobreescribirá al guardar)")
                    e_resp.delete(0, "end")
                    e_resp.insert(0, r['responsable'])
                    if r['datos']:
                        import json
                        datos_cargados = r['datos']
                        if isinstance(datos_cargados, str):
                            datos_cargados = json.loads(datos_cargados)
                        for key, field in self.resonador_fields.items():
                            set_field_val(field, datos_cargados.get(key))
                else:
                    lbl_estado_existente.configure(text="✓ Turno libre, sin registro previo.")
                    e_resp.delete(0, "end")
                    e_resp.insert(0, default_resp)
                    for key, field in self.resonador_fields.items():
                        if key in obs_dia:
                            set_field_val(field, obs_dia[key])
                        else:
                            set_field_val(field, None)

        var_tipo.trace_add("write", verificar_existente)
        var_turno.trace_add("write", verificar_existente)
        cal_fecha.bind("<<DateEntrySelected>>", verificar_existente)
        verificar_existente()

        def toggle_secadores(*args):
            turno = var_turno.get()
            tipo = var_tipo.get()
            is_disabled = (turno == "Noche" and tipo == "Gases Medicinales")
            for key in ['aire_s1_p', 'aire_s1_t', 'aire_s1_p_evap', 'aire_s1_t_evap',
                         'aire_s2_p', 'aire_s2_t', 'aire_s2_p_evap', 'aire_s2_t_evap']:
                if key in self.gases_fields:
                    field = self.gases_fields[key]
                    if is_disabled:
                        field.delete(0, "end")
                        field.configure(state="disabled", fg_color="#E5E5EA")
                    else:
                        field.configure(state="normal", fg_color=C_BG)

        var_turno.trace_add("write", toggle_secadores)
        var_tipo.trace_add("write", toggle_secadores)

        def toggle_campos(*args):
            tipo = var_tipo.get()
            if tipo == "Resonador Magnético":
                f_resonador.pack(fill="x", padx=10, pady=10)
                f_gases.pack_forget()
            else:
                f_resonador.pack_forget()
                f_gases.pack(fill="x", padx=10, pady=10)
            toggle_secadores()

        var_tipo.trace_add("write", toggle_campos)
        toggle_campos()

        def guardar():
            if not e_resp.get().strip():
                messagebox.showwarning("Aviso", "Indique el responsable.")
                return

            tipo_prot = var_tipo.get()
            turno_sel = var_turno.get()
            fecha_str = cal_fecha.get_date().strftime("%Y-%m-%d")
            nombre_base = "Protocolo_Gases_" if tipo_prot == "Gases Medicinales" else "Protocolo_Resonador_"
            nombre_excel = f"{nombre_base}{fecha_str.replace('-','')}.xlsx"
            ruta_salida = os.path.join(CARPETAS["protocolos"], nombre_excel)
            plantilla_base = "plantilla_gases.xlsx" if tipo_prot == "Gases Medicinales" else "plantilla_rasonador.xlsx"
            ruta_plantilla = obtener_ruta_plantilla(plantilla_base)

            if not os.path.exists(ruta_plantilla) and not os.path.exists(ruta_salida):
                messagebox.showerror("Error", f"No se encuentra la plantilla '{plantilla_base}'.")
                return

            conn = None
            try:
                wb = openpyxl.load_workbook(ruta_salida) if os.path.exists(ruta_salida) else openpyxl.load_workbook(ruta_plantilla)
                ws = wb.active
                grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                if tipo_prot == "Gases Medicinales":
                    # Mapear filas de turnos según el diseño del usuario
                    if turno_sel == "Mañana":
                        row_ox = 16
                        row_vac = 26
                        row_comp = 35
                        row_sec = 41
                    elif turno_sel == "Tarde":
                        row_ox = 17
                        row_vac = 27
                        row_comp = 36
                        row_sec = 42
                    else:  # Noche
                        row_ox = 18
                        row_vac = 28
                        row_comp = 37
                        row_sec = None  # No secador row for night shift

                    # Escribir la fecha actual y el responsable
                    escribir_en_celda_segura(ws, 'L4', fecha_str, centrar=True)
                    escribir_en_celda_segura(ws, 'M5', e_resp.get(), centrar=True)

                    # Escribir la hora actual en la columna B de cada sección activa ( Hrs. )
                    from datetime import datetime
                    hora_actual = datetime.now().strftime("%H:%M")
                    
                    escribir_en_celda_segura(ws, f'B{row_ox}', hora_actual, centrar=True)
                    escribir_en_celda_segura(ws, f'B{row_vac}', hora_actual, centrar=True)
                    escribir_en_celda_segura(ws, f'B{row_comp}', hora_actual, centrar=True)
                    if row_sec:
                        escribir_en_celda_segura(ws, f'B{row_sec}', hora_actual, centrar=True)

                    # Asegurar restauración explícita de Hrs. para evitar que se borren
                    escribir_en_celda_segura(ws, 'B12', 'Hrs.', centrar=True)
                    escribir_en_celda_segura(ws, 'B22', 'Hrs.', centrar=True)
                    escribir_en_celda_segura(ws, 'B32', 'Hrs.', centrar=True)
                    escribir_en_celda_segura(ws, 'B38', 'Hrs.', centrar=True)

                    def w(celda, val):
                        escribir_en_celda_segura(ws, celda, val, centrar=True)

                    fraccion_map = {
                        "0% (0/5)": "0/5",
                        "20% (1/5)": "1/5",
                        "40% (2/5)": "2/5",
                        "60% (3/5)": "3/5",
                        "80% (4/5)": "4/5",
                        "100% (5/5)": "5/5"
                    }
                    def get_aceite_val(field):
                        val_raw = get_field_val(field)
                        return fraccion_map.get(val_raw, val_raw)

                    # --- 1. Oxígeno Central ---
                    w(f'D{row_ox}', get_field_val(self.gases_fields['ox_tanque_mano']))
                    w(f'H{row_ox}', get_field_val(self.gases_fields['ox_tanque_nivel']))
                    w(f'L{row_ox}', get_field_val(self.gases_fields['ox_red_p_ent']))
                    w(f'O{row_ox}', get_field_val(self.gases_fields['ox_red_p_sal']))
                    w(f'R{row_ox}', get_field_val(self.gases_fields['ox_res_rd_ent']))
                    w(f'U{row_ox}', get_field_val(self.gases_fields['ox_res_rd_sal']))
                    w(f'X{row_ox}', get_field_val(self.gases_fields['ox_res_ri_ent']))
                    w(f'AA{row_ox}', get_field_val(self.gases_fields['ox_res_ri_sal']))
                    w(f'AD{row_ox}', get_field_val(self.gases_fields['ox_res_emerg_ent']))
                    w(f'AG{row_ox}', get_field_val(self.gases_fields['ox_res_emerg_sal']))
                    w(f'AJ{row_ox}', get_field_val(self.gases_fields['ox_res_sw_ent']))
                    w(f'AM{row_ox}', get_field_val(self.gases_fields['ox_res_sw_sal']))
                    w(f'AP{row_ox}', get_field_val(self.gases_fields['ox_res_sw_emerg_ent']))
                    w(f'AS{row_ox}', get_field_val(self.gases_fields['ox_res_sw_emerg_sal']))

                    # --- 2. Vacío Medicinal ---
                    w(f'D{row_vac}', get_field_val(self.gases_fields['vac_tanque']))
                    w(f'H{row_vac}', get_field_val(self.gases_fields['vac_red']))
                    
                    w(f'L{row_vac}', get_field_val(self.gases_fields['vac_b1_principal']))
                    w(f'P{row_vac}', get_field_val(self.gases_fields['vac_b1_marca']))
                    w(f'T{row_vac}', get_aceite_val(self.gases_fields['vac_b1_aceite']))
                    
                    w(f'X{row_vac}', get_field_val(self.gases_fields['vac_b2_principal']))
                    w(f'AB{row_vac}', get_field_val(self.gases_fields['vac_b2_marca']))
                    w(f'AF{row_vac}', get_aceite_val(self.gases_fields['vac_b2_aceite']))
                    
                    w(f'AJ{row_vac}', get_field_val(self.gases_fields['vac_b3_principal']))
                    w(f'AN{row_vac}', get_field_val(self.gases_fields['vac_b3_marca']))
                    w(f'AR{row_vac}', get_aceite_val(self.gases_fields['vac_b3_aceite']))

                    # --- 3. Aire Comprimido (Compresores) ---
                    w(f'D{row_comp}', get_field_val(self.gases_fields['aire_c1_p']))
                    w(f'G{row_comp}', get_field_val(self.gases_fields['aire_c1_t']))
                    
                    t_c1_m1 = get_field_val(self.gases_fields['aire_c1_m1'])
                    w(f'J{row_comp}', t_c1_m1[0])
                    w(f'L{row_comp}', t_c1_m1[1])
                    w(f'M{row_comp}', t_c1_m1[2])
                    
                    t_c1_m2 = get_field_val(self.gases_fields['aire_c1_m2'])
                    w(f'N{row_comp}', t_c1_m2[0])
                    w(f'P{row_comp}', t_c1_m2[1])
                    w(f'Q{row_comp}', t_c1_m2[2])
                    
                    t_c1_m3 = get_field_val(self.gases_fields['aire_c1_m3'])
                    w(f'R{row_comp}', t_c1_m3[0])
                    w(f'T{row_comp}', t_c1_m3[1])
                    w(f'U{row_comp}', t_c1_m3[2])
                    
                    t_c1_m4 = get_field_val(self.gases_fields['aire_c1_m4'])
                    w(f'V{row_comp}', t_c1_m4[0])
                    w(f'X{row_comp}', t_c1_m4[1])
                    w(f'Y{row_comp}', t_c1_m4[2])

                    # Compresor 2
                    w(f'Z{row_comp}', get_field_val(self.gases_fields['aire_c2_p']))
                    w(f'AC{row_comp}', get_field_val(self.gases_fields['aire_c2_t']))
                    
                    t_c2_m1 = get_field_val(self.gases_fields['aire_c2_m1'])
                    w(f'AF{row_comp}', t_c2_m1[0])
                    w(f'AH{row_comp}', t_c2_m1[1])
                    w(f'AI{row_comp}', t_c2_m1[2])
                    
                    t_c2_m2 = get_field_val(self.gases_fields['aire_c2_m2'])
                    w(f'AJ{row_comp}', t_c2_m2[0])
                    w(f'AL{row_comp}', t_c2_m2[1])
                    w(f'AM{row_comp}', t_c2_m2[2])
                    
                    t_c2_m3 = get_field_val(self.gases_fields['aire_c2_m3'])
                    w(f'AN{row_comp}', t_c2_m3[0])
                    w(f'AP{row_comp}', t_c2_m3[1])
                    w(f'AQ{row_comp}', t_c2_m3[2])
                    
                    t_c2_m4 = get_field_val(self.gases_fields['aire_c2_m4'])
                    w(f'AR{row_comp}', t_c2_m4[0])
                    w(f'AT{row_comp}', t_c2_m4[1])
                    w(f'AU{row_comp}', t_c2_m4[2])

                    # --- 4. Secadores (Solo Mañana y Tarde) ---
                    if row_sec:
                        w(f'D{row_sec}', get_field_val(self.gases_fields['aire_s1_p']))
                        w(f'I{row_sec}', get_field_val(self.gases_fields['aire_s1_t']))
                        w(f'N{row_sec}', get_field_val(self.gases_fields['aire_s1_p_evap']))
                        w(f'T{row_sec}', get_field_val(self.gases_fields['aire_s1_t_evap']))
                        
                        w(f'Z{row_sec}', get_field_val(self.gases_fields['aire_s2_p']))
                        w(f'AE{row_sec}', get_field_val(self.gases_fields['aire_s2_t']))
                        w(f'AJ{row_sec}', get_field_val(self.gases_fields['aire_s2_p_evap']))
                        w(f'AP{row_sec}', get_field_val(self.gases_fields['aire_s2_t_evap']))

                    # --- 5. Observaciones Generales ---
                    w('F46', self.txt_obs_gases.get("1.0", "end-1c").strip())
                else:
                    escribir_en_celda_segura(ws, 'L4', fecha_str, centrar=True)
                    escribir_en_celda_segura(ws, 'M5', e_resp.get(), centrar=True)
                    
                    celdas_turno = {
                        "Mañana": {"comp": "B13", "flujo": "B25", "temp": "P37", "flow": "P40"},
                        "Tarde": {"comp": "L13", "flujo": "K25", "temp": "U37", "flow": "U40"},
                        "Noche": {"comp": "V13", "flujo": "T25", "temp": "Z37", "flow": "Z40"}
                    }
                    
                    c_turno = celdas_turno[turno_sel]
                    
                    def w(celda, val):
                        escribir_en_celda_segura(ws, celda, val, centrar=True)

                    # Escribir los valores del turno
                    w(c_turno["comp"], get_field_val(self.resonador_fields['res_comp_estado']))
                    w(c_turno["flujo"], get_field_val(self.resonador_fields['res_flujo_estado']))
                    w(c_turno["temp"], get_field_val(self.resonador_fields['res_temp_valor']))
                    w(c_turno["flow"], get_field_val(self.resonador_fields['res_flow_valor']))



                    # Escribir las observaciones diarias (compartidas)
                    w("AF13", get_field_val(self.resonador_fields['res_comp_obs']))
                    w("AC25", get_field_val(self.resonador_fields['res_flujo_obs']))
                    w("AE37", get_field_val(self.resonador_fields['res_temp_obs']))
                    w("AE40", get_field_val(self.resonador_fields['res_flow_obs']))

                wb.save(ruta_salida)

                conn = obtener_conexion()
                cur = conn.cursor()
                import json
                if tipo_prot == "Gases Medicinales":
                    datos_gases = {}
                    for key, field in self.gases_fields.items():
                        datos_gases[key] = get_field_val(field)
                    datos_gases['observaciones_gases'] = self.txt_obs_gases.get("1.0", "end-1c").strip()
                    datos_json = json.dumps(datos_gases)
                else:
                    datos_res = {}
                    for key, field in self.resonador_fields.items():
                        datos_res[key] = get_field_val(field)
                    datos_json = json.dumps(datos_res)

                cur.execute("""
                    INSERT INTO protocolos (fecha, tipo_protocolo, turno, responsable, datos, ruta_excel)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (fecha, tipo_protocolo, turno) DO UPDATE
                    SET responsable = EXCLUDED.responsable, datos = EXCLUDED.datos, ruta_excel = EXCLUDED.ruta_excel;
                """, (fecha_str, tipo_prot, turno_sel, e_resp.get(), datos_json, os.path.abspath(ruta_salida)))
                conn.commit()
                cur.close(); conn.close()

                self.app.cargar_datos_memoria()
                self.app.actualizar_boton_alertas()
                self.refrescar_datos()
                messagebox.showinfo("Éxito", f"Turno {turno_sel} registrado y Excel actualizado.")
                v.destroy()
            except PermissionError:
                if conn:
                    conn.rollback(); conn.close()
                messagebox.showerror("Error de Permisos", "El archivo Excel está abierto o bloqueado por otra aplicación (como Microsoft Excel). Por favor, cierra Excel antes de continuar.")
            except Exception as e:
                if conn: 
                    conn.rollback(); conn.close()
                import traceback
                traceback.print_exc()
                messagebox.showerror("Error", f"Ocurrió un error al guardar:\n{str(e)}")

        ctk.CTkButton(sf, text="Guardar Turno y Actualizar Excel", font=ctk.CTkFont(weight="bold", size=14), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, height=45, command=guardar).pack(pady=20, padx=10, fill="x")

    def abrir_excel_protocolo(self):
        sel = self.tabla_prot.focus()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione un registro.")
            return
        val = self.tabla_prot.item(sel, "values")
        fecha, tipo = val[0], val[1]
        nombre_base = "Protocolo_Gases_" if tipo == "Gases Medicinales" else "Protocolo_Resonador_"
        ruta = os.path.abspath(os.path.join(CARPETAS["protocolos"], f"{nombre_base}{fecha.replace('-','')}.xlsx"))
        if os.path.exists(ruta): os.startfile(ruta)
        else: messagebox.showerror("Error", "No se encontró el archivo Excel.")

    def exportar_pdf_protocolo(self):
        sel = self.tabla_prot.focus()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione un registro.")
            return
        val = self.tabla_prot.item(sel, "values")
        fecha, tipo = val[0], val[1]
        nombre_base = "Protocolo_Gases_" if tipo == "Gases Medicinales" else "Protocolo_Resonador_"
        ruta_excel = os.path.abspath(os.path.join(CARPETAS["protocolos"], f"{nombre_base}{fecha.replace('-','')}.xlsx"))
        if not os.path.exists(ruta_excel):
            messagebox.showerror("Error", "No hay Excel para exportar.")
            return
        ruta_pdf = filedialog.asksaveasfilename(initialdir=CARPETAS["protocolos"], initialfile=f"{nombre_base}{fecha}.pdf", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if not ruta_pdf: return
        messagebox.showinfo("Exportando", "Por favor espera, generando PDF...")
        rango = "$A$1:$AV$63"
        exito = exportar_excel_a_pdf(ruta_excel, ruta_pdf, rango_impresion=rango)
        if exito: os.startfile(os.path.abspath(ruta_pdf))
        else: messagebox.showerror("Error", "Fallo al generar el documento PDF.")