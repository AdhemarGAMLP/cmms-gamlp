# main.py
import os
import sys
import io
import base64

# Asegurar que el directorio de trabajo sea siempre el de main.py
dir_actual = os.path.dirname(os.path.abspath(__file__))
os.chdir(dir_actual)
if dir_actual not in sys.path:
    sys.path.insert(0, dir_actual)

import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import psycopg2.extras
import socket
import qrcode
import json
from PIL import Image
import fitz  # PyMuPDF
import pythoncom
import win32com.client
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import threading

# Importar constantes de diseño y backend centralizado
from estilos import *
from config import CONFIG, CARPETAS
from database import (
    inicializar_bd, 
    obtener_conexion, 
    calcular_proximos_mantenimientos,
    guardar_cache_local_datos,
    cargar_cache_local_datos,
    guardar_mantenimiento_offline_cola,
    sincronizar_mantenimientos_offline_cola,
    obtener_firma_datos_db,
    ejecutar_en_segundo_plano,
    comprimir_imagen_base64,
    cargar_imagen_pil
)

from auth import inicializar_usuarios, login
from web_server import arrancar_hilo_web
from excel_utils import obtener_ruta_plantilla, escribir_en_celda_segura, exportar_excel_a_pdf


# Importar las vistas modulares del subpaquete vistas
from vistas.inventario import VistaInventario
from vistas.catalogo import VistaCatalogo
from vistas.repuestos import VistaRepuestos
from vistas.cronograma import VistaCronograma
from vistas.historial import VistaHistorial
from vistas.protocolos import VistaProtocolos
from vistas.areas import VistaAreas
from vistas.analisis import VistaAnalisis
from vistas.respaldos import VistaRespaldos
from vistas.usuarios import VistaUsuarios

# ========================================================
# VERSIÓN DEL SISTEMA
# ========================================================
VERSION_APP = "v1.0"


# ========================================================
# SELECTOR TERRITORIAL Y DE SEDE (PREVIO AL ACCESO)
# ========================================================
class VentanaSelectorSede(ctk.CTkToplevel):
    def __init__(self, parent, usuario, on_confirmar_callback):
        super().__init__(parent)
        self.parent = parent
        self.usuario = usuario
        self.on_confirmar_callback = on_confirmar_callback
        
        self.title("SGEM GAMLP - Selección de Sede Territorial")
        self.geometry("540x660")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        # Centrar ventana
        self.update_idletasks()
        w = 540
        h = 660
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        
        # Cargar datos jerárquicos de la base de datos
        from database import obtener_jerarquia_sedes_db
        self.sedes_data = obtener_jerarquia_sedes_db()

        self.construir_ui()

    def construir_ui(self):
        f_top = ctk.CTkFrame(self, fg_color="transparent")
        f_top.pack(pady=(20, 10), padx=30, fill="x")
        
        ctk.CTkLabel(f_top, text="🏥 Selector de Centro y Red de Salud", font=ctk.CTkFont(size=20, weight="bold"), text_color=C_BLUE).pack()
        ctk.CTkLabel(f_top, text="Selecciona la ubicación territorial para filtrar el inventario\no accede de forma general a todo el municipio:", font=ctk.CTkFont(size=11), text_color=C_SUBTEXT).pack(pady=(4, 0))

        # Tarjeta de Controles en Cascada
        card = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        card.pack(padx=30, pady=10, fill="both", expand=True)

        # 1. DEPARTAMENTO
        ctk.CTkLabel(card, text="🗺️ 1. Departamento:", font=ctk.CTkFont(size=12, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=25, pady=(15, 2))
        deptos_nombres = [d["nombre"] for d in self.sedes_data.get("departamentos", [])]
        if not deptos_nombres:
            deptos_nombres = ["La Paz"]
        self.combo_depto = ctk.CTkComboBox(card, values=deptos_nombres, command=self.on_depto_cambiado, height=38, corner_radius=10, border_color=C_BORDER, fg_color=C_BG)
        self.combo_depto.pack(padx=25, fill="x", pady=(0, 10))
        if "La Paz" in deptos_nombres:
            self.combo_depto.set("La Paz")

        # 2. MUNICIPIO
        ctk.CTkLabel(card, text="🏛️ 2. Municipio:", font=ctk.CTkFont(size=12, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=25, pady=(0, 2))
        self.combo_mun = ctk.CTkComboBox(card, values=["GAMLP"], command=self.on_mun_cambiado, height=38, corner_radius=10, border_color=C_BORDER, fg_color=C_BG)
        self.combo_mun.pack(padx=25, fill="x", pady=(0, 10))

        # 3. RED DE SALUD
        ctk.CTkLabel(card, text="🌐 3. Red de Salud:", font=ctk.CTkFont(size=12, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=25, pady=(0, 2))
        self.combo_red = ctk.CTkComboBox(card, values=["[ Todas las Redes (Acceso General GAMLP) ]"], command=self.on_red_cambiada, height=38, corner_radius=10, border_color=C_BORDER, fg_color=C_BG)
        self.combo_red.pack(padx=25, fill="x", pady=(0, 10))

        # 4. CENTRO DE SALUD
        ctk.CTkLabel(card, text="🏥 4. Centro de Salud / Hospital:", font=ctk.CTkFont(size=12, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=25, pady=(0, 2))
        self.combo_centro = ctk.CTkComboBox(card, values=["[ Todos los Centros de GAMLP ]"], command=lambda e: self.actualizar_resumen(), height=38, corner_radius=10, border_color=C_BORDER, fg_color=C_BG)
        self.combo_centro.pack(padx=25, fill="x", pady=(0, 12))

        # Badge Informativo de Selección
        self.f_resumen = ctk.CTkFrame(card, fg_color="#F1F5F9", corner_radius=10, border_width=1, border_color="#CBD5E1")
        self.f_resumen.pack(padx=25, fill="x", pady=(0, 15))
        self.lbl_resumen = ctk.CTkLabel(self.f_resumen, text="📍 Vista: Acceso General", font=ctk.CTkFont(size=12, weight="bold"), text_color=C_BLUE)
        self.lbl_resumen.pack(pady=8, padx=12)

        # Botón de Acceso
        btn_ingresar = ctk.CTkButton(self, text="Ingresar al Sistema ➔", font=ctk.CTkFont(size=14, weight="bold"), height=46, corner_radius=12, fg_color=C_BLUE, hover_color=C_BLUE_HOVER, command=self.confirmar_seleccion)
        btn_ingresar.pack(padx=30, pady=(5, 20), fill="x")

        # Inicializar cascada con La Paz
        self.on_depto_cambiado(self.combo_depto.get())

    def on_depto_cambiado(self, depto_sel):
        depto_obj = next((d for d in self.sedes_data.get("departamentos", []) if d["nombre"] == depto_sel), None)
        depto_id = depto_obj["id"] if depto_obj else None

        if depto_sel == "La Paz":
            muns = ["GAMLP"]
        else:
            muns = [f"Capital {depto_sel}"]
            
        self.combo_mun.configure(values=muns)
        self.combo_mun.set(muns[0])
        self.on_mun_cambiado(self.combo_mun.get())

    def on_mun_cambiado(self, mun_sel):
        redes = [r["nombre"] for r in self.sedes_data.get("redes", [])]
        if not redes:
            redes = [
                "RED 1-SUR OESTE (MACRODISTRITO COTAHUMA)",
                "RED 2-NOR OESTE (MACRODISTRITO MAX PAREDES)",
                "RED 3-NORTE CENTRAL (MACRODISTRITO PERIFERICA CENTRAL)",
                "RED 4-SAN ANTONIO (MACRODISTRITO SAN ANTONIO)",
                "RED 5-SUR (MACRODISTRITO SUR)"
            ]
        redes_con_todas = ["[ Todas las Redes (Acceso General GAMLP) ]"] + redes

        self.combo_red.configure(values=redes_con_todas)
        self.combo_red.set(redes_con_todas[0])
        self.on_red_cambiada(self.combo_red.get())

    def on_red_cambiada(self, red_sel):
        if red_sel.startswith("[ Todas"):
            todos_los_centros = sorted(list(set(c["nombre"] for c in self.sedes_data.get("centros", []))))
            centros_vals = ["[ Todos los Centros de GAMLP ]"] + todos_los_centros
            self.combo_centro.configure(values=centros_vals)
            self.combo_centro.set("[ Todos los Centros de GAMLP ]")
        else:
            red_obj = next((r for r in self.sedes_data.get("redes", []) if r["nombre"] == red_sel), None)
            red_id = red_obj["id"] if red_obj else None
            centros = [c["nombre"] for c in self.sedes_data.get("centros", []) if c.get("red_salud_id") == red_id]
            if not centros:
                centros = ["CENTRO DE SALUD CENTRAL"]
            centros_con_todos = ["[ Todos los Centros de la Red ]"] + sorted(centros)
            self.combo_centro.configure(values=centros_con_todos)
            self.combo_centro.set(centros_con_todos[0])
            
        self.actualizar_resumen()

    def actualizar_resumen(self):
        dep = self.combo_depto.get()
        mun = self.combo_mun.get()
        red = self.combo_red.get()
        cen = self.combo_centro.get()

        if cen.startswith("[ Todos los Centros de GAMLP"):
            res = f"🌐 GAMLP • Acceso General (Todas las Redes)"
        elif cen.startswith("[ Todos los Centros de la Red"):
            res = f"🌐 {red} (Todos los Centros)"
        else:
            # Si se seleccionó un centro específico en modo Todas las Redes, identificar su red
            cen_obj = next((c for c in self.sedes_data.get("centros", []) if c["nombre"] == cen), None)
            if cen_obj and red.startswith("[ Todas"):
                red_padre = next((r for r in self.sedes_data.get("redes", []) if r["id"] == cen_obj.get("red_salud_id")), None)
                red_txt = red_padre["nombre"] if red_padre else "GAMLP"
                res = f"📍 {red_txt} • {cen}"
            else:
                res = f"📍 {red} • {cen}"

        self.lbl_resumen.configure(text=res)

    def confirmar_seleccion(self):
        dep = self.combo_depto.get()
        mun = self.combo_mun.get()
        red = self.combo_red.get()
        cen = self.combo_centro.get()

        cen_obj = next((c for c in self.sedes_data.get("centros", []) if c["nombre"] == cen), None)
        
        # Si eligió un centro específico pero la red estaba en [ Todas las Redes ], resolver su red real
        if cen_obj and red.startswith("[ Todas"):
            red_obj = next((r for r in self.sedes_data.get("redes", []) if r["id"] == cen_obj.get("red_salud_id")), None)
            red_nombre = red_obj["nombre"] if red_obj else red
        else:
            red_obj = next((r for r in self.sedes_data.get("redes", []) if r["nombre"] == red), None)
            red_nombre = red_obj["nombre"] if red_obj else red

        depto_obj = next((d for d in self.sedes_data.get("departamentos", []) if d["nombre"] == dep), None)
        mun_obj = next((m for m in self.sedes_data.get("municipios", []) if m["nombre"] == mun), None)

        contexto = {
            "departamento": dep,
            "departamento_id": depto_obj["id"] if depto_obj else None,
            "municipio": mun,
            "municipio_id": mun_obj["id"] if mun_obj else None,
            "red_salud": red_nombre,
            "red_salud_id": red_obj["id"] if red_obj else None,
            "centro_salud": cen,
            "centro_salud_id": cen_obj["id"] if cen_obj else None,
            "es_global": cen.startswith("[ Todos"),
            "resumen_texto": self.lbl_resumen.cget("text")
        }

        self.destroy()
        if self.on_confirmar_callback:
            self.on_confirmar_callback(contexto)


# ========================================================
# INTERFAZ DE LOGIN
# ========================================================
class VentanaLogin(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"SGEM GAMLP {VERSION_APP} - Iniciar Sesión")
        self.geometry("400x420")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        self.usuario_autenticado = None
        self.contexto_sede = None

        ctk.CTkLabel(self, text="🏛️ SGEM GAMLP", font=ctk.CTkFont(size=24, weight="bold"), text_color=C_BLUE).pack(pady=(25, 2))
        ctk.CTkLabel(self, text=f"Sistema de Gestión de Equipamiento Médico ({VERSION_APP})", font=ctk.CTkFont(size=12), text_color=C_SUBTEXT).pack(pady=(0, 12))
        
        marco = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        marco.pack(padx=30, pady=5, fill="both", expand=True)

        ctk.CTkLabel(marco, text="Usuario:", font=ctk.CTkFont(size=12, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=25, pady=(15, 2))
        self.e_user = ctk.CTkEntry(marco, placeholder_text="Ingrese su usuario", width=300, height=38, corner_radius=10, border_color=C_BORDER, fg_color=C_BG)
        self.e_user.pack(padx=25, pady=(0, 8))
        
        ctk.CTkLabel(marco, text="Contraseña:", font=ctk.CTkFont(size=12, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=25, pady=(0, 2))
        self.e_pass = ctk.CTkEntry(marco, placeholder_text="••••••••", show="*", width=300, height=38, corner_radius=10, border_color=C_BORDER, fg_color=C_BG)
        self.e_pass.pack(padx=25, pady=(0, 12))
        self.e_pass.bind("<Return>", lambda e: self.intentar_login())
        
        ctk.CTkButton(marco, text="Ingresar al Sistema", command=self.intentar_login, height=40, corner_radius=10, font=ctk.CTkFont(weight="bold", size=13), fg_color=C_BLUE, hover_color=C_BLUE_HOVER).pack(padx=25, pady=(5, 8), fill="x")
        
        ctk.CTkButton(marco, text="⚙️ Configurar Servidor (IP / Nube)", command=self.abrir_config_servidor, height=28, fg_color="transparent", text_color=C_SUBTEXT, hover_color=C_CARD_HOVER, font=ctk.CTkFont(size=11)).pack(pady=(0, 10))


    def intentar_login(self):
        u = login(self.e_user.get(), self.e_pass.get())
        if u:
            self.usuario_autenticado = u
            self.withdraw()
            
            def al_confirmar_sede(contexto):
                self.contexto_sede = contexto
                self.destroy()
                
            v_sel = VentanaSelectorSede(self, u, al_confirmar_sede)
            v_sel.protocol("WM_DELETE_WINDOW", lambda: (v_sel.destroy(), self.deiconify()))
        else:
            messagebox.showerror("Acceso Denegado", "Usuario o contraseña incorrectos.\n\n(Verifique las credenciales o la conexión al servidor en ⚙️ Configurar Servidor)")

    def abrir_config_servidor(self):
        from config import CONFIG, guardar_config
        import psycopg2
        
        v_cfg = ctk.CTkToplevel(self)
        v_cfg.title("Configuración de Servidor de Base de Datos")
        v_cfg.geometry("420x420")
        v_cfg.transient(self)
        v_cfg.grab_set()
        v_cfg.configure(fg_color=C_BG)
        
        ctk.CTkLabel(v_cfg, text="Conexión con el Servidor Central", font=ctk.CTkFont(size=16, weight="bold"), text_color=C_TEXT).pack(pady=(20, 5))
        ctk.CTkLabel(v_cfg, text="Si esta es una PC cliente, ingresa la IP del Servidor Central:", font=ctk.CTkFont(size=11), text_color=C_SUBTEXT).pack(pady=(0, 15), padx=20)
        
        f_campos = ctk.CTkFrame(v_cfg, fg_color=C_CARD, corner_radius=10)
        f_campos.pack(padx=25, fill="both", expand=True)
        
        ctk.CTkLabel(f_campos, text="IP / Host del Servidor:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w", padx=20, pady=(15, 2))
        e_host = ctk.CTkEntry(f_campos, width=320)
        e_host.insert(0, str(CONFIG.get("db_host", "localhost")))
        e_host.pack(padx=20, pady=(0, 10))
        
        ctk.CTkLabel(f_campos, text="Puerto:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w", padx=20, pady=(0, 2))
        e_port = ctk.CTkEntry(f_campos, width=320)
        e_port.insert(0, str(CONFIG.get("db_port", "5433")))
        e_port.pack(padx=20, pady=(0, 10))
        
        ctk.CTkLabel(f_campos, text="Nombre de Base de Datos:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w", padx=20, pady=(0, 2))
        e_name = ctk.CTkEntry(f_campos, width=320)
        e_name.insert(0, str(CONFIG.get("db_name", "postgres")))
        e_name.pack(padx=20, pady=(0, 15))
        
        lbl_status = ctk.CTkLabel(v_cfg, text="", font=ctk.CTkFont(size=12, weight="bold"))
        lbl_status.pack(pady=5)
        
        def probar_conexion():
            try:
                test_conn = psycopg2.connect(
                    dbname=e_name.get().strip(),
                    user=CONFIG.get("db_user", "postgres"),
                    password=CONFIG.get("db_password", "1234"),
                    host=e_host.get().strip(),
                    port=e_port.get().strip(),
                    connect_timeout=4
                )
                test_conn.close()
                lbl_status.configure(text="✅ Conexión con el Servidor exitosa.", text_color=C_GREEN_HOVER)
                return True
            except Exception as ex:
                lbl_status.configure(text=f"❌ Error al conectar: {ex}", text_color=C_RED_HOVER)
                return False
                
        def guardar():
            nueva_cfg = dict(CONFIG)
            nueva_cfg["db_host"] = e_host.get().strip()
            nueva_cfg["db_port"] = e_port.get().strip()
            nueva_cfg["db_name"] = e_name.get().strip()
            if guardar_config(nueva_cfg):
                messagebox.showinfo("Guardado", "Configuración de conexión actualizada con éxito.")
                v_cfg.destroy()
            else:
                messagebox.showerror("Error", "No se pudo guardar la configuración.")
                
        f_btns = ctk.CTkFrame(v_cfg, fg_color="transparent")
        f_btns.pack(pady=15, padx=25, fill="x")
        
        ctk.CTkButton(f_btns, text="Probar Conexión", command=probar_conexion, fg_color=C_BG, text_color=C_BLUE, hover_color=C_BORDER, height=36, corner_radius=8).pack(side="left", expand=True, padx=5)
        ctk.CTkButton(f_btns, text="Guardar Cambios", command=guardar, fg_color=C_BLUE, hover_color=C_BLUE_HOVER, height=36, corner_radius=8, font=ctk.CTkFont(weight="bold")).pack(side="right", expand=True, padx=5)


# ========================================================
# NÚCLEO PRINCIPAL DEL SOFTWARE
# ========================================================
class SistemaMantenimiento(ctk.CTk):
    def __init__(self, usuario, contexto_sede=None):
        super().__init__()
        self.usuario_actual = usuario
        self.contexto_sede = contexto_sede or {"resumen_texto": "🌐 Acceso General GAMLP", "es_global": True}
        self.es_jefe = usuario.get("rol") == "jefe"
        self.modo_offline = False
        
        self.title(f"SGEM GAMLP {VERSION_APP} - Sistema de Gestión de Equipamiento Médico | GAMLP (Rol: {usuario['rol'].upper()})")
        self.geometry("1300x800")

        self.after(100, lambda: self.state('zoomed'))
        ctk.set_appearance_mode("light")

        self.configure(fg_color=C_BG)

        self.ip_local = self.obtener_ip_local()
        self.alertas_activas = []
        self.alertas_ignoradas = set()
        self._calendario_sucio = True
        
        # Asegurar que todas las carpetas del sistema existan desde el arranque (auto-sanación)
        for folder_path in CARPETAS.values():
            os.makedirs(folder_path, exist_ok=True)
            
        arrancar_hilo_web(self)
        self.cargar_datos_memoria()
        self.configurar_estilo_ttk()
        self.crear_interfaz_base()
        
        self.vistas = {}
        self.crear_vistas_modulares()
        self.datos_sucios = False
        self._ejecutando = True
        self.chequear_datos_sucios()
        self.iniciar_sincronizacion_background()
        self.mostrar_vista("Inventario")
        self.actualizar_boton_alertas()
        self.verificar_y_ejecutar_backup_auto()
        self.protocol("WM_DELETE_WINDOW", self.al_cerrar_aplicacion)
        self.sidebar_mode = None
        self.bind("<Configure>", self.al_redimensionar)

    def obtener_ip_local(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "127.0.0.1"

    def configurar_estilo_ttk(self):
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", 
                        background=C_CARD, 
                        foreground=C_TEXT, 
                        rowheight=38, 
                        fieldbackground=C_CARD, 
                        borderwidth=0, 
                        font=('Segoe UI', 10))
        style.map('Treeview', 
                  background=[('selected', C_BLUE)], 
                  foreground=[('selected', '#FFFFFF')])
        style.configure("Treeview.Heading", 
                        background="#F1F5F9", 
                        foreground="#475569", 
                        font=('Segoe UI', 10, 'bold'), 
                        borderwidth=0, 
                        relief="flat")
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        # Diseño y comportamiento de Scrollbar minimalista
        style.layout("Vertical.TScrollbar", [
            ('Vertical.Scrollbar.trough', {
                'children': [
                    ('Vertical.Scrollbar.thumb', {'expand': '1', 'sticky': 'nswe'})
                ],
                'sticky': 'ns'
            })
        ])
        style.configure("Vertical.TScrollbar", 
                        background="#CBD5E1", 
                        troughcolor=C_CARD, 
                        bordercolor=C_CARD, 
                        thickness=8, 
                        relief="flat")
        style.map("Vertical.TScrollbar", 
                  background=[('pressed', '#64748B'), ('active', '#94A3B8')])

        style.layout("Horizontal.TScrollbar", [
            ('Horizontal.Scrollbar.trough', {
                'children': [
                    ('Horizontal.Scrollbar.thumb', {'expand': '1', 'sticky': 'nswe'})
                ],
                'sticky': 'ew'
            })
        ])
        style.configure("Horizontal.TScrollbar", 
                        background="#CBD5E1", 
                        troughcolor=C_CARD, 
                        bordercolor=C_CARD, 
                        thickness=8, 
                        relief="flat")
        style.map("Horizontal.TScrollbar", 
                  background=[('pressed', '#64748B'), ('active', '#94A3B8')])


    def iniciar_sincronizacion_background(self):
        """Monitorea estado de conectividad a internet y sincronización en la nube en segundo plano."""
        def _hilo_sync():
            import time, socket
            while getattr(self, "_ejecutando", True):
                time.sleep(5)
                # 1. Comprobar conectividad real a internet
                online = False
                try:
                    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    s.settimeout(2.5)
                    s.connect(("8.8.8.8", 53))
                    s.close()
                    online = True
                except Exception:
                    online = False
                
                estado_anterior = getattr(self, "modo_offline", False)
                nuevo_estado_offline = not online

                if estado_anterior != nuevo_estado_offline:
                    self.modo_offline = nuevo_estado_offline
                    if not nuevo_estado_offline:
                        print("[INFO] ¡Reconexión con Internet y Servidor detectada! Sincronizando...")
                        self.after(0, self._aplicar_reconeccion_online)
                    else:
                        print("[WARN] Conexión perdida a Internet. Cambiando a Modo Offline...")
                        self.after(0, self.actualizar_estado_offline_ui)
                elif not nuevo_estado_offline:
                    try:
                        firma_actual = obtener_firma_datos_db()
                        if firma_actual and firma_actual != getattr(self, "ultima_firma_db", None):
                            primera_vez = getattr(self, "ultima_firma_db", None) is None
                            self.ultima_firma_db = firma_actual
                            if not primera_vez:
                                print("[INFO] ¡Cambio detectado en la Base de Datos Central! Sincronizando...")
                                self.after(0, self._aplicar_datos_sincronizados)
                    except Exception:
                        pass

        t = threading.Thread(target=_hilo_sync, daemon=True)
        t.start()

    def _aplicar_reconeccion_online(self):
        self.cargar_datos_memoria()
        self.actualizar_estado_offline_ui()
        for nombre_v, vista_v in self.vistas.items():
            if hasattr(vista_v, 'refrescar_datos'):
                vista_v.refrescar_datos()

    def _aplicar_datos_sincronizados(self):
        self.cargar_datos_memoria()
        vista_activa = getattr(self, "vista_actual_nombre", "Inventario")
        if vista_activa in self.vistas and hasattr(self.vistas[vista_activa], 'refrescar_datos'):
            self.vistas[vista_activa].refrescar_datos()
        if vista_activa == "Cronograma" and "Cronograma" in self.vistas:
            try:
                self.vistas["Cronograma"].dibujar_mes(self.vistas["Cronograma"].anio_actual, self.vistas["Cronograma"].mes_actual)
                self.vistas["Cronograma"].dibujar_anio(self.vistas["Cronograma"].anio_vista)
            except:
                pass
        self.actualizar_boton_alertas()

    def chequear_datos_sucios(self):
        if getattr(self, 'datos_sucios', False):
            self.datos_sucios = False
            self._aplicar_datos_sincronizados()
        self.after(1500, self.chequear_datos_sucios)


    def actualizar_estado_offline_ui(self):
        if hasattr(self, 'lbl_estado_conexion') and self.lbl_estado_conexion.winfo_exists():
            if getattr(self, 'modo_offline', False):
                self.lbl_estado_conexion.configure(text="🔴 Desconectado (Offline)", text_color="#DC2626")
            else:
                self.lbl_estado_conexion.configure(text="🟢 Conectado", text_color="#16A34A")
        if hasattr(self, 'lbl_modo_offline') and self.lbl_modo_offline.winfo_exists():
            if getattr(self, 'modo_offline', False):
                self.lbl_modo_offline.pack(before=self.lbl_estado_conexion, pady=(0, 5), padx=15, fill="x")
            else:
                self.lbl_modo_offline.pack_forget()

    def cargar_datos_memoria(self):
        self.datos = {"catalogo": [], "repuestos": [], "equipos": [], "protocolos": [], "areas": []}
        self.eventos_calendario = {}
        self.hoy = datetime.now().date()
        self.hora_actual = datetime.now().hour
        self.alertas_activas = []

        conn = obtener_conexion()
        if not conn: 
            # Fallback a Caché Local de Lectura
            cache_datos = cargar_cache_local_datos()
            if cache_datos:
                self.datos = cache_datos
                self.modo_offline = True
                print("[INFO] Modo Offline Activo: Datos cargados desde la caché local.")
            else:
                self.modo_offline = True
                return
        else:
            self.modo_offline = False
            # Sincronizar cola de intervenciones guardadas offline
            try:
                sinc, _ = sincronizar_mantenimientos_offline_cola()
                if sinc > 0:
                    print(f"[OK] Sincronizados {sinc} mantenimientos offline con PostgreSQL.")
            except Exception as se:
                print(f"[WARN] Error sincronizando cola: {se}")

            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            cur.execute("SELECT * FROM catalogo ORDER BY nombre ASC")
            self.datos["catalogo"] = [dict(r) for r in cur.fetchall()]
            
            cur.execute("SELECT * FROM repuestos")
            self.datos["repuestos"] = [dict(r) for r in cur.fetchall()]

            cur.execute("SELECT * FROM areas ORDER BY piso DESC, nombre ASC")
            self.datos["areas"] = [dict(r) for r in cur.fetchall()]

            try:
                cur.execute("SELECT * FROM protocolos ORDER BY fecha DESC, turno ASC")
                self.datos["protocolos"] = [dict(r) for r in cur.fetchall()]
            except:
                conn.rollback()

            try:
                cur.execute("SELECT * FROM historial_intervenciones ORDER BY COALESCE(fecha_entrega, fecha) DESC, COALESCE(hora_entrega, '00:00') DESC, id DESC")
                todas_inter = [dict(h) for h in cur.fetchall()]
                hist_por_equipo = {}
                for h in todas_inter:
                    hist_por_equipo.setdefault(h['equipo_id'], []).append(h)

                cur.execute("SELECT * FROM equipos")
                eqs_db = [dict(r) for r in cur.fetchall()]

                for eq in eqs_db:
                    eq['historial_intervenciones'] = hist_por_equipo.get(eq['id'], [])
                    self.datos["equipos"].append(eq)
            except Exception as e:
                print("[WARN] Error cargando equipos/historial:", e)

            cur.close()
            conn.close()

            # Guardar copia fresca en caché local
            guardar_cache_local_datos(self.datos)

        try:
            for eq in self.datos.get("equipos", []):


                
                # Alertas de vencimiento de garantía (1 mes antes)
                f_venc = eq.get("fecha_vencimiento_garantia")
                dias_gar = -9999
                if eq.get("garantia") == "Con Garantía" and f_venc:
                    if isinstance(f_venc, str):
                        try:
                            f_venc = datetime.strptime(f_venc, "%Y-%m-%d").date()
                        except:
                            f_venc = None
                    if f_venc:
                        dias_gar = (f_venc - self.hoy).days
                        if 0 <= dias_gar <= 30:
                            msg = f"🛡️ GARANTÍA POR VENCER: {eq['id']} - {eq['nombre']} (Vence en {dias_gar} días)"
                            if msg not in self.alertas_ignoradas:
                                self.alertas_activas.append(msg)
                                
                crit = str(eq.get("criticidad") or "Riesgo Medio")
                meses = 3 if "Alto" in crit else (4 if "Medio" in crit else 6)
                
                # Calcular la fecha de inicio original f_start
                f_reg = eq.get("fecha_adquisicion") or eq.get("fecha_registro", self.hoy)
                if isinstance(f_reg, datetime): 
                    f_reg = f_reg.date()
                elif isinstance(f_reg, str):
                    try: 
                        f_reg = datetime.strptime(f_reg, "%Y-%m-%d").date()
                    except: 
                        f_reg = eq.get("fecha_registro", self.hoy)
                        if isinstance(f_reg, str):
                            try: f_reg = datetime.strptime(f_reg, "%Y-%m-%d").date()
                            except: f_reg = self.hoy
                        elif isinstance(f_reg, datetime):
                            f_reg = f_reg.date()

                f_venc_g_date = None
                esta_en_garantia = False
                if eq.get("garantia") == "Con Garantía" and eq.get("fecha_vencimiento_garantia"):
                    f_venc_g = eq.get("fecha_vencimiento_garantia")
                    if isinstance(f_venc_g, str):
                        try: f_venc_g_date = datetime.strptime(f_venc_g, "%Y-%m-%d").date()
                        except: f_venc_g_date = None
                    elif isinstance(f_venc_g, datetime):
                        f_venc_g_date = f_venc_g.date()
                    else:
                        f_venc_g_date = f_venc_g
                    
                    if f_venc_g_date and f_venc_g_date >= self.hoy:
                        esta_en_garantia = True

                # 1. Obtener la próxima fecha de mantenimiento no completada
                f_prox = None
                estado = "Al Día"
                if eq.get("estado") == "Baja":
                    estado = "Dado de Baja"
                    f_prox = None
                else:
                    proximos = calcular_proximos_mantenimientos(eq, cantidad=1, hoy=self.hoy)
                    if proximos:
                        f_prox = proximos[0]
                        dias_restantes = (f_prox - self.hoy).days
                        limit_date = date(f_prox.year, f_prox.month, 1) + relativedelta(months=+1, day=5)
                        estado = "Vencido" if self.hoy > limit_date else ("Por Vencer" if dias_restantes <= 30 else "Al Día")
                    else:
                        f_prox = None
                        estado = "Al Día"


                eq['f_prox'] = f_prox
                if estado == "Por Vencer" and not esta_en_garantia and eq.get("estado") != "Baja":
                    msg = f"⏳ Próximo a vencer: {eq['id']} - {eq['nombre']} ({dias_restantes} días)"
                    if msg not in self.alertas_ignoradas:
                        self.alertas_activas.append(msg)

                # 2. Generar y evaluar todos los slots para eventos_calendario (años 2026 y 2027)
                if eq.get("estado") != "Baja":
                    f_iter = f_reg
                    while True:
                        f_iter = f_iter + relativedelta(months=+meses)
                        if f_iter.year > 2027:
                            break
                        
                        # Si está en período de garantía, no se muestran mantenimientos internos en el calendario
                        if f_venc_g_date and f_iter <= f_venc_g_date:
                            continue
                    
                        slot_is_completed = False
                        completed_m = None
                        for m in eq.get("historial_intervenciones", []):
                            if m["tipo"] == "Preventivo":
                                m_prog = m.get("fecha_programada")
                                if m_prog:
                                    if isinstance(m_prog, str):
                                        try: m_prog_d = datetime.strptime(m_prog, "%Y-%m-%d").date()
                                        except: m_prog_d = None
                                    else:
                                        m_prog_d = m_prog
                                    if m_prog_d == f_iter:
                                        slot_is_completed = True
                                        completed_m = m
                                        break
                                else:
                                    m_f = m.get("fecha")
                                    if isinstance(m_f, str):
                                        try: m_f_d = datetime.strptime(m_f, "%Y-%m-%d").date()
                                        except: m_f_d = None
                                    else:
                                        m_f_d = m_f
                                    if m_f_d and m_f_d.year == f_iter.year and m_f_d.month == f_iter.month:
                                        slot_is_completed = True
                                        completed_m = m
                                        break
                        
                        if slot_is_completed:
                            f_real = completed_m.get("fecha") if completed_m else None
                            if isinstance(f_real, str):
                                try: f_real = datetime.strptime(f_real, "%Y-%m-%d").date()
                                except: f_real = None
                            elif isinstance(f_real, datetime):
                                f_real = f_real.date()
                            
                            lim_date = date(f_iter.year, f_iter.month, 1) + relativedelta(months=+1, day=5)
                            if f_real and f_real <= lim_date:
                                est_slot = "Realizado a Tiempo"
                            else:
                                est_slot = "Realizado Fuera de Fecha"
                        elif eq.get("estado") == "Baja":
                            est_slot = "Dado de Baja"
                        else:
                            lim_date = date(f_iter.year, f_iter.month, 1) + relativedelta(months=+1, day=5)
                            if self.hoy > lim_date:
                                est_slot = "Vencido"
                            elif f_iter.year == self.hoy.year and f_iter.month == self.hoy.month:
                                est_slot = "Pendiente Este Mes"
                            else:
                                est_slot = "Futuro"
                            
                        self.eventos_calendario.setdefault(f_iter, []).append({'eq': eq['nombre'], 'estado': est_slot, 'id': eq['id'], 'f_prox': f_iter})
        except Exception as e:
            print("Error parsing equipos/alertas:", e)

        # Alertas de protocolos de Gases y Resonador desactivadas a solicitud
        # prot_hoy = [p for p in self.datos["protocolos"] if p['fecha'] == self.hoy]
        # tipos = ['Gases Medicinales', 'Resonador Magnético']
        # turnos_req = []
        # if self.hora_actual >= 8: turnos_req.append('Mañana')
        # if self.hora_actual >= 14: turnos_req.append('Tarde')
        # if self.hora_actual >= 23: turnos_req.append('Noche')
        # for t_req in turnos_req:
        #     for tipo in tipos:
        #         if not any(p['tipo_protocolo'] == tipo and p['turno'] == t_req for p in prot_hoy):
        #             msg = f"🚨 FALTA PROTOCOLO: {tipo} (Turno: {t_req})"
        #             if msg not in self.alertas_ignoradas:
        #                 self.alertas_activas.append(msg)

        for area in self.datos["areas"]:
            area_name = area.get("nombre", "General")
            # Sanitizar el nombre del área para usarlo como carpeta de forma segura
            area_folder = "".join([c for c in area_name if c.isalnum() or c==' ']).strip()
            if area_folder:
                os.makedirs(os.path.join(CARPETAS["areas"], area_folder, "equipos"), exist_ok=True)
                os.makedirs(os.path.join(CARPETAS["areas"], area_folder, "mantenimientos"), exist_ok=True)

        # Actualizar botón de alertas en el sidebar si ya fue creado
        if hasattr(self, 'btn_alertas'):
            self.actualizar_boton_alertas()

    def crear_interfaz_base(self):
        self.sidebar = ctk.CTkFrame(self, width=240, corner_radius=0, fg_color=C_CARD, border_width=1, border_color=C_BORDER)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        
        # 1. Área Superior Fija
        self.top_sidebar = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.top_sidebar.pack(side="top", fill="x", pady=(10, 2))
        
        self.lbl_logo = ctk.CTkLabel(self.top_sidebar, text="SGEM GAMLP", font=ctk.CTkFont(size=20, weight="bold"), text_color=C_TEXT)
        self.lbl_logo.pack(pady=(12, 4), padx=15)

        # Indicador de Modo Offline / Lectura
        self.lbl_modo_offline = ctk.CTkLabel(
            self.top_sidebar, text="⚠️ MODO SIN CONEXIÓN\n(Solo Lectura)", 
            fg_color="#C0392B", text_color="white", corner_radius=6, 
            font=ctk.CTkFont(size=10, weight="bold"), height=26
        )
        if getattr(self, "modo_offline", False):
            self.lbl_modo_offline.pack(pady=(0, 6), padx=12, fill="x")

        # Indicador de Estado de Conexión en Tiempo Real
        self.lbl_estado_conexion = ctk.CTkLabel(
            self.top_sidebar, 
            text="🔴 Desconectado (Offline)" if getattr(self, "modo_offline", False) else "🟢 Conectado", 
            font=ctk.CTkFont(size=10, weight="bold"), 
            text_color="#DC2626" if getattr(self, "modo_offline", False) else "#16A34A"
        )
        self.lbl_estado_conexion.pack(pady=(0, 1))
        
        self.lbl_name = ctk.CTkLabel(self.top_sidebar, text="Rudel Adhemar Santos Medina", font=ctk.CTkFont(size=10, slant="italic"), text_color=C_SUBTEXT)
        self.lbl_name.pack(pady=(0, 4))

        # Badge de Sede / Centro de Salud Activo
        self.f_sede_badge = ctk.CTkFrame(self.top_sidebar, fg_color="#F8FAFC", corner_radius=8, border_width=1, border_color="#CBD5E1")
        self.f_sede_badge.pack(pady=(0, 6), padx=10, fill="x")
        
        sede_txt = self.contexto_sede.get("resumen_texto", "🌐 Acceso General GAMLP") if hasattr(self, "contexto_sede") and self.contexto_sede else "🌐 Acceso General GAMLP"
        self.lbl_sede_badge = ctk.CTkLabel(self.f_sede_badge, text=sede_txt, font=ctk.CTkFont(size=10, weight="bold"), text_color=C_BLUE, wraplength=200)
        self.lbl_sede_badge.pack(pady=(4, 2), padx=6)
        
        self.btn_cambiar_sede = ctk.CTkButton(self.f_sede_badge, text="🔄 Cambiar Sede", font=ctk.CTkFont(size=10, weight="bold"), height=22, fg_color=C_CARD, text_color=C_TEXT, hover_color=C_BORDER, corner_radius=6, command=self.abrir_selector_sede_rapido)
        self.btn_cambiar_sede.pack(pady=(0, 4), padx=6, fill="x")

        self.btn_alertas = ctk.CTkButton(self.top_sidebar, text="🔔 Alertas (0)", height=34, font=ctk.CTkFont(weight="bold", size=12), fg_color=C_BG, text_color=C_TEXT, command=self.mostrar_ventana_alertas)
        self.btn_alertas.pack(pady=(0, 6), padx=12, fill="x")

        # 2. Área Inferior Fija
        self.bottom_sidebar = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.bottom_sidebar.pack(side="bottom", fill="x", pady=(5, 10))
        
        self.btn_bottom_mantenimiento = ctk.CTkButton(self.bottom_sidebar, text="✚ Mantenimiento", height=38, corner_radius=8, font=ctk.CTkFont(size=13, weight="bold"), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, command=self.modulo_mantenimiento)
        self.btn_bottom_mantenimiento.pack(pady=2, padx=12, fill="x")

        # 3. Área Central (Scrollable Frame para que quepan todos los botones en cualquier pantalla)
        self.scroll_sidebar = ctk.CTkScrollableFrame(self.sidebar, fg_color="transparent", scrollbar_button_color=C_CARD, scrollbar_button_hover_color=C_BORDER)
        self.scroll_sidebar.pack(side="top", fill="both", expand=True, padx=2, pady=2)

        btn_estilo = {"fg_color": "transparent", "text_color": C_TEXT, "hover_color": C_BG, "anchor": "center", "height": 34, "font": ctk.CTkFont(size=13, weight="bold")}
        
        self.btn_nav_inv = ctk.CTkButton(self.scroll_sidebar, text="📦 Inventario", command=lambda: self.mostrar_vista("Inventario"), **btn_estilo)
        self.btn_nav_inv.pack(pady=1, padx=8, fill="x")
        
        self.btn_nav_cat = ctk.CTkButton(self.scroll_sidebar, text="🩺 Equipos Médicos", command=lambda: self.mostrar_vista("Catalogo"), **btn_estilo)
        self.btn_nav_cat.pack(pady=1, padx=8, fill="x")
        
        self.btn_nav_rep = ctk.CTkButton(self.scroll_sidebar, text="🔧 Repuestos", command=lambda: self.mostrar_vista("Repuestos"), **btn_estilo)
        self.btn_nav_rep.pack(pady=1, padx=8, fill="x")
        
        self.btn_nav_cro = ctk.CTkButton(self.scroll_sidebar, text="📅 Cronograma", command=lambda: self.mostrar_vista("Cronograma"), **btn_estilo)
        self.btn_nav_cro.pack(pady=1, padx=8, fill="x")
        
        self.btn_nav_hist = ctk.CTkButton(self.scroll_sidebar, text="📋 Mantenimientos", command=lambda: self.mostrar_vista("Historial"), **btn_estilo)
        self.btn_nav_hist.pack(pady=1, padx=8, fill="x")
        
        self.btn_nav_analisis = ctk.CTkButton(self.scroll_sidebar, text="📊 Análisis", command=lambda: self.mostrar_vista("Analisis"), **btn_estilo)
        self.btn_nav_analisis.pack(pady=1, padx=8, fill="x")
        
        # Módulo de Protocolos: Oculto del menú operativo pero conservado
        self.btn_nav_prot = ctk.CTkButton(self.scroll_sidebar, text="📝 Protocolos", command=lambda: self.mostrar_vista("Protocolos"), **btn_estilo)
        # self.btn_nav_prot.pack(pady=1, padx=8, fill="x")
        
        self.btn_nav_areas = ctk.CTkButton(self.scroll_sidebar, text="📍 Áreas", command=lambda: self.mostrar_vista("Areas"), **btn_estilo)
        self.btn_nav_areas.pack(pady=1, padx=8, fill="x")
        
        self.btn_nav_respaldos = ctk.CTkButton(self.scroll_sidebar, text="💾 Respaldos", command=lambda: self.mostrar_vista("Respaldos"), **btn_estilo)
        self.btn_nav_respaldos.pack(pady=1, padx=8, fill="x")

        self.btn_nav_usuarios = None
        if self.es_jefe:
            self.btn_nav_usuarios = ctk.CTkButton(self.scroll_sidebar, text="👥 Usuarios", command=lambda: self.mostrar_vista("Usuarios"), **btn_estilo)
            self.btn_nav_usuarios.pack(pady=1, padx=8, fill="x")

        self.botones_nav = [
            self.btn_nav_inv, self.btn_nav_cat, self.btn_nav_rep, self.btn_nav_cro, 
            self.btn_nav_hist, self.btn_nav_analisis, self.btn_nav_areas, 
            self.btn_nav_respaldos
        ]
        if self.btn_nav_usuarios:
            self.botones_nav.append(self.btn_nav_usuarios)

        self.contenedor_principal = ctk.CTkFrame(self, fg_color=C_BG)
        self.contenedor_principal.pack(side="right", fill="both", expand=True)

    def abrir_selector_sede_rapido(self):
        def al_cambiar_sede(nuevo_contexto):
            self.contexto_sede = nuevo_contexto
            self.lbl_sede_badge.configure(text=nuevo_contexto.get("resumen_texto", "🌐 Acceso General GAMLP"))
            # Refrescar todas las vistas modulares
            for nombre_v, v in self.vistas.items():
                if hasattr(v, "refrescar_datos"):
                    v.refrescar_datos()
            messagebox.showinfo("Sede Actualizada", f"Sede activa cambiada a:\n{nuevo_contexto.get('resumen_texto')}")

        VentanaSelectorSede(self, self.usuario_actual, al_cambiar_sede)

    def equipo_cumple_sede_activa(self, eq):
        """Verifica si un equipo pertenece a la sede territorial activa seleccionada."""
        contexto = getattr(self, "contexto_sede", None)
        if not contexto or contexto.get("es_global", True):
            return True
        cen_id = contexto.get("centro_salud_id")
        cen_nom = contexto.get("centro_salud")
        red_id = contexto.get("red_salud_id")
        red_nom = contexto.get("red_salud")

        if cen_id or (cen_nom and not str(cen_nom).startswith("[ Todos")):
            if cen_id and eq.get("centro_salud_id") == cen_id:
                return True
            if cen_nom:
                cen_clean = str(cen_nom).strip().lower()
                if str(eq.get("centro_salud_nombre", "")).strip().lower() == cen_clean:
                    return True
                if str(eq.get("servicio", "")).strip().lower() == cen_clean:
                    return True
            return False
        elif red_id or (red_nom and not str(red_nom).startswith("[ Todas")):
            if red_id and eq.get("red_salud_id") == red_id:
                return True
            if red_nom:
                red_clean = str(red_nom).strip().lower()
                if str(eq.get("red_salud_nombre", "")).strip().lower() == red_clean:
                    return True
            return False
        return True


    def crear_vistas_modulares(self):
        self.vistas["Inventario"] = VistaInventario(self.contenedor_principal, self)
        self.vistas["Catalogo"] = VistaCatalogo(self.contenedor_principal, self)
        self.vistas["Repuestos"] = VistaRepuestos(self.contenedor_principal, self)
        self.vistas["Cronograma"] = VistaCronograma(self.contenedor_principal, self)
        self.vistas["Historial"] = VistaHistorial(self.contenedor_principal, self)
        self.vistas["Analisis"] = VistaAnalisis(self.contenedor_principal, self)
        self.vistas["Protocolos"] = VistaProtocolos(self.contenedor_principal, self)
        self.vistas["Areas"] = VistaAreas(self.contenedor_principal, self)
        self.vistas["Respaldos"] = VistaRespaldos(self.contenedor_principal, self)
        if self.es_jefe:
            self.vistas["Usuarios"] = VistaUsuarios(self.contenedor_principal, self)

    def mostrar_vista(self, nombre):
        self.vista_actual_nombre = nombre
        for btn in self.botones_nav:
            btn.configure(fg_color="transparent", text_color=C_TEXT)
            
        for vista in self.vistas.values():
            vista.pack_forget()
            
        self.vistas[nombre].pack(fill="both", expand=True)
        
        if hasattr(self.vistas[nombre], 'refrescar_datos'):
            self.vistas[nombre].refrescar_datos()

            
        mapa_botones = {
            "Inventario": self.btn_nav_inv,
            "Catalogo": self.btn_nav_cat,
            "Repuestos": self.btn_nav_rep,
            "Cronograma": self.btn_nav_cro,
            "Historial": self.btn_nav_hist,
            "Analisis": self.btn_nav_analisis,
            "Protocolos": self.btn_nav_prot,
            "Areas": self.btn_nav_areas,
            "Respaldos": self.btn_nav_respaldos
        }
        if self.btn_nav_usuarios:
            mapa_botones["Usuarios"] = self.btn_nav_usuarios
            
        btn_sel = mapa_botones.get(nombre)
        if btn_sel:
            btn_sel.configure(fg_color=C_BLUE_LIGHT, text_color=C_BLUE)


        if nombre == "Cronograma":
            if self._calendario_sucio: 
                self.vistas["Cronograma"].dibujar_mes(self.vistas["Cronograma"].anio_actual, self.vistas["Cronograma"].mes_actual)
                self.vistas["Cronograma"].dibujar_anio(self.vistas["Cronograma"].anio_vista)
                self.vistas["Cronograma"].refrescar_datos()
                self._calendario_sucio = False

    def actualizar_boton_alertas(self):
        c = len(self.alertas_activas)
        self.btn_alertas.configure(text=f"🔔 Alertas ({c})")
        if c > 0: 
            self.btn_alertas.configure(fg_color=C_RED, hover_color=C_RED_HOVER, text_color="white")
        else: 
            self.btn_alertas.configure(fg_color=C_BG, text_color=C_TEXT)

    def mostrar_ventana_alertas(self):
        v = ctk.CTkToplevel(self)
        v.title("Alertas Activas")
        v.geometry("600x500")
        v.transient(self)
        v.grab_set()
        v.configure(fg_color=C_BG)
        
        ctk.CTkLabel(v, text="🔔 Alertas del Sistema", font=ctk.CTkFont(size=18, weight="bold"), text_color=C_TEXT).pack(pady=15)
        
        sf = ctk.CTkScrollableFrame(v, fg_color=C_CARD, corner_radius=12)
        sf.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        def refrescar_lista_alertas():
            # Limpiar frame
            for widget in sf.winfo_children():
                widget.destroy()
                
            if not self.alertas_activas:
                ctk.CTkLabel(sf, text="✅ Todo está al día. No hay alertas activas.", font=ctk.CTkFont(size=14), text_color=C_SUBTEXT).pack(pady=40)
            else:
                for a in list(self.alertas_activas):
                    f_row = ctk.CTkFrame(sf, fg_color="transparent")
                    f_row.pack(fill="x", pady=6, padx=5)
                    
                    lbl = ctk.CTkLabel(f_row, text=a, font=ctk.CTkFont(size=12), text_color=C_TEXT, anchor="w", justify="left")
                    lbl.pack(side="left", fill="x", expand=True, padx=5)
                    
                    def ignorar_alerta(alerta_text=a):
                        self.alertas_ignoradas.add(alerta_text)
                        if alerta_text in self.alertas_activas:
                            self.alertas_activas.remove(alerta_text)
                        self.actualizar_boton_alertas()
                        refrescar_lista_alertas()
                        
                    btn = ctk.CTkButton(f_row, text="Ignorar", width=70, height=26, fg_color=C_RED, hover_color=C_RED_HOVER, text_color="white", font=ctk.CTkFont(size=11, weight="bold"), corner_radius=6, command=ignorar_alerta)
                    btn.pack(side="right", padx=5)
                    
        refrescar_lista_alertas()

    def centrar_ventana_segura(self, toplevel, ancho_deseado, alto_deseado):
        """Calcula dimensiones seguras según la resolución del monitor y centra la ventana sin salirse de la pantalla."""
        toplevel.update_idletasks()
        sw = toplevel.winfo_screenwidth()
        sh = toplevel.winfo_screenheight()
        
        # Dejar márgenes de seguridad para barra de tareas de Windows y marcos de ventana
        max_w = max(400, sw - 40)
        max_h = max(350, sh - 80)
        
        ancho_final = min(ancho_deseado, max_w)
        alto_final = min(alto_deseado, max_h)
        
        x = max(10, (sw - ancho_final) // 2)
        y = max(10, (sh - alto_final) // 2)
        
        toplevel.geometry(f"{ancho_final}x{alto_final}+{x}+{y}")

    # ========================================================
    # FORMULARIO MAESTRO DE REGISTRO DE EQUIPOS
    # ========================================================
    def abrir_formulario_equipo(self, eq_edit=None):
        vent = ctk.CTkToplevel(self)
        vent.title("Ficha de Equipo")
        vent.transient(self)
        vent.grab_set()
        vent.configure(fg_color=C_BG)
        self.centrar_ventana_segura(vent, 800, 680)

        
        ctk.CTkLabel(vent, text="Ficha Técnica Institucional", font=ctk.CTkFont(size=22, weight="bold"), text_color=C_TEXT).pack(pady=15)
        
        sf = ctk.CTkScrollableFrame(vent, fg_color=C_CARD, corner_radius=12)
        sf.pack(pady=5, padx=20, fill="both", expand=True)
        
        ctk.CTkLabel(sf, text="1. Identificación y Ubicación", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(10, 5))
        
        # Red de Salud y Centro de Salud Oficiales GAMLP
        from database import obtener_jerarquia_sedes_db
        sedes_form_data = obtener_jerarquia_sedes_db()
        redes_opts = [r["nombre"] for r in sedes_form_data.get("redes", [])]
        if not redes_opts:
            redes_opts = [
                "RED 1-SUR OESTE (MACRODISTRITO COTAHUMA)",
                "RED 2-NOR OESTE (MACRODISTRITO MAX PAREDES)",
                "RED 3-NORTE CENTRAL (MACRODISTRITO PERIFERICA CENTRAL)",
                "RED 4-SAN ANTONIO (MACRODISTRITO SAN ANTONIO)",
                "RED 5-SUR (MACRODISTRITO SUR)"
            ]

        ctk.CTkLabel(sf, text="Red de Salud:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        combo_red_form = ctk.CTkComboBox(sf, width=500, values=redes_opts)
        combo_red_form.pack(pady=(0, 5))

        ctk.CTkLabel(sf, text="Centro de Salud / Establecimiento:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        combo_centro_form = ctk.CTkComboBox(sf, width=500, values=["Seleccione Red"])
        combo_centro_form.pack(pady=(0, 5))

        def al_cambiar_red_form(red_sel):
            red_obj = next((r for r in sedes_form_data.get("redes", []) if r["nombre"] == red_sel), None)
            red_id = red_obj["id"] if red_obj else None
            centros = [c["nombre"] for c in sedes_form_data.get("centros", []) if c.get("red_salud_id") == red_id]
            if not centros:
                centros = ["CENTRO DE SALUD"]
            combo_centro_form.configure(values=centros)
            combo_centro_form.set(centros[0])

        combo_red_form.configure(command=al_cambiar_red_form)

        # Pre-seleccionar según contexto de sede activa o por defecto
        sede_activa = getattr(self, "contexto_sede", {}) or {}
        red_default = sede_activa.get("red_salud")
        if red_default and red_default in redes_opts:
            combo_red_form.set(red_default)
        else:
            combo_red_form.set(redes_opts[0])
        al_cambiar_red_form(combo_red_form.get())

        centro_default = sede_activa.get("centro_salud")
        if centro_default and not str(centro_default).startswith("[ Todos"):
            combo_centro_form.set(centro_default)

        val_cat = [f"{c['nombre']} - {c.get('marca', '')} - {c.get('modelo', '')}" for c in self.datos["catalogo"]]
        ctk.CTkLabel(sf, text="Modelo de Catálogo:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        combo_tipo = ctk.CTkComboBox(sf, width=500, values=val_cat if val_cat else ["No hay modelos"])
        combo_tipo.pack(pady=(0, 5))
        habilitar_autocompletado(combo_tipo, val_cat)
        
        ctk.CTkLabel(sf, text="Unidad:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        val_areas = [a["nombre"] for a in self.datos["areas"]]
        combo_area = ctk.CTkComboBox(sf, width=500, values=val_areas if val_areas else ["No hay áreas"])
        combo_area.pack(pady=(0, 5))
        combo_area.configure(state="disabled")
        habilitar_autocompletado(combo_area, val_areas)

        ctk.CTkLabel(sf, text="Servicio (Ej. Rayos X):", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        e_servicio = ctk.CTkEntry(sf, placeholder_text="Servicio (Ej. Rayos X)", width=500)
        e_servicio.pack(pady=(0, 5))

        ctk.CTkLabel(sf, text="Código de Activos Fijos:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        e_id = ctk.CTkEntry(sf, placeholder_text="Código de Activo Fijo (único)", width=500)
        e_id.pack(pady=(0, 5))
        
        ctk.CTkLabel(sf, text="Número de Serie:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        e_serie = ctk.CTkEntry(sf, placeholder_text="Número de Serie del Equipo", width=500)
        e_serie.pack(pady=(0, 5))

        def al_seleccionar_tipo(val_sel):
            if not val_sel or "-" not in val_sel:
                return
            partes = [p.strip() for p in val_sel.split("-")]
            if len(partes) >= 1:
                nom = partes[0]
                mrc = partes[1] if len(partes) > 1 else ""
                mdl = partes[2] if len(partes) > 2 else ""
                match = next((c for c in self.datos.get("catalogo", []) if c["nombre"].strip().lower() == nom.lower() and (c.get("marca") or "").strip().lower() == mrc.lower() and (c.get("modelo") or "").strip().lower() == mdl.lower()), None)
                if match and match.get("area"):
                    combo_area.configure(state="normal")
                    combo_area.set(match["area"])
                    combo_area.configure(state="disabled")
                
                # Buscar en equipos existentes para auto-rellenar procedencia, fabricante, proveedor y año de fabricación
                eq_match = next((eq for eq in self.datos.get("equipos", []) if 
                                 eq["nombre"].strip().lower() == nom.lower() and 
                                 (eq.get("marca") or "").strip().lower() == mrc.lower() and 
                                 (eq.get("modelo") or "").strip().lower() == mdl.lower()), None)
                
                if eq_match:
                    if eq_match.get("procedencia"):
                        e_procedencia.delete(0, "end"); e_procedencia.insert(0, eq_match.get("procedencia") or "")
                    if eq_match.get("fabricante"):
                        e_fabricante.delete(0, "end"); e_fabricante.insert(0, eq_match.get("fabricante") or "")
                    if eq_match.get("proveedor"):
                        e_proveedor.delete(0, "end"); e_proveedor.insert(0, eq_match.get("proveedor") or "")
                    if eq_match.get("anio_fab"):
                        e_anio.delete(0, "end"); e_anio.insert(0, eq_match.get("anio_fab") or "")
                    
        combo_tipo.configure(command=al_seleccionar_tipo)
 
        ctk.CTkLabel(sf, text="2. Adquisición", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(15, 5))
        
        ctk.CTkLabel(sf, text="Procedencia:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        e_procedencia = ctk.CTkEntry(sf, placeholder_text="Procedencia", width=500)
        e_procedencia.pack(pady=(0, 5))
        
        ctk.CTkLabel(sf, text="Fabricante Original:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        e_fabricante = ctk.CTkEntry(sf, placeholder_text="Fabricante Original", width=500)
        e_fabricante.pack(pady=(0, 5))
        
        ctk.CTkLabel(sf, text="Proveedor Local:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        e_proveedor = ctk.CTkEntry(sf, placeholder_text="Proveedor Local", width=500)
        e_proveedor.pack(pady=(0, 5))
        
        ctk.CTkLabel(sf, text="Año de Fabricación:", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w", padx=150, pady=(5, 0))
        e_anio = ctk.CTkEntry(sf, placeholder_text="Año de Fabricación", width=500)
        e_anio.pack(pady=(0, 5))
        
        f_cal = ctk.CTkFrame(sf, fg_color="transparent")
        f_cal.pack(pady=5, anchor="w", padx=150)
        
        ctk.CTkLabel(f_cal, text="Fecha de Instalación:", text_color=C_SUBTEXT).pack(side="left", padx=5)
        cal_adq = DateEntry(f_cal, width=15, font=('Segoe UI', 11), background=C_BLUE, foreground='white', borderwidth=0, date_pattern='y-mm-dd')
        cal_adq.pack(side="left")

        ctk.CTkLabel(sf, text="3. Especificaciones", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(15, 5))
        
        frame_checks = ctk.CTkFrame(sf, fg_color=C_BG, corner_radius=10)
        frame_checks.pack(fill="x", pady=5)

        self.var_electrico = ctk.StringVar()
        self.var_electronico = ctk.StringVar()
        self.var_mecanico = ctk.StringVar()
        self.var_hidraulico = ctk.StringVar()
        self.var_neumatico = ctk.StringVar()
        self.var_vapor = ctk.StringVar()
        
        ft = ctk.CTkFrame(frame_checks, fg_color="transparent")
        ft.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(ft, text="Tecnología:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkCheckBox(ft, text="Eléctrico", variable=self.var_electrico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Electrónico", variable=self.var_electronico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Mecánico", variable=self.var_mecanico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Hidráulico", variable=self.var_hidraulico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Neumático", variable=self.var_neumatico, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(ft, text="Vapor", variable=self.var_vapor, onvalue="X", offvalue="").pack(anchor="w", pady=2)

        self.var_compra = ctk.StringVar()
        self.var_comodato = ctk.StringVar()
        self.var_donacion = ctk.StringVar()
        
        fa = ctk.CTkFrame(frame_checks, fg_color="transparent")
        fa.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(fa, text="Adquisición:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkCheckBox(fa, text="Compra", variable=self.var_compra, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fa, text="Comodato", variable=self.var_comodato, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fa, text="Donación", variable=self.var_donacion, onvalue="X", offvalue="").pack(anchor="w", pady=2)

        self.var_fijo = ctk.StringVar()
        self.var_movil = ctk.StringVar()
        self.var_portatil = ctk.StringVar()
        
        fti = ctk.CTkFrame(frame_checks, fg_color="transparent")
        fti.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(fti, text="Tipo:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkCheckBox(fti, text="Fijo", variable=self.var_fijo, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fti, text="Móvil", variable=self.var_movil, onvalue="X", offvalue="").pack(anchor="w", pady=2)
        ctk.CTkCheckBox(fti, text="Portátil", variable=self.var_portatil, onvalue="X", offvalue="").pack(anchor="w", pady=2)

        self.var_garantia = ctk.StringVar()
        
        fg = ctk.CTkFrame(frame_checks, fg_color="transparent")
        fg.pack(side="left", expand=True, padx=5, pady=10)
        ctk.CTkLabel(fg, text="Garantía:", font=ctk.CTkFont(weight="bold", size=12)).pack(anchor="w")
        ctk.CTkRadioButton(fg, text="Sí", variable=self.var_garantia, value="Con Garantía").pack(anchor="w", pady=2)
        ctk.CTkRadioButton(fg, text="No", variable=self.var_garantia, value="Sin Garantía").pack(anchor="w", pady=2)
        
        # Fecha de inicio y vencimiento de garantía
        self.f_gar_frame = ctk.CTkFrame(fg, fg_color="transparent")
        
        f_ini = ctk.CTkFrame(self.f_gar_frame, fg_color="transparent")
        f_ini.pack(anchor="w", pady=2)
        ctk.CTkLabel(f_ini, text="Inicio:    ", font=ctk.CTkFont(size=10, weight="bold"), text_color=C_TEXT).pack(side="left", padx=(0, 5))
        self.cal_gar_inicio = DateEntry(f_ini, width=12, font=('Segoe UI', 10), background=C_BLUE, foreground='white', borderwidth=0, date_pattern='y-mm-dd')
        self.cal_gar_inicio.pack(side="left")
        
        f_fin = ctk.CTkFrame(self.f_gar_frame, fg_color="transparent")
        f_fin.pack(anchor="w", pady=2)
        ctk.CTkLabel(f_fin, text="Vence el:", font=ctk.CTkFont(size=10, weight="bold"), text_color=C_TEXT).pack(side="left", padx=(0, 5))
        self.cal_gar = DateEntry(f_fin, width=12, font=('Segoe UI', 10), background=C_BLUE, foreground='white', borderwidth=0, date_pattern='y-mm-dd')
        self.cal_gar.pack(side="left")
        
        self.lbl_restante_gar = ctk.CTkLabel(fg, text="", font=ctk.CTkFont(size=10, weight="bold"), text_color=C_BLUE)
        self.lbl_restante_gar.pack(anchor="w", pady=(2, 0))

        def actualizar_restante_gar(*args):
            if self.var_garantia.get() == "Con Garantía":
                self.f_gar_frame.pack(anchor="w", pady=(5, 0))
                try:
                    fecha_ini = self.cal_gar_inicio.get_date()
                    fecha_venc = self.cal_gar.get_date()
                    from datetime import date
                    hoy = date.today()
                    
                    if fecha_venc < hoy:
                        self.lbl_restante_gar.configure(text="⚠️ Garantía Vencida", text_color=C_RED)
                    elif hoy < fecha_ini:
                        from dateutil.relativedelta import relativedelta
                        diff = relativedelta(fecha_ini, hoy)
                        parts = []
                        if diff.years > 0: parts.append(f"{diff.years} {'año' if diff.years == 1 else 'años'}")
                        if diff.months > 0: parts.append(f"{diff.months} {'mes' if diff.months == 1 else 'meses'}")
                        if diff.days > 0: parts.append(f"{diff.days} {'día' if diff.days == 1 else 'días'}")
                        dur = ", ".join(parts) if parts else "0 días"
                        self.lbl_restante_gar.configure(text=f"Aún no de alta (inicia en: {dur})", text_color=C_ORANGE)
                    else:
                        from dateutil.relativedelta import relativedelta
                        diff = relativedelta(fecha_venc, hoy)
                        parts = []
                        if diff.years > 0:
                            parts.append(f"{diff.years} {'año' if diff.years == 1 else 'años'}")
                        if diff.months > 0:
                            parts.append(f"{diff.months} {'mes' if diff.months == 1 else 'meses'}")
                        if diff.days > 0:
                            parts.append(f"{diff.days} {'día' if diff.days == 1 else 'días'}")
                        duracion = ", ".join(parts) if parts else "Hoy"
                        self.lbl_restante_gar.configure(text=f"Resta: {duracion}", text_color=C_BLUE)
                except:
                    self.lbl_restante_gar.configure(text="")
            else:
                self.f_gar_frame.pack_forget()
                self.lbl_restante_gar.configure(text="")
                
        self.var_garantia.trace_add("write", actualizar_restante_gar)
        self.cal_gar_inicio.bind("<<DateEntrySelected>>", actualizar_restante_gar)
        self.cal_gar.bind("<<DateEntrySelected>>", actualizar_restante_gar)

        # Contenedor externo para el costo del equipo - SIEMPRE PRESENTE para mantener posición
        # Solo el contenido interno (label + entry) se muestra u oculta
        self.f_costo_container = ctk.CTkFrame(sf, fg_color="transparent", height=0)
        self.f_costo_container.pack(fill="x", pady=0)
        
        self.lbl_costo_title = ctk.CTkLabel(self.f_costo_container, text="Costo del Equipo (Bs.):", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT)
        self.e_costo = ctk.CTkEntry(self.f_costo_container, placeholder_text="Costo del Equipo (Bs.)", width=500)
        
        def actualizar_costo_visibilidad(*args):
            if self.var_compra.get() == "X":
                self.lbl_costo_title.pack(anchor="w", padx=150, pady=(5, 0))
                self.e_costo.pack(pady=(0, 5))
            else:
                self.lbl_costo_title.pack_forget()
                self.e_costo.pack_forget()
                self.e_costo.delete(0, "end")
                self.e_costo.insert(0, "0")

        def on_compra_change(*args):
            if self.var_compra.get() == "X":
                self.var_comodato.set("")
                self.var_donacion.set("")
            actualizar_costo_visibilidad()

        def on_comodato_change(*args):
            if self.var_comodato.get() == "X":
                self.var_compra.set("")
                self.var_donacion.set("")
            actualizar_costo_visibilidad()

        def on_donacion_change(*args):
            if self.var_donacion.get() == "X":
                self.var_compra.set("")
                self.var_comodato.set("")
            actualizar_costo_visibilidad()

        self.var_compra.trace_add("write", on_compra_change)
        self.var_comodato.trace_add("write", on_comodato_change)
        self.var_donacion.trace_add("write", on_donacion_change)
        actualizar_costo_visibilidad()

        ctk.CTkLabel(sf, text="4. Criticidad", font=ctk.CTkFont(weight="bold", size=14), text_color=C_BLUE).pack(anchor="w", pady=(15, 5))
        
        p_cat = [
            "Intercambiabilidad", 
            "Funcion Clinica", 
            "Frecuencia de Uso", 
            "Impacto en el Servicio", 
            "Mantenibilidad", 
            "Historial de Fallas", 
            "Complejidad Teconologica", 
            "Valor de compra", 
            "Exigencia Normativa", 
            "Seguridad Operacional", 
            "Vulnerabilidad Ambiental", 
            "Riesgo a explosiones", 
            "Edad del equipo"
        ]
        f_cat = ctk.CTkFrame(sf, fg_color=C_BG, corner_radius=10)
        f_cat.pack(fill="x", pady=5, padx=20)
        
        # Configurar columnas con pesos y anchos mínimos
        f_cat.grid_columnconfigure(0, weight=3, minsize=200)
        f_cat.grid_columnconfigure(1, weight=1, minsize=80)
        f_cat.grid_columnconfigure(2, weight=1, minsize=80)
        f_cat.grid_columnconfigure(3, weight=1, minsize=80)
        
        # Cabecera de columnas para las opciones
        ctk.CTkLabel(f_cat, text="", anchor="w").grid(row=0, column=0, padx=(20, 10), pady=6, sticky="w")
        for col_idx, text in [(1, "I"), (2, "II"), (3, "III")]:
            ctk.CTkLabel(f_cat, text=text, font=ctk.CTkFont(weight="bold", size=13), text_color=C_BLUE).grid(row=0, column=col_idx, sticky="nsew", pady=6)
        
        self.variables_cat = []
        for i, param in enumerate(p_cat):
            ctk.CTkLabel(f_cat, text=param, anchor="w", font=ctk.CTkFont(size=12)).grid(row=i+1, column=0, padx=(20, 10), pady=4, sticky="w")
            var = ctk.StringVar(value="")
            self.variables_cat.append(var)
            
            for col_idx, val in [(1, "1"), (2, "2"), (3, "3")]:
                f_btn = ctk.CTkFrame(f_cat, fg_color="transparent")
                f_btn.grid(row=i+1, column=col_idx, sticky="nsew", pady=4)
                rb = ctk.CTkRadioButton(f_btn, text="", variable=var, value=val, width=20, height=20)
                rb.pack(expand=True)

        combo_estado = ctk.CTkComboBox(sf, values=["Operativo", "Baja"], width=500)
        combo_estado.pack(pady=(20,5))
        
        # FUNCIONALIDAD DE FOTO CON COMPRESIÓN Y SINCRONIZACIÓN EN LA NUBE
        ruta_foto = ctk.StringVar(value="")
        
        def seleccionar_foto():
            r = filedialog.askopenfilename(filetypes=[("Imágenes", "*.jpg;*.jpeg;*.png;*.webp;*.bmp")])
            if r:
                b64 = comprimir_imagen_base64(r)
                if b64:
                    ruta_foto.set(b64)
                    btn_foto.configure(text="✅ Fotografía Comprimida y Sincronizada", fg_color="#16A34A", hover_color="#15803D")
                else:
                    ruta_foto.set(r)
                    btn_foto.configure(text="✅ Fotografía Adjuntada", fg_color=C_BLUE, hover_color=C_BLUE_HOVER)
                
        btn_foto = ctk.CTkButton(sf, text="📸 Adjuntar Fotografía del Equipo", command=seleccionar_foto, fg_color=C_BLUE, hover_color=C_BLUE_HOVER, width=500, font=ctk.CTkFont(weight="bold", size=14), height=35)
        btn_foto.pack(pady=10)

        # --- SECCIÓN ADICIONAL PARA DATOS TÉCNICOS Y CONTEXTO OPERACIONAL (EXCEL) ---
        self.f_adicionales_container = ctk.CTkFrame(sf, fg_color="transparent")
        self.adicionales_visible = False
        
        # Grid para los 9 campos técnicos oficiales
        f_grid = ctk.CTkFrame(self.f_adicionales_container, fg_color="transparent")
        f_grid.pack(fill="x", pady=5)
        f_grid.grid_columnconfigure(0, weight=1)
        f_grid.grid_columnconfigure(1, weight=1)
        
        def crear_campo_grid(row, col, lbl_txt, ph_txt):
            f_item = ctk.CTkFrame(f_grid, fg_color="transparent")
            f_item.grid(row=row, column=col, padx=10, pady=5, sticky="ew")
            ctk.CTkLabel(f_item, text=lbl_txt, font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT).pack(anchor="w")
            e_field = ctk.CTkEntry(f_item, placeholder_text=ph_txt, height=30)
            e_field.pack(fill="x", pady=2)
            return e_field

        self.e_voltaje = crear_campo_grid(0, 0, "Voltaje:", "Voltaje (ej: 220V)")
        self.e_corriente = crear_campo_grid(0, 1, "Corriente:", "Corriente (ej: 5A)")
        self.e_potencia = crear_campo_grid(1, 0, "Potencia Consumida:", "Potencia (ej: 500W)")
        self.e_vida_util = crear_campo_grid(1, 1, "Vida Útil Estimada:", "Vida útil (ej: 10 años)")
        self.e_peso = crear_campo_grid(2, 0, "Peso:", "Peso (ej: 50 kg)")
        self.e_dimensiones = crear_campo_grid(2, 1, "Dimensiones:", "Dimensiones (ej: 120x80x90 cm)")
        self.e_bateria_respaldo = crear_campo_grid(3, 0, "Batería de Respaldo:", "Batería (ej: 12V 7Ah / Sí / No)")
        self.e_version_software = crear_campo_grid(3, 1, "Versión Software:", "Versión (ej: v2.4.1)")
        self.e_suministro_gases = crear_campo_grid(4, 0, "Suministro de Gases:", "Gases (ej: O2 / Aire / Vacío)")

        # Helper para los 9 campos de texto largo
        def crear_campo_texto(label_text, placeholder):
            lbl = ctk.CTkLabel(self.f_adicionales_container, text=label_text, font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT)
            lbl.pack(anchor="w", padx=10, pady=(8, 2))
            txt = ctk.CTkTextbox(self.f_adicionales_container, height=55, border_width=1, border_color=C_BORDER, fg_color=C_BG)
            txt.pack(fill="x", padx=10, pady=(0, 5))
            return txt

        self.txt_contexto = crear_campo_texto("Contexto Operacional:", "Describa el contexto operacional...")
        self.txt_funciones = crear_campo_texto("Funciones del Equipo Médico:", "Describa las funciones del equipo...")
        self.txt_acciones_prev = crear_campo_texto("Acciones Preventivas:", "Describa las acciones preventivas...")
        self.txt_acciones_falla = crear_campo_texto("Insumos / Accesorios:", "Describa insumos y accesorios...")
        self.txt_fallas_func = crear_campo_texto("Fallas Comunes:", "Describa las fallas comunes...")
        self.txt_causas_fallo = crear_campo_texto("Causas de Fallo en el Equipo:", "Describa las causas de fallo...")
        self.txt_efectos_fallo = crear_campo_texto("Consecuencias de Fallo en el Equipo:", "Describa las consecuencias de fallo...")
        self.txt_efecto_entorno = crear_campo_texto("Acciones Correctivas Comunes:", "Describa las acciones correctivas...")
        self.txt_observaciones_ficha = crear_campo_texto("Observaciones:", "Observaciones adicionales...")

        def toggle_datos_adicionales():
            if self.adicionales_visible:
                self.f_adicionales_container.pack_forget()
                btn_adicionales.configure(text="▼ Mostrar Datos Técnicos y Contexto (Excel)")
                self.adicionales_visible = False
            else:
                self.f_adicionales_container.pack(fill="x", pady=10)
                btn_adicionales.configure(text="▲ Ocultar Datos Técnicos y Contexto (Excel)")
                self.adicionales_visible = True

        btn_adicionales = ctk.CTkButton(sf, text="▼ Mostrar Datos Técnicos y Contexto (Excel)", command=toggle_datos_adicionales, fg_color=C_BLUE, hover_color=C_BLUE_HOVER, width=500, font=ctk.CTkFont(weight="bold", size=13), height=35)
        btn_adicionales.pack(pady=10)

        if eq_edit:
            red_nom_edit = eq_edit.get("red_salud_nombre") or eq_edit.get("red_salud")
            if red_nom_edit and red_nom_edit in redes_opts:
                combo_red_form.set(red_nom_edit)
                al_cambiar_red_form(red_nom_edit)
            
            cen_nom_edit = eq_edit.get("centro_salud_nombre") or eq_edit.get("centro_salud")
            if cen_nom_edit:
                combo_centro_form.set(cen_nom_edit)

            e_id.insert(0, eq_edit["id"] or "")
            e_id.configure(state="disabled")
            e_serie.insert(0, eq_edit.get("numero_serie") or "")
            e_serie.configure(state="disabled")
            
            e_servicio.insert(0, eq_edit.get("servicio") or "")
            combo_area.configure(state="normal")
            combo_area.set(eq_edit.get("area") or "")
            combo_area.configure(state="disabled")
            e_procedencia.insert(0, eq_edit.get("procedencia") or "")
            e_fabricante.insert(0, eq_edit.get("fabricante") or "")
            e_proveedor.insert(0, eq_edit.get("proveedor") or "")
            e_anio.insert(0, eq_edit.get("anio_fab") or "")
            self.e_costo.insert(0, str(eq_edit.get("costo") or 0))
            
            combo_estado.set(eq_edit.get("estado","Operativo"))
            combo_tipo.set(f"{eq_edit['nombre']} - {eq_edit.get('marca','')} - {eq_edit.get('modelo','')}")
            
            self.var_electrico.set(eq_edit.get("t_elec", ""))
            self.var_electronico.set(eq_edit.get("t_elco", ""))
            self.var_mecanico.set(eq_edit.get("t_mec", ""))
            self.var_hidraulico.set(eq_edit.get("t_hid", ""))
            self.var_neumatico.set(eq_edit.get("t_neu", ""))
            self.var_vapor.set(eq_edit.get("t_vap", ""))
            
            self.var_compra.set(eq_edit.get("a_comp", ""))
            self.var_comodato.set(eq_edit.get("a_como", ""))
            self.var_donacion.set(eq_edit.get("a_don", ""))
            
            self.var_fijo.set(eq_edit.get("te_fijo", ""))
            self.var_movil.set(eq_edit.get("te_mov", ""))
            self.var_portatil.set(eq_edit.get("te_por", ""))
            
            self.var_garantia.set(eq_edit.get("garantia", "Sin Garantía"))
            f_venc = eq_edit.get("fecha_vencimiento_garantia")
            if f_venc:
                if isinstance(f_venc, str):
                    try:
                        f_venc = datetime.strptime(f_venc, "%Y-%m-%d").date()
                    except:
                        f_venc = None
                if f_venc:
                    self.cal_gar.set_date(f_venc)
            
            f_ini = eq_edit.get("fecha_inicio_garantia")
            if f_ini:
                if isinstance(f_ini, str):
                    try:
                        f_ini = datetime.strptime(f_ini, "%Y-%m-%d").date()
                    except:
                        f_ini = None
                if f_ini:
                    self.cal_gar_inicio.set_date(f_ini)
                    
            actualizar_restante_gar()
            
            foto_guardada = eq_edit.get("foto", "")
            if foto_guardada:
                if not foto_guardada.startswith("data:image") and os.path.exists(foto_guardada):
                    foto_guardada = comprimir_imagen_base64(foto_guardada)
                ruta_foto.set(foto_guardada)
                btn_foto.configure(text="✅ Fotografía Existente en Sistema", fg_color="#16A34A", hover_color="#15803D")
            
            c_det = eq_edit.get("categorizacion_detalle") or []
            if isinstance(c_det, str):
                c_det = json.loads(c_det)
                
            for i, val in enumerate(c_det):
                if i < len(self.variables_cat):
                    if val in ("I","II","III"):
                        newval = "3" if val == "I" else ("2" if val == "II" else "1")
                        self.variables_cat[i].set(newval)
                    else:
                        self.variables_cat[i].set(str(val))

            # Cargar datos adicionales oficiales
            self.e_voltaje.insert(0, eq_edit.get("voltaje") or "")
            self.e_corriente.insert(0, eq_edit.get("corriente") or "")
            self.e_potencia.insert(0, eq_edit.get("potencia") or "")
            self.e_vida_util.insert(0, eq_edit.get("vida_util") or eq_edit.get("temperatura") or "")
            self.e_peso.insert(0, eq_edit.get("peso") or "")
            self.e_dimensiones.insert(0, eq_edit.get("dimensiones") or "")
            self.e_bateria_respaldo.insert(0, eq_edit.get("bateria_respaldo") or eq_edit.get("resolucion") or "")
            self.e_version_software.insert(0, eq_edit.get("version_software") or eq_edit.get("humedad") or "")
            self.e_suministro_gases.insert(0, eq_edit.get("suministro_gases") or "")
            
            self.txt_contexto.delete("1.0", "end")
            self.txt_contexto.insert("1.0", eq_edit.get("contexto_operacional") or "")
            
            self.txt_funciones.delete("1.0", "end")
            self.txt_funciones.insert("1.0", eq_edit.get("funciones_equipo") or "")
            
            self.txt_acciones_prev.delete("1.0", "end")
            self.txt_acciones_prev.insert("1.0", eq_edit.get("acciones_preventivas") or "")
            
            self.txt_acciones_falla.delete("1.0", "end")
            self.txt_acciones_falla.insert("1.0", eq_edit.get("acciones_falla") or "")
            
            self.txt_fallas_func.delete("1.0", "end")
            self.txt_fallas_func.insert("1.0", eq_edit.get("fallas_funcionales") or "")
            
            self.txt_causas_fallo.delete("1.0", "end")
            self.txt_causas_fallo.insert("1.0", eq_edit.get("causas_fallo") or "")
            
            self.txt_efectos_fallo.delete("1.0", "end")
            self.txt_efectos_fallo.insert("1.0", eq_edit.get("efectos_fallo") or "")
            
            self.txt_efecto_entorno.delete("1.0", "end")
            self.txt_efecto_entorno.insert("1.0", eq_edit.get("efecto_entorno") or "")
            
            self.txt_observaciones_ficha.delete("1.0", "end")
            self.txt_observaciones_ficha.insert("1.0", eq_edit.get("observaciones") or "")
        else:
            self.var_garantia.set("Sin Garantía")
            actualizar_restante_gar()

        def guardar():
            if not e_id.get(): 
                return
                
            # Validar si el ID o Serie ya existen al crear un equipo nuevo (Validación instantánea en memoria)
            id_val = e_id.get().strip()
            serie_val = e_serie.get().strip()
            if eq_edit is None:
                if any(str(eq.get("id", "")).strip().lower() == id_val.lower() for eq in self.datos.get("equipos", [])):
                    messagebox.showerror("Error de Duplicidad", f"Ya existe un equipo registrado con el Código de Activo Fijo '{id_val}'.")
                    return
                if serie_val and any(str(eq.get("numero_serie", "")).strip().lower() == serie_val.lower() for eq in self.datos.get("equipos", [])):
                    messagebox.showerror("Error de Duplicidad", f"Ya existe un equipo registrado con el Número de Serie '{serie_val}'.")
                    return
                 
            ts = combo_tipo.get().split(" - ")
            n_nom = ts[0]
            n_mar = ts[1] if len(ts)>1 else ""
            n_mod = ts[2] if len(ts)>2 else ""
            
            puntajes = []
            for v in self.variables_cat:
                if v.get():
                    puntajes.append(int(v.get()))
                else:
                    puntajes.append(0)
                    
            puntaje_total = sum(puntajes)
            
            if puntaje_total >= 30:
                criticidad_final = "Riesgo Alto"
            elif puntaje_total >= 20:
                criticidad_final = "Riesgo Medio"
            else:
                criticidad_final = "Riesgo Bajo"
                
            detalles_cat = [v.get() for v in self.variables_cat]
            
            f_gar_val = self.cal_gar.get_date() if self.var_garantia.get() == "Con Garantía" else None
            f_gar_ini_val = self.cal_gar_inicio.get_date() if self.var_garantia.get() == "Con Garantía" else None
            
            try:
                costo_val = float(self.e_costo.get().strip().replace(",", "."))
                if costo_val < 0:
                    raise ValueError()
            except:
                costo_val = 0.0

            def _get_txt(widget):
                try:
                    return widget.get("1.0", "end-1c").strip()
                except Exception:
                    return ""

            def _get_entry(widget):
                try:
                    return widget.get().strip()
                except Exception:
                    return ""

            red_sel_val = combo_red_form.get().strip()
            cen_sel_val = combo_centro_form.get().strip()
            red_obj_val = next((r for r in sedes_form_data.get("redes", []) if r["nombre"] == red_sel_val), None)
            cen_obj_val = next((c for c in sedes_form_data.get("centros", []) if c["nombre"] == cen_sel_val), None)
            red_id_val = red_obj_val["id"] if red_obj_val else None
            cen_id_val = cen_obj_val["id"] if cen_obj_val else None

            # Construir objeto equipo para actualización inmediata en memoria
            eq_dict = {
                "id": e_id.get().strip(),
                "nombre": n_nom,
                "marca": n_mar,
                "modelo": n_mod,
                "servicio": e_servicio.get().strip(),
                "area": combo_area.get(),
                "red_salud_id": red_id_val,
                "red_salud_nombre": red_sel_val,
                "centro_salud_id": cen_id_val,
                "centro_salud_nombre": cen_sel_val,
                "municipio_nombre": "GAMLP",
                "departamento_nombre": "La Paz",
                "procedencia": e_procedencia.get().strip(),
                "fabricante": e_fabricante.get().strip(),
                "proveedor": e_proveedor.get().strip(),
                "anio_fab": e_anio.get().strip(),
                "t_elec": self.var_electrico.get(),
                "t_elco": self.var_electronico.get(),
                "t_mec": self.var_mecanico.get(),
                "t_hid": self.var_hidraulico.get(),
                "t_neu": self.var_neumatico.get(),
                "t_vap": self.var_vapor.get(),
                "a_comp": self.var_compra.get(),
                "a_como": self.var_comodato.get(),
                "a_don": self.var_donacion.get(),
                "te_fijo": self.var_fijo.get(),
                "te_mov": self.var_movil.get(),
                "te_por": self.var_portatil.get(),
                "garantia": self.var_garantia.get(),
                "criticidad": criticidad_final,
                "categorizacion_detalle": json.dumps(detalles_cat),
                "estado": combo_estado.get(),
                "fecha_adquisicion": str(cal_adq.get_date()),
                "fecha_registro": datetime.now().strftime("%Y-%m-%d"),
                "foto": ruta_foto.get(),
                "fecha_vencimiento_garantia": str(f_gar_val) if f_gar_val else None,
                "numero_serie": e_serie.get().strip(),
                "fecha_inicio_garantia": str(f_gar_ini_val) if f_gar_ini_val else None,
                "costo": costo_val,
                "voltaje": _get_entry(self.e_voltaje),
                "corriente": _get_entry(self.e_corriente),
                "potencia": _get_entry(self.e_potencia),
                "vida_util": _get_entry(self.e_vida_util),
                "temperatura": _get_entry(self.e_vida_util),
                "peso": _get_entry(self.e_peso),
                "dimensiones": _get_entry(self.e_dimensiones),
                "bateria_respaldo": _get_entry(self.e_bateria_respaldo),
                "resolucion": _get_entry(self.e_bateria_respaldo),
                "version_software": _get_entry(self.e_version_software),
                "humedad": _get_entry(self.e_version_software),
                "suministro_gases": _get_entry(self.e_suministro_gases),
                "contexto_operacional": _get_txt(self.txt_contexto),
                "funciones_equipo": _get_txt(self.txt_funciones),
                "acciones_preventivas": _get_txt(self.txt_acciones_prev),
                "acciones_falla": _get_txt(self.txt_acciones_falla),
                "fallas_funcionales": _get_txt(self.txt_fallas_func),
                "causas_fallo": _get_txt(self.txt_causas_fallo),
                "efectos_fallo": _get_txt(self.txt_efectos_fallo),
                "efecto_entorno": _get_txt(self.txt_efecto_entorno),
                "observaciones": _get_txt(self.txt_observaciones_ficha),
                "historial_intervenciones": eq_edit.get("historial_intervenciones", []) if eq_edit else []
            }

            # 1. Actualizar memoria y caché de inmediato (0 ms)
            if eq_edit:
                for idx_e, ex in enumerate(self.datos.get("equipos", [])):
                    if str(ex.get("id")) == str(eq_edit.get("id")):
                        self.datos["equipos"][idx_e] = eq_dict
                        break
            else:
                self.datos.setdefault("equipos", []).append(eq_dict)

            guardar_cache_local_datos(self.datos)
            self._calendario_sucio = True
            self.vistas["Inventario"].refrescar_datos()
            vent.destroy()

            # 2. Guardar en PostgreSQL en segundo plano sin congelar la pantalla
            def _guardar_equipo_db(eq_data):
                conn = obtener_conexion()
                if conn:
                    try:
                        cur = conn.cursor()
                        # Resolver IDs reales en base de datos para evitar cualquier desface de llaves foráneas
                        red_id = None
                        cen_id = None
                        if eq_data.get("red_salud_nombre"):
                            cur.execute("SELECT id FROM redes_salud WHERE nombre = %s OR codigo = %s OR nombre ILIKE %s LIMIT 1;", 
                                        (eq_data["red_salud_nombre"], eq_data.get("red_salud_nombre", ""), f"%{eq_data['red_salud_nombre']}%"))
                            r_row = cur.fetchone()
                            if r_row: red_id = r_row[0]
                            
                        if eq_data.get("centro_salud_nombre"):
                            cur.execute("SELECT id FROM centros_salud WHERE nombre = %s OR nombre ILIKE %s LIMIT 1;", 
                                        (eq_data["centro_salud_nombre"], f"%{eq_data['centro_salud_nombre']}%"))
                            c_row = cur.fetchone()
                            if c_row: cen_id = c_row[0]

                        sql_q = """
                            INSERT INTO equipos (
                                id, nombre, marca, modelo, servicio, area, procedencia, fabricante, proveedor, anio_fab,
                                t_elec, t_elco, t_mec, t_hid, t_neu, t_vap, a_comp, a_como, a_don, te_fijo, te_mov, te_por, garantia, criticidad, categorizacion_detalle, estado, fecha_adquisicion, fecha_registro, foto, fecha_vencimiento_garantia, numero_serie, fecha_inicio_garantia, costo,
                                voltaje, corriente, potencia, vida_util, temperatura, peso, dimensiones, bateria_respaldo, resolucion, version_software, humedad, suministro_gases, contexto_operacional, funciones_equipo, acciones_preventivas, acciones_falla, fallas_funcionales, causas_fallo, efectos_fallo, efecto_entorno, observaciones,
                                red_salud_id, red_salud_nombre, centro_salud_id, centro_salud_nombre, municipio_nombre, departamento_nombre
                            )
                            VALUES (
                                %(id)s, %(nombre)s, %(marca)s, %(modelo)s, %(servicio)s, %(area)s, %(procedencia)s, %(fabricante)s, %(proveedor)s, %(anio_fab)s,
                                %(t_elec)s, %(t_elco)s, %(t_mec)s, %(t_hid)s, %(t_neu)s, %(t_vap)s, %(a_comp)s, %(a_como)s, %(a_don)s, %(te_fijo)s, %(te_mov)s, %(te_por)s, %(garantia)s, %(criticidad)s, %(categorizacion_detalle)s, %(estado)s, %(fecha_adquisicion)s, %(fecha_registro)s, %(foto)s, %(fecha_vencimiento_garantia)s, %(numero_serie)s, %(fecha_inicio_garantia)s, %(costo)s,
                                %(voltaje)s, %(corriente)s, %(potencia)s, %(vida_util)s, %(temperatura)s, %(peso)s, %(dimensiones)s, %(bateria_respaldo)s, %(resolucion)s, %(version_software)s, %(humedad)s, %(suministro_gases)s, %(contexto_operacional)s, %(funciones_equipo)s, %(acciones_preventivas)s, %(acciones_falla)s, %(fallas_funcionales)s, %(causas_fallo)s, %(efectos_fallo)s, %(efecto_entorno)s, %(observaciones)s,
                                %(red_salud_id)s, %(red_salud_nombre)s, %(centro_salud_id)s, %(centro_salud_nombre)s, %(municipio_nombre)s, %(departamento_nombre)s
                            )
                            ON CONFLICT (id) DO UPDATE SET
                                nombre=EXCLUDED.nombre, marca=EXCLUDED.marca, modelo=EXCLUDED.modelo, servicio=EXCLUDED.servicio, area=EXCLUDED.area, procedencia=EXCLUDED.procedencia, fabricante=EXCLUDED.fabricante, proveedor=EXCLUDED.proveedor, anio_fab=EXCLUDED.anio_fab,
                                t_elec=EXCLUDED.t_elec, t_elco=EXCLUDED.t_elco, t_mec=EXCLUDED.t_mec, t_hid=EXCLUDED.t_hid, t_neu=EXCLUDED.t_neu, t_vap=EXCLUDED.t_vap, a_comp=EXCLUDED.a_comp, a_como=EXCLUDED.a_como, a_don=EXCLUDED.a_don,
                                te_fijo=EXCLUDED.te_fijo, te_mov=EXCLUDED.te_mov, te_por=EXCLUDED.te_por, garantia=EXCLUDED.garantia, criticidad=EXCLUDED.criticidad, categorizacion_detalle=EXCLUDED.categorizacion_detalle, estado=EXCLUDED.estado, fecha_adquisicion=EXCLUDED.fecha_adquisicion, foto=EXCLUDED.foto, fecha_vencimiento_garantia=EXCLUDED.fecha_vencimiento_garantia, numero_serie=EXCLUDED.numero_serie, fecha_inicio_garantia=EXCLUDED.fecha_inicio_garantia, costo=EXCLUDED.costo,
                                voltaje=EXCLUDED.voltaje, corriente=EXCLUDED.corriente, potencia=EXCLUDED.potencia, vida_util=EXCLUDED.vida_util, temperatura=EXCLUDED.temperatura, peso=EXCLUDED.peso, dimensiones=EXCLUDED.dimensiones, bateria_respaldo=EXCLUDED.bateria_respaldo, resolucion=EXCLUDED.resolucion, version_software=EXCLUDED.version_software, humedad=EXCLUDED.humedad, suministro_gases=EXCLUDED.suministro_gases, contexto_operacional=EXCLUDED.contexto_operacional, funciones_equipo=EXCLUDED.funciones_equipo, acciones_preventivas=EXCLUDED.acciones_preventivas, acciones_falla=EXCLUDED.acciones_falla, fallas_funcionales=EXCLUDED.fallas_funcionales, causas_fallo=EXCLUDED.causas_fallo, efectos_fallo=EXCLUDED.efectos_fallo, efecto_entorno=EXCLUDED.efecto_entorno, observaciones=EXCLUDED.observaciones,
                                red_salud_id=EXCLUDED.red_salud_id, red_salud_nombre=EXCLUDED.red_salud_nombre, centro_salud_id=EXCLUDED.centro_salud_id, centro_salud_nombre=EXCLUDED.centro_salud_nombre, municipio_nombre=EXCLUDED.municipio_nombre, departamento_nombre=EXCLUDED.departamento_nombre;
                        """
                        cur.execute(sql_q, {**eq_data, "red_salud_id": red_id, "centro_salud_id": cen_id})
                        conn.commit()
                        cur.close()
                        conn.close()
                        print(f"[OK] Equipo {eq_data.get('id')} guardado y sincronizado con éxito en la base de datos central.")
                    except Exception as err:
                        print(f"[ERROR] Error al guardar equipo en PostgreSQL: {err}")
                        if conn:
                            conn.rollback()
                            conn.close()

            ejecutar_en_segundo_plano(_guardar_equipo_db, dict(eq_dict))
                
        # Botón de guardar fijo al final de la ventana, fuera de la zona de scroll
        btn_txt = "Actualizar Ficha de Equipo" if eq_edit else "Guardar Equipo"
        ctk.CTkButton(vent, text=btn_txt, font=ctk.CTkFont(weight="bold", size=14), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, height=45, command=guardar).pack(pady=(5, 15), padx=20, fill="x")
        vent.after(100, lambda: sf._parent_canvas.yview_moveto(0.0))

    # ========================================================
    # PREVISUALIZACIÓN DE HOJA DE VIDA Y EXCEL
    # ========================================================
    def abrir_hoja_vida_click(self, event=None, equipo_id=None):
        # Debounce de 0.5 segundos para evitar doble ejecución rápida (por ej. en pantallas táctiles)
        ahora = datetime.now()
        ultimo_click = getattr(self, "_ultimo_click_hv", None)
        if ultimo_click and (ahora - ultimo_click).total_seconds() < 0.5:
            return
        self._ultimo_click_hv = ahora

        # Evitar abrir múltiples ventanas de Ficha Técnica simultáneamente
        if hasattr(self, "window_ficha_tecnica") and self.window_ficha_tecnica.winfo_exists():
            try:
                self.window_ficha_tecnica.lift()
                self.window_ficha_tecnica.focus_force()
            except:
                pass
            return

        item_id = equipo_id
        if not item_id:
            vista_inv = self.vistas.get("Inventario")
            if vista_inv and hasattr(vista_inv, "obtener_id_seleccionado"):
                item_id = vista_inv.obtener_id_seleccionado()
            elif vista_inv and hasattr(vista_inv, "tabla_inv"):
                sel = vista_inv.tabla_inv.selection() or ([vista_inv.tabla_inv.focus()] if vista_inv.tabla_inv.focus() else [])
                if sel:
                    vals = vista_inv.tabla_inv.item(sel[0], "values")
                    item_id = vals[4] if len(vals) > 4 else vals[0]
            elif vista_inv and hasattr(vista_inv, "tabla"):
                sel = vista_inv.tabla.selection()
                if sel:
                    vals = vista_inv.tabla.item(sel[0], "values")
                    item_id = vals[4] if len(vals) > 4 else vals[0]
                    
        if not item_id:
            return

        eq_act = next((e for e in self.datos["equipos"] if str(e.get("id")) == str(item_id)), None)
        if not eq_act:
            eq_act = next((e for e in self.datos["equipos"] if str(e.get("id", "")).strip().lower() == str(item_id).strip().lower()), None)
        if not eq_act: 
            return
        
        v_hv = ctk.CTkToplevel(self)
        self.window_ficha_tecnica = v_hv
        v_hv.title(f"Ficha Técnica - {eq_act.get('nombre', '')} ({eq_act.get('id', '')})")
        v_hv.configure(fg_color=C_BG)
        self.centrar_ventana_segura(v_hv, 1100, 735)
        v_hv.transient(self)
        v_hv.lift()
        v_hv.focus_force()
        
        # Habilitar redimensionado del diálogo
        v_hv.resizable(True, True)
        
        frame_izq = ctk.CTkFrame(v_hv, fg_color="transparent")
        frame_izq.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Sanitizar el ID para evitar caracteres inválidos en rutas de archivos de Windows (ej. '|')
        id_sanitizado = "".join([c for c in str(eq_act['id']) if c.isalnum() or c in ('-', '_')]).strip()
        area_name = eq_act.get("area", "General")
        area_folder = "".join([c for c in area_name if c.isalnum() or c==' ']).strip()
        dir_ficha = os.path.join(CARPETAS["areas"], area_folder, "equipos")
        os.makedirs(dir_ficha, exist_ok=True)
        ruta_excel_guardado = os.path.join(dir_ficha, f"Ficha_Tecnica_{id_sanitizado}.xlsx")
        ruta_excel_absoluta = os.path.abspath(ruta_excel_guardado)
        
        def actualizar_excel_datos_silencioso():
            ruta_plantilla = obtener_ruta_plantilla("plantilla_ficha.xlsx")
            if not os.path.exists(ruta_plantilla):
                raise FileNotFoundError(f"No se encontró la plantilla en:\n{ruta_plantilla}")
                
            wb = openpyxl.load_workbook(ruta_plantilla)
            hoja = wb.active
            
            def escribir(celda, valor):
                try:
                    if valor is not None:
                        hoja[celda].value = valor
                except: 
                    pass
                
            def escribir_rcm(celda, texto):
                try:
                    hoja[celda].value = texto if texto else ''
                except:
                    pass
                
            # 1. Datos Generales de Identificación y Territorio
            escribir('K4', eq_act.get('red_salud_nombre', ''))
            escribir('K5', eq_act.get('centro_salud_nombre', ''))
            escribir('K8', eq_act.get('nombre', ''))
            escribir('H11', eq_act.get('area', ''))
            escribir('H12', eq_act.get('servicio', ''))
            escribir('H13', eq_act.get('marca', ''))
            escribir('H14', eq_act.get('modelo', ''))
            escribir('H15', str(eq_act.get('id', '')))
            escribir('H16', eq_act.get('procedencia', ''))
            escribir('H17', eq_act.get('fabricante', ''))
            escribir('H18', eq_act.get('garantia', ''))
            escribir('H19', eq_act.get('proveedor', ''))
            escribir('H20', eq_act.get('numero_serie', ''))
            escribir('H21', str(eq_act.get('anio_fab') or ''))
            escribir('H22', str(eq_act.get('fecha_adquisicion') or ''))

            # 2. Datos Técnicos Oficiales del Equipo
            escribir('E24', eq_act.get('voltaje', '') or '')
            escribir('G25', eq_act.get('corriente', '') or '')
            escribir('I26', eq_act.get('potencia', '') or '')
            escribir('H27', eq_act.get('vida_util', '') or eq_act.get('temperatura', '') or '')
            escribir('D28', eq_act.get('peso', '') or '')
            escribir('F29', eq_act.get('dimensiones', '') or '')
            escribir('H30', eq_act.get('bateria_respaldo', '') or eq_act.get('resolucion', '') or '')
            escribir('H31', eq_act.get('version_software', '') or eq_act.get('humedad', '') or '')
            escribir('H32', eq_act.get('suministro_gases', '') or '')

            # 3. Existencia de Repuestos con Cantidad/Stock Actual
            cat_str = f"{eq_act['nombre']} - {eq_act.get('marca', '')} - {eq_act.get('modelo', '')}"
            repuestos_equipo = [r for r in self.datos["repuestos"] if r.get("tipo_equipo") == cat_str]
            for idx, r in enumerate(repuestos_equipo[:5]):
                cell_row = 34 + idx
                nom_rep = r.get("nombre_repuesto", "")
                cant_rep = r.get("cantidad", 0)
                txt_rep = f"{nom_rep}  (Cantidad: {cant_rep})" if cant_rep is not None else nom_rep
                escribir(f'C{cell_row}', txt_rep)

            # 4. Tecnología Predominante (X)
            escribir('S25', eq_act.get('t_elec', ''))
            escribir('S27', eq_act.get('t_elco', ''))
            escribir('S29', eq_act.get('t_mec', ''))
            escribir('Z25', eq_act.get('t_hid', ''))
            escribir('Z27', eq_act.get('t_neu', ''))
            escribir('Z29', eq_act.get('t_vap', ''))

            # 5. Tipo Adquisición y Tipo de Equipo (X)
            escribir('S33', eq_act.get('a_comp', ''))
            escribir('S35', eq_act.get('a_como', ''))
            escribir('S37', eq_act.get('a_don', ''))
            escribir('Y33', eq_act.get('te_fijo', ''))
            escribir('Y35', eq_act.get('te_mov', ''))
            escribir('Y37', eq_act.get('te_por', ''))

            # 6. Categorización
            cat_data = eq_act.get("categorizacion_detalle") or []
            if isinstance(cat_data, str):
                try: cat_data = json.loads(cat_data)
                except: cat_data = []
                
            for i in range(13):
                valor = str(cat_data[i]) if i < len(cat_data) else ""
                if valor in ("1", "I"): 
                    escribir(f'AK{24+i}', 'X')
                elif valor in ("2", "II"): 
                    escribir(f'AM{24+i}', 'X')
                elif valor in ("3", "III"): 
                    escribir(f'AO{24+i}', 'X')

            # 7. Tablas RCM y Observaciones
            escribir_rcm('B41', eq_act.get('contexto_operacional'))
            escribir_rcm('L41', eq_act.get('funciones_equipo'))
            escribir_rcm('V41', eq_act.get('acciones_preventivas'))
            escribir_rcm('AE41', eq_act.get('acciones_falla'))

            escribir_rcm('B49', eq_act.get('fallas_funcionales'))
            escribir_rcm('L49', eq_act.get('causas_fallo'))
            escribir_rcm('V49', eq_act.get('efectos_fallo'))
            escribir_rcm('AE49', eq_act.get('efecto_entorno'))

            escribir_rcm('B58', eq_act.get('observaciones'))
                    
            try:
                puntajes_int = []
                for x in cat_data:
                    if str(x).isdigit():
                        puntajes_int.append(int(x))
                    elif str(x) == "I":
                        puntajes_int.append(1)
                    elif str(x) == "II":
                        puntajes_int.append(2)
                    elif str(x) == "III":
                        puntajes_int.append(3)
                    else:
                        puntajes_int.append(0)
                puntaje_total = sum(puntajes_int)
            except:
                puntaje_total = 0
                
            if puntaje_total >= 30 or eq_act.get("criticidad") == "Riesgo Alto":
                escribir('AO37', 'X')
                escribir('AB38', "3 veces al año")
            elif puntaje_total >= 20 or eq_act.get("criticidad") == "Riesgo Medio":
                escribir('AM37', 'X')
                escribir('AB38', "2 veces al año")
            else:
                escribir('AK37', 'X')
                escribir('AB38', "1 vez al año")
                
            foto_path = eq_act.get('foto')
            if foto_path:
                try:
                    foto_str = str(foto_path).strip()
                    if foto_str.startswith("data:image") or len(foto_str) > 200:
                        foto_b64 = foto_str.split(",", 1)[1] if "," in foto_str else foto_str
                        img_bytes = base64.b64decode(foto_b64)
                        img_stream = io.BytesIO(img_bytes)
                        img = ExcelImage(img_stream)
                    elif os.path.exists(foto_str):
                        img = ExcelImage(foto_str)
                    else:
                        img = None
                    if img:
                        img.width = 230
                        img.height = 205
                        hoja.add_image(img, 'Y10')
                except Exception as e:
                    print(f"Aviso: No se pudo inyectar la imagen en el Excel: {e}")

            try:
                wb.save(ruta_excel_guardado)
            except PermissionError:
                raise PermissionError("El archivo Excel está abierto o bloqueado por otra aplicación (como Microsoft Excel). Por favor, cierra Excel antes de continuar.")

        # Calcular MTTR y MTBF
        intervenciones = eq_act.get("historial_intervenciones", [])
        
        tiempos_rep = [float(i.get("tiempo_reparacion", 0) or 0) for i in intervenciones if i.get("tiempo_reparacion") is not None]
        total_intervenciones_con_tiempo = len(tiempos_rep)
        mttr_val = sum(tiempos_rep) / total_intervenciones_con_tiempo if total_intervenciones_con_tiempo > 0 else 0.0
        
        downtime_dias = 0
        correctivas_count = 0
        for i in intervenciones:
            f_rec = i.get("fecha")
            f_ent = i.get("fecha_entrega")
            if i.get("tipo") == "Correctivo":
                correctivas_count += 1
            if f_rec and f_ent:
                if isinstance(f_rec, str):
                    try: f_rec = datetime.strptime(f_rec, "%Y-%m-%d").date()
                    except: f_rec = None
                if isinstance(f_ent, str):
                    try: f_ent = datetime.strptime(f_ent, "%Y-%m-%d").date()
                    except: f_ent = None
                if f_rec and f_ent:
                    downtime_dias += max(0, (f_ent - f_rec).days)
                    
        f_ini_eq = eq_act.get("fecha_adquisicion") or eq_act.get("fecha_registro")
        if isinstance(f_ini_eq, str):
            try: f_ini_eq = datetime.strptime(f_ini_eq, "%Y-%m-%d").date()
            except: f_ini_eq = date.today()
        elif isinstance(f_ini_eq, datetime):
            f_ini_eq = f_ini_eq.date()
        elif not f_ini_eq:
            f_ini_eq = date.today()
            
        dias_totales = max(1, (date.today() - f_ini_eq).days)
        tiempo_operativo = max(0, dias_totales - downtime_dias)
        
        if correctivas_count > 0:
            mtbf_val = tiempo_operativo / correctivas_count
            mtbf_str = f"{mtbf_val:.1f} días"
        else:
            mtbf_str = f"{tiempo_operativo} días (Sin fallas)"
            
        mttr_str = f"{mttr_val:.1f} horas"

        url_base = CONFIG.get("url_base_web", "https://cmms-gamlp.onrender.com")
        enl = f"{url_base}/equipo/{eq_act['id']}"
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(enl)
        qr.make(fit=True)
        img_qr = qr.make_image(fill_color="black", back_color="white").get_image()
        v_hv.geometry("1100x880")
        
        m_info = ctk.CTkFrame(frame_izq, fg_color=C_CARD, corner_radius=12, border_width=0)
        m_info.pack(pady=(5, 10), fill="x")
        
        # --- FUNCIONES DE ACCIÓN ---
        def abrir_web():
            import webbrowser
            webbrowser.open(enl)

        def btn_ver_excel():
            try:
                actualizar_excel_datos_silencioso()
                os.startfile(ruta_excel_absoluta)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                
        def btn_descargar_pdf():
            import shutil
            ruta_pdf = filedialog.asksaveasfilename(initialdir=dir_ficha, initialfile=f"Ficha_{eq_act['id']}.pdf", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
            if ruta_pdf:
                try:
                    actualizar_excel_datos_silencioso()
                    excel = win32com.client.DispatchEx("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    
                    wb_pdf = excel.Workbooks.Open(ruta_excel_absoluta, UpdateLinks=False, ReadOnly=True)
                    ws = wb_pdf.ActiveSheet
                    
                    try:
                        ws.PageSetup.Zoom = False
                        ws.PageSetup.FitToPagesWide = 1
                        ws.PageSetup.FitToPagesTall = 1
                        ws.PageSetup.PrintArea = "$A$1:$AQ$67"
                        ws.PageSetup.CenterHorizontally = True
                        ws.PageSetup.LeftMargin = 10
                        ws.PageSetup.RightMargin = 10
                        ws.PageSetup.TopMargin = 10
                        ws.PageSetup.BottomMargin = 10
                    except:
                        pass
                    
                    ws.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
                    wb_pdf.Close(False)
                    excel.Quit()
                    os.startfile(ruta_pdf)
                except Exception as e:
                    messagebox.showerror("Error", f"Fallo al generar PDF: {e}")
                finally:
                    pythoncom.CoUninitialize()

        def abrir_manuales():
            dir_man = os.path.join(CARPETAS["manuales"], f"{eq_act['nombre']}_{id_sanitizado}")
            os.makedirs(dir_man, exist_ok=True)
            os.startfile(dir_man)
        def abrir_videos():
            dir_vid = os.path.join(CARPETAS["videos"], f"{eq_act['nombre']}_{id_sanitizado}")
            os.makedirs(dir_vid, exist_ok=True)
            os.startfile(dir_vid)
            
        # --- CREACIÓN DE WIDGETS ---
        m_info.pack_configure(padx=8, pady=(5, 8))
        
        # Grid maestro superior de 3 columnas
        m_info.grid_columnconfigure(0, weight=0, minsize=165)
        m_info.grid_columnconfigure(1, weight=1)
        m_info.grid_columnconfigure(2, weight=0, minsize=175)
        
        # 1. Columna Izquierda: Código QR
        c_img = ctk.CTkFrame(m_info, fg_color="transparent")
        c_img.grid(row=0, column=0, padx=(12, 8), pady=10, sticky="nsew")
        
        lbl_qr = ctk.CTkLabel(c_img, image=ctk.CTkImage(light_image=img_qr, size=(135, 135)), text="", cursor="hand2")
        lbl_qr.pack(pady=(0, 4))
        lbl_qr.bind("<Button-1>", lambda event: abrir_web())
        
        def descargar_qr_etiqueta():
            """Genera y descarga una etiqueta PNG con QR + nombre + Código AF + ubicación."""
            from PIL import ImageDraw, ImageFont
            try:
                qr_size = 280
                margin = 14
                line_h = 22
                n_lines = 3
                total_h = qr_size + margin + n_lines * line_h + margin
                sticker = Image.new('RGB', (qr_size, total_h), 'white')
                qr_grande = qr.make_image(fill_color='black', back_color='white').get_image()
                qr_grande = qr_grande.resize((qr_size, qr_size), Image.LANCZOS)
                sticker.paste(qr_grande, (0, 0))
                draw = ImageDraw.Draw(sticker)
                try:
                    font_nom = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", 16)
                    font_sub = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 13)
                except:
                    font_nom = font_sub = ImageFont.load_default()
                id_raw = str(eq_act['id'])
                id_mostrar = id_raw.split('|')[-1].strip() if '|' in id_raw else id_raw
                lineas = [
                    (eq_act.get('nombre', ''), font_nom),
                    (id_mostrar, font_sub),
                    (eq_act.get('area', eq_act.get('servicio', '')), font_sub),
                ]
                y = qr_size + margin
                for texto, fnt in lineas:
                    try:
                        bbox = draw.textbbox((0, 0), texto, font=fnt)
                        w = bbox[2] - bbox[0]
                    except:
                        w = len(texto) * 8
                    x = max(0, (qr_size - w) // 2)
                    draw.text((x, y), texto, fill='black', font=fnt)
                    y += line_h
                ruta_dest = filedialog.asksaveasfilename(
                    initialfile=f"QR_{id_sanitizado}.png",
                    defaultextension=".png",
                    filetypes=[("PNG", "*.png")],
                    title="Guardar Código QR"
                )
                if ruta_dest:
                    sticker.save(ruta_dest, 'PNG')
                    os.startfile(ruta_dest)
            except Exception as ex:
                messagebox.showerror("Error", f"No se pudo generar el QR: {ex}")

        btn_web_qr = ctk.CTkButton(c_img, text="🌐 Abrir Web", font=ctk.CTkFont(weight="bold", size=11), fg_color="transparent", text_color=C_BLUE, border_width=1, border_color=C_BLUE, corner_radius=8, height=28, command=abrir_web)
        btn_web_qr.pack(pady=(2, 2), fill="x")
        btn_dl_qr = ctk.CTkButton(c_img, text="📥 Descargar QR", font=ctk.CTkFont(weight="bold", size=11), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, corner_radius=8, height=28, command=descargar_qr_etiqueta)
        btn_dl_qr.pack(pady=(2, 0), fill="x")

        # 2. Columna Central: Información Completa y Equilibrada del Equipo
        i_txt = ctk.CTkFrame(m_info, fg_color="transparent")
        i_txt.grid(row=0, column=1, padx=(4, 8), pady=8, sticky="nsew")
        
        ctk.CTkLabel(i_txt, text=eq_act['nombre'], font=ctk.CTkFont(size=20, weight="bold"), text_color=C_BLUE, anchor="w").pack(fill="x", pady=(0, 3))
        
        def _crear_badge(padre, texto, color_bg, color_txt="#FFFFFF"):
            b = ctk.CTkFrame(padre, fg_color=color_bg, corner_radius=6, height=22)
            b.pack(side="left", padx=(0, 5))
            ctk.CTkLabel(b, text=f" {texto} ", font=ctk.CTkFont(size=10, weight="bold"), text_color=color_txt).pack(padx=6, pady=1)

        # --- FILA 1: Red y Centro de Salud (izq) + Botones Ficha Excel / Ficha PDF (der) ---
        f_row1 = ctk.CTkFrame(i_txt, fg_color="transparent")
        f_row1.pack(fill="x", pady=(0, 3))

        f_badges_r1 = ctk.CTkFrame(f_row1, fg_color="transparent")
        f_badges_r1.pack(side="left", fill="x", expand=True)

        red_nombre_full = eq_act.get('red_salud_nombre') or eq_act.get('red_salud') or 'Red GAMLP'
        centro_nombre_full = eq_act.get('centro_salud_nombre') or eq_act.get('centro_salud') or 'Centro de Salud'

        if "RED 1" in red_nombre_full.upper(): red_badge = "🌐 RED 1 - SUR OESTE"
        elif "RED 2" in red_nombre_full.upper(): red_badge = "🌐 RED 2 - NOR OESTE"
        elif "RED 3" in red_nombre_full.upper(): red_badge = "🌐 RED 3 - NORTE CENTRAL"
        elif "RED 4" in red_nombre_full.upper(): red_badge = "🌐 RED 4 - SAN ANTONIO"
        elif "RED 5" in red_nombre_full.upper(): red_badge = "🌐 RED 5 - SUR"
        else: red_badge = f"🌐 {red_nombre_full}"

        _crear_badge(f_badges_r1, red_badge, "#E2E8F0", C_BLUE)
        _crear_badge(f_badges_r1, f"🏥 {centro_nombre_full}", "#E2E8F0", C_TEXT)

        f_btns_r1 = ctk.CTkFrame(f_row1, fg_color="transparent")
        f_btns_r1.pack(side="right")

        btn_excel = ctk.CTkButton(f_btns_r1, text="📄 Ficha Excel", font=ctk.CTkFont(weight="bold", size=11), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, corner_radius=7, height=24, width=105, command=btn_ver_excel)
        btn_excel.pack(side="left", padx=2)
        btn_pdf = ctk.CTkButton(f_btns_r1, text="⬇ Ficha PDF", font=ctk.CTkFont(weight="bold", size=11), fg_color=C_PURPLE, hover_color="#963ECA", corner_radius=7, height=24, width=105, command=btn_descargar_pdf)
        btn_pdf.pack(side="left", padx=2)

        # --- FILA 2: AF, Servicio, Riesgo y Mttos/Año (izq) + Botones Manuales / Videos (der) ---
        f_row2 = ctk.CTkFrame(i_txt, fg_color="transparent")
        f_row2.pack(fill="x", pady=(0, 4))

        f_badges_r2 = ctk.CTkFrame(f_row2, fg_color="transparent")
        f_badges_r2.pack(side="left", fill="x", expand=True)

        _crear_badge(f_badges_r2, f"🔑 {eq_act['id']}", C_BG, C_TEXT)
        _crear_badge(f_badges_r2, f"📍 {eq_act.get('servicio', 'Servicio')}", C_BG, C_TEXT)
        
        crit_txt = str(eq_act.get('criticidad') or 'Riesgo Medio')
        crit_color = C_RED if "Alto" in crit_txt else (C_ORANGE if "Medio" in crit_txt else C_GREEN)
        _crear_badge(f_badges_r2, f"⚡ {crit_txt}", crit_color, "#FFFFFF")

        # Frecuencia de mantenimiento anual
        cat_data_score = eq_act.get("categorizacion_detalle") or []
        if isinstance(cat_data_score, str):
            try: cat_data_score = json.loads(cat_data_score)
            except: cat_data_score = []
        p_ints = []
        for x in cat_data_score:
            if str(x).isdigit(): p_ints.append(int(x))
            elif str(x) == "I": p_ints.append(1)
            elif str(x) == "II": p_ints.append(2)
            elif str(x) == "III": p_ints.append(3)
        p_total = sum(p_ints)
        if p_total >= 30: mttos_str = "3 Mttos/Año"
        elif p_total >= 20: mttos_str = "2 Mttos/Año"
        else: mttos_str = "1 Mtto/Año"
        _crear_badge(f_badges_r2, f"🗓️ {mttos_str}", "#E0F2FE", C_BLUE)

        f_btns_r2 = ctk.CTkFrame(f_row2, fg_color="transparent")
        f_btns_r2.pack(side="right")

        btn_man = ctk.CTkButton(f_btns_r2, text="📁 Manuales", font=ctk.CTkFont(weight="bold", size=11), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, corner_radius=7, height=24, width=105, command=abrir_manuales)
        btn_man.pack(side="left", padx=2)
        btn_vid = ctk.CTkButton(f_btns_r2, text="🎥 Videos", font=ctk.CTkFont(weight="bold", size=11), fg_color=C_PURPLE, hover_color="#963ECA", corner_radius=7, height=24, width=105, command=abrir_videos)
        btn_vid.pack(side="left", padx=2)

        # Grid compacto de 2 columnas con todos los datos clave
        f_grid_datos = ctk.CTkFrame(i_txt, fg_color=C_BG, corner_radius=10, border_width=1, border_color=C_BORDER)
        f_grid_datos.pack(fill="both", expand=True, pady=(2, 0), padx=0)
        
        f_grid_datos.grid_columnconfigure(0, weight=0)
        f_grid_datos.grid_columnconfigure(1, weight=1)
        f_grid_datos.grid_columnconfigure(2, weight=0)
        f_grid_datos.grid_columnconfigure(3, weight=1)

        datos_matriz = [
            ("Marca:", eq_act.get('marca', '-'), "Modelo:", eq_act.get('modelo', '-')),
            ("Código / AF:", str(eq_act['id']), "N° Serie:", str(eq_act.get('numero_serie', '-'))),
            ("Garantía:", str(eq_act.get('garantia', 'Sin Garantía')), "Año Fab.:", str(eq_act.get('anio_fab', '-'))),
            ("MTTR (Rep.):", mttr_str, "MTBF (Fallas):", mtbf_str)
        ]
        
        for r_idx, (l1, v1, l2, v2) in enumerate(datos_matriz):
            ctk.CTkLabel(f_grid_datos, text=l1, font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT, anchor="w").grid(row=r_idx, column=0, padx=(10, 4), pady=2, sticky="w")
            ctk.CTkLabel(f_grid_datos, text=str(v1), font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT, anchor="w").grid(row=r_idx, column=1, padx=(0, 10), pady=2, sticky="w")
            ctk.CTkLabel(f_grid_datos, text=l2, font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT, anchor="w").grid(row=r_idx, column=2, padx=(10, 4), pady=2, sticky="w")
            ctk.CTkLabel(f_grid_datos, text=str(v2), font=ctk.CTkFont(size=11, weight="bold"), text_color=C_TEXT, anchor="w").grid(row=r_idx, column=3, padx=(0, 10), pady=2, sticky="w")

        # 3. Columna Derecha: Foto CUADRADA del equipo (más grande y sin espacio vacío)
        c_der_col = ctk.CTkFrame(m_info, fg_color="transparent")
        c_der_col.grid(row=0, column=2, padx=(4, 12), pady=6, sticky="nsew")
        
        foto_raw = eq_act.get("foto")
        img_pil = cargar_imagen_pil(foto_raw)
        if img_pil:
            try:
                ctk_img = ctk.CTkImage(light_image=img_pil, size=(165, 165))
                lbl_foto = ctk.CTkLabel(c_der_col, image=ctk_img, text="")
                lbl_foto.pack(expand=True, fill="both")
            except Exception:
                f_placeholder = ctk.CTkFrame(c_der_col, width=165, height=165, fg_color=C_BG, corner_radius=8)
                f_placeholder.pack_propagate(False)
                f_placeholder.pack(expand=True)
                ctk.CTkLabel(f_placeholder, text="📷 Sin Imagen", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT).pack(expand=True)
        else:
            f_placeholder = ctk.CTkFrame(c_der_col, width=165, height=165, fg_color=C_BG, corner_radius=8)
            f_placeholder.pack_propagate(False)
            f_placeholder.pack(expand=True)
            ctk.CTkLabel(f_placeholder, text="📷 Sin Imagen", font=ctk.CTkFont(size=11, weight="bold"), text_color=C_SUBTEXT).pack(expand=True)

        # =========================================================================
        # SECCIÓN INFERIOR: 4 CUADRANTES DE INFORMACIÓN Y REPUESTOS
        # =========================================================================
        f_cuadrantes = ctk.CTkFrame(frame_izq, fg_color="transparent")
        f_cuadrantes.pack(fill="both", expand=True, pady=(4, 10))
        
        f_cuadrantes.grid_rowconfigure(0, weight=1, minsize=190)
        f_cuadrantes.grid_rowconfigure(1, weight=1, minsize=190)
        f_cuadrantes.grid_columnconfigure(0, weight=1)
        f_cuadrantes.grid_columnconfigure(1, weight=1)
        
        # --- CUADRANTE 1: Historial de Mantenimientos ---
        c1 = ctk.CTkFrame(f_cuadrantes, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        c1.grid(row=0, column=0, padx=6, pady=6, sticky="nsew")
        
        f_c1_header = ctk.CTkFrame(c1, fg_color="transparent")
        f_c1_header.pack(fill="x", padx=10, pady=6)
        ctk.CTkLabel(f_c1_header, text="📋 Historial de Intervenciones", font=ctk.CTkFont(size=13, weight="bold"), text_color=C_BLUE).pack(side="left")
        def ir_a_mantenimientos():
            v_hv.destroy()
            self.mostrar_vista("Historial")
        ctk.CTkButton(f_c1_header, text="Ver más →", font=ctk.CTkFont(size=11), fg_color="transparent", text_color=C_BLUE, hover_color=C_BG, height=20, width=65, command=ir_a_mantenimientos).pack(side="right")
        
        tab_c1 = ttk.Treeview(c1, columns=("Fecha", "Tipo", "Detalle"), show="headings", height=4)
        for c in ("Fecha", "Tipo", "Detalle"):
            tab_c1.heading(c, text=c)
            tab_c1.column(c, anchor="w" if c=="Detalle" else "center", width=85 if c!="Detalle" else 180)
        tab_c1.pack(padx=8, pady=(0, 8), fill="both", expand=True)
        
        intervenciones_lista = eq_act.get("historial_intervenciones", [])
        for inter in intervenciones_lista:
            tab_c1.insert("", "end", values=(inter["fecha"], inter["tipo"], inter.get("trabajo", inter.get("detalle", ""))))
        if not intervenciones_lista:
            tab_c1.insert("", "end", values=("-", "Sin registros", "No se han realizado intervenciones"))
            
        # --- CUADRANTE 2: Siguientes Mantenimientos Programados ---
        c2 = ctk.CTkFrame(f_cuadrantes, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        c2.grid(row=0, column=1, padx=6, pady=6, sticky="nsew")
        
        f_c2_header = ctk.CTkFrame(c2, fg_color="transparent")
        f_c2_header.pack(fill="x", padx=10, pady=6)
        ctk.CTkLabel(f_c2_header, text="📅 Siguientes Mantenimientos", font=ctk.CTkFont(size=13, weight="bold"), text_color=C_BLUE).pack(side="left")
        def ir_a_cronograma():
            v_hv.destroy()
            self.mostrar_vista("Cronograma")
        ctk.CTkButton(f_c2_header, text="Ver más →", font=ctk.CTkFont(size=11), fg_color="transparent", text_color=C_BLUE, hover_color=C_BG, height=20, width=65, command=ir_a_cronograma).pack(side="right")
        
        tab_c2 = ttk.Treeview(c2, columns=("Numero", "Fecha Programada", "Estado"), show="headings", height=4)
        tab_c2.heading("Numero", text="Nº")
        tab_c2.heading("Fecha Programada", text="Fecha Programada")
        tab_c2.heading("Estado", text="Estado")
        tab_c2.column("Numero", anchor="center", width=35)
        tab_c2.column("Fecha Programada", anchor="center", width=130)
        tab_c2.column("Estado", anchor="center", width=90)
        tab_c2.pack(padx=8, pady=(0, 8), fill="both", expand=True)
        
        siguientes_mantenimientos = calcular_proximos_mantenimientos(eq_act, cantidad=3, hoy=self.hoy)
        for idx, f_prog in enumerate(siguientes_mantenimientos):
            estado_lbl = "Vencido" if f_prog < date.today() else "Planificado"
            tab_c2.insert("", "end", values=(f"{idx+1}º", f_prog.strftime("%d / %m / %Y"), estado_lbl))
        if not siguientes_mantenimientos:
            tab_c2.insert("", "end", values=("-", "No programado", "Sin proyección"))

        # --- OBTENER REPUESTOS COMPATIBLES CON ESTE EQUIPO ---
        eq_nom_low = str(eq_act.get('nombre', '')).strip().lower()
        cat_full_low = f"{eq_act.get('nombre', '')} - {eq_act.get('marca', '')} - {eq_act.get('modelo', '')}".strip().lower()
        
        repuestos_compatibles = []
        for r in self.datos.get("repuestos", []):
            r_tipo_low = str(r.get("tipo_equipo", "")).strip().lower()
            if (r_tipo_low == cat_full_low or 
                r_tipo_low == eq_nom_low or 
                (eq_nom_low and eq_nom_low in r_tipo_low) or 
                (r_tipo_low and r_tipo_low in cat_full_low) or
                r_tipo_low == "general / multiuso" or
                r_tipo_low == "vacio"):
                repuestos_compatibles.append(r)
                
        rep_en_stock = [r for r in repuestos_compatibles if str(r.get("estado_disponibilidad", "En Stock")).strip().lower() != "requerido" and int(r.get("cantidad", 0) or 0) > 0]
        rep_requeridos = [r for r in repuestos_compatibles if str(r.get("estado_disponibilidad", "En Stock")).strip().lower() == "requerido"]

        # --- CUADRANTE 3: Repuestos en Stock Disponibles ---
        c3 = ctk.CTkFrame(f_cuadrantes, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        c3.grid(row=1, column=0, padx=6, pady=6, sticky="nsew")
        
        f_c3_header = ctk.CTkFrame(c3, fg_color="transparent")
        f_c3_header.pack(fill="x", padx=10, pady=6)
        ctk.CTkLabel(f_c3_header, text=f"📦 Repuestos en Stock ({len(rep_en_stock)})", font=ctk.CTkFont(size=13, weight="bold"), text_color=C_GREEN).pack(side="left")
        
        def ir_a_repuestos_tab():
            v_hv.destroy()
            self.mostrar_vista("Repuestos")
        ctk.CTkButton(f_c3_header, text="Administrar →", font=ctk.CTkFont(size=11), fg_color="transparent", text_color=C_BLUE, hover_color=C_BG, height=20, width=75, command=ir_a_repuestos_tab).pack(side="right")
        
        tab_c3 = ttk.Treeview(c3, columns=("Repuesto", "Modelo / P/N", "Stock", "Costo"), show="headings", height=4)
        tab_c3.heading("Repuesto", text="Repuesto")
        tab_c3.heading("Modelo / P/N", text="Modelo / P/N")
        tab_c3.heading("Stock", text="Stock Disp.")
        tab_c3.heading("Costo", text="Costo (Bs.)")
        tab_c3.column("Repuesto", anchor="w", width=140)
        tab_c3.column("Modelo / P/N", anchor="center", width=90)
        tab_c3.column("Stock", anchor="center", width=65)
        tab_c3.column("Costo", anchor="center", width=70)
        tab_c3.pack(padx=8, pady=(0, 8), fill="both", expand=True)
        
        for r in rep_en_stock:
            c_val = float(r.get("costo", 0) or 0)
            c_str = f"{c_val:.2f}" if c_val > 0 else "-"
            tab_c3.insert("", "end", values=(r.get("nombre_repuesto", ""), r.get("modelo_parte", "-") or "-", r.get("cantidad", 0), c_str))
        if not rep_en_stock:
            tab_c3.insert("", "end", values=("Sin repuestos en stock", "-", "0", "-"))

        # --- CUADRANTE 4: Repuestos Requeridos (Necesarios) ---
        c4 = ctk.CTkFrame(f_cuadrantes, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        c4.grid(row=1, column=1, padx=6, pady=6, sticky="nsew")
        
        f_c4_header = ctk.CTkFrame(c4, fg_color="transparent")
        f_c4_header.pack(fill="x", padx=10, pady=6)
        ctk.CTkLabel(f_c4_header, text=f"⚠️ Repuestos Requeridos ({len(rep_requeridos)})", font=ctk.CTkFont(size=13, weight="bold"), text_color=C_ORANGE).pack(side="left")
        
        def solicitar_repuesto_equipo():
            v_hv.destroy()
            self.mostrar_vista("Repuestos")
            if "Repuestos" in self.vistas:
                self.vistas["Repuestos"].abrir_formulario_repuesto(estado_inicial="Requerido")
                
        ctk.CTkButton(f_c4_header, text="✚ Solicitar", font=ctk.CTkFont(size=11, weight="bold"), fg_color=C_ORANGE, hover_color="#D97706", height=20, width=70, command=solicitar_repuesto_equipo).pack(side="right")
        
        tab_c4 = ttk.Treeview(c4, columns=("Repuesto", "Modelo / P/N", "Cant", "Motivo"), show="headings", height=4)
        tab_c4.heading("Repuesto", text="Repuesto Requerido")
        tab_c4.heading("Modelo / P/N", text="Modelo / P/N")
        tab_c4.heading("Cant", text="Cant. Req.")
        tab_c4.heading("Motivo", text="Motivo / Obs.")
        tab_c4.column("Repuesto", anchor="w", width=130)
        tab_c4.column("Modelo / P/N", anchor="center", width=85)
        tab_c4.column("Cant", anchor="center", width=65)
        tab_c4.column("Motivo", anchor="w", width=120)
        tab_c4.pack(padx=8, pady=(0, 8), fill="both", expand=True)
        
        for r in rep_requeridos:
            tab_c4.insert("", "end", values=(r.get("nombre_repuesto", ""), r.get("modelo_parte", "-") or "-", r.get("cantidad", 0), r.get("observaciones", "-") or "-"))
        if not rep_requeridos:
            tab_c4.insert("", "end", values=("Sin requerimientos pendientes", "-", "0", "-"))



    def verificar_autorizacion_jefe(self, password_plano):
        from auth import verificar_password, MASTER_PASS
        if password_plano == MASTER_PASS:
            return True
        conn = obtener_conexion()
        if not conn: 
            return False
        try:
            cur = conn.cursor()
            cur.execute("SELECT password_hash FROM usuarios WHERE rol = 'jefe' AND activo = TRUE")
            rows = cur.fetchall()
            cur.close()
            conn.close()
            for r in rows:
                if verificar_password(password_plano, r[0]):
                    return True
        except:
            pass
        return False


    def verificar_y_ejecutar_backup_auto(self):
        import os
        import sys
        import json
        import shutil
        import threading
        from datetime import datetime, timedelta
        from database import crear_backup_json
        
        def tarea_backup():
            try:
                # 1. Directorios de destino (oficial y local/interno)
                dirs_respaldos = [CARPETAS["respaldos"]]
                dir_local = os.path.join(os.path.dirname(os.path.abspath(__file__)), "respaldos")
                if dir_local not in dirs_respaldos:
                    dirs_respaldos.append(dir_local)
                if hasattr(sys, 'frozen'):
                    dir_exe = os.path.dirname(sys.executable)
                    dirs_respaldos.append(os.path.join(dir_exe, "_internal", "respaldos"))
                    dirs_respaldos.append(os.path.join(dir_exe, "respaldos"))

                for d in dirs_respaldos:
                    try: os.makedirs(d, exist_ok=True)
                    except: pass
                
                dir_principal = CARPETAS["respaldos"]
                meta_path = os.path.join(dir_principal, "metadata.json")
                necesita_backup = False
                ultimo_fecha = None
                
                if os.path.exists(meta_path):
                    try:
                        with open(meta_path, "r", encoding="utf-8") as f:
                            meta = json.load(f)
                            fecha_str = meta.get("ultimo_respaldo_auto")
                            if fecha_str:
                                ultimo_fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    except Exception as e:
                        print("Error al leer metadata.json:", e)
                        necesita_backup = True
                else:
                    necesita_backup = True
                
                if not necesita_backup and ultimo_fecha:
                    if self.hoy - ultimo_fecha >= timedelta(days=7):
                        necesita_backup = True
                
                if necesita_backup:
                    hoy_str = self.hoy.strftime("%Y-%m-%d")
                    nombre_archivo = f"respaldo_automatico_{hoy_str}.json"
                    destino_oficial = os.path.join(dir_principal, nombre_archivo)
                    
                    if crear_backup_json(destino_oficial):
                        with open(meta_path, "w", encoding="utf-8") as f:
                            json.dump({"ultimo_respaldo_auto": hoy_str}, f, indent=2)
                        
                        # Replicar a todas las carpetas locales/internas secundarias
                        for d_sec in dirs_respaldos:
                            if d_sec != dir_principal and os.path.exists(d_sec):
                                try:
                                    shutil.copy2(destino_oficial, os.path.join(d_sec, nombre_archivo))
                                    shutil.copy2(meta_path, os.path.join(d_sec, "metadata.json"))
                                except:
                                    pass
                        print(f"[OK] Respaldo automatico semanal creado en ambas ubicaciones: {nombre_archivo}")
            except Exception as ex:
                print("Error en hilo de backup automático:", ex)
                
        threading.Thread(target=tarea_backup, daemon=True).start()

    def al_cerrar_aplicacion(self):
        self._ejecutando = False
        import os
        import sys
        import shutil
        from database import crear_backup_json
        try:
            dirs_respaldos = [CARPETAS["respaldos"]]
            dir_local = os.path.join(os.path.dirname(os.path.abspath(__file__)), "respaldos")
            if dir_local not in dirs_respaldos:
                dirs_respaldos.append(dir_local)
            if hasattr(sys, 'frozen'):
                dir_exe = os.path.dirname(sys.executable)
                dirs_respaldos.append(os.path.join(dir_exe, "_internal", "respaldos"))
                dirs_respaldos.append(os.path.join(dir_exe, "respaldos"))

            for d in dirs_respaldos:
                try: os.makedirs(d, exist_ok=True)
                except: pass
            
            hoy_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            nombre_archivo = f"respaldo_cierre_{hoy_str}.json"
            destino_oficial = os.path.join(CARPETAS["respaldos"], nombre_archivo)
            
            if crear_backup_json(destino_oficial):
                for d_sec in dirs_respaldos:
                    if d_sec != CARPETAS["respaldos"] and os.path.exists(d_sec):
                        try:
                            shutil.copy2(destino_oficial, os.path.join(d_sec, nombre_archivo))
                        except:
                            pass
                print(f"[OK] Respaldo de cierre creado en ambas ubicaciones: {nombre_archivo}")
        except Exception as e:
            print("Error al crear backup al cerrar:", e)
        finally:
            try:
                self.destroy()
            except:
                pass
            os._exit(0)



    def al_redimensionar(self, event):
        if event.widget != self:
            return
        
        h = self.winfo_height()
        nuevo_modo = "compact" if h < 780 else "spacious"
        
        if self.sidebar_mode != nuevo_modo:
            self.sidebar_mode = nuevo_modo
            self.aplicar_modo_sidebar(nuevo_modo)

    def aplicar_modo_sidebar(self, modo):
        if modo == "compact":
            self.lbl_logo.configure(font=ctk.CTkFont(size=16, weight="bold"))
            self.top_sidebar.pack_configure(pady=(10, 2))
            if hasattr(self, 'lbl_estado_conexion'):
                self.lbl_estado_conexion.pack_configure(pady=(0, 2))
            self.lbl_name.pack_configure(pady=(0, 6))
            self.btn_alertas.configure(height=34)
            self.btn_alertas.pack_configure(pady=(0, 8))
            self.btn_bottom_mantenimiento.configure(height=36, font=ctk.CTkFont(size=12, weight="bold"))
            self.bottom_sidebar.pack_configure(pady=8)
            
            for btn in self.botones_nav:
                btn.configure(height=34, font=ctk.CTkFont(size=12, weight="bold"))
                btn.pack_configure(pady=1)
        else:
            self.lbl_logo.configure(font=ctk.CTkFont(size=20, weight="bold"))
            self.top_sidebar.pack_configure(pady=(20, 5))
            if hasattr(self, 'lbl_estado_conexion'):
                self.lbl_estado_conexion.pack_configure(pady=(0, 2))
            self.lbl_name.pack_configure(pady=(0, 10))
            self.btn_alertas.configure(height=38)
            self.btn_alertas.pack_configure(pady=(0, 10))
            self.btn_bottom_mantenimiento.configure(height=42, font=ctk.CTkFont(size=13, weight="bold"))
            self.bottom_sidebar.pack_configure(pady=15)
            
            for btn in self.botones_nav:
                btn.configure(height=40, font=ctk.CTkFont(size=14, weight="bold"))
                btn.pack_configure(pady=2)

    # ========================================================
    # FORMULARIO: MANTENIMIENTO MAESTRO 
    # ========================================================
    def modulo_mantenimiento(self):
        v = ctk.CTkToplevel(self)
        v.title("Mantenimiento y Hoja de Trabajo")
        self.centrar_ventana_segura(v, 750, 700)
        v.transient(self)
        v.grab_set()
        v.configure(fg_color=C_BG)
        
        # Declaración previa de variables para evitar errores de referencia
        var_tipo = ctk.StringVar(value="Preventivo")
        var_cond = ctk.StringVar(value="Óptimo")
        var_est = ctk.StringVar(value="Bueno")
        var_usar_repuesto = ctk.BooleanVar(value=False)
        ruta_ht_excel_act = ctk.StringVar(value="")
        
        ctk.CTkLabel(v, text="Registrar Intervención y Generar Hoja", font=ctk.CTkFont(size=20, weight="bold"), text_color=C_TEXT).pack(pady=15)
        
        sf = ctk.CTkScrollableFrame(v, fg_color=C_CARD, corner_radius=12)
        sf.pack(fill="both", expand=True, padx=20, pady=10)
        
        # --- LISTA DE MANTENIMIENTOS PENDIENTES DEL CRONOGRAMA ---
        from datetime import date, timedelta
        from dateutil.relativedelta import relativedelta
        
        nombres_meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        
        pendientes_data = []
        opciones_pendientes = ["Ninguno / Mantenimiento Aparte"]
        
        for eq in self.datos.get("equipos", []):
            if eq.get("estado") == "Baja":
                continue
            
            proximos = calcular_proximos_mantenimientos(eq, cantidad=1, hoy=self.hoy)
            if not proximos:
                continue
            
            f_prox = proximos[0]
            fecha_inicio_mes = date(f_prox.year, f_prox.month, 1)
            fecha_activacion = fecha_inicio_mes - timedelta(days=5)
            
            if self.hoy >= fecha_activacion:
                limit_date = date(f_prox.year, f_prox.month, 1) + relativedelta(months=+1, day=5)
                es_v = self.hoy > limit_date
                if es_v:
                    lbl_op = f"⚠️ [VENCIDO] {eq['id']} - {eq['nombre']} (Venció: {f_prox.strftime('%Y-%m-%d')} - {nombres_meses[f_prox.month - 1]})"
                else:
                    lbl_op = f"{eq['id']} - {eq['nombre']} (Vence: {f_prox.strftime('%Y-%m-%d')} - {nombres_meses[f_prox.month - 1]})"
                opciones_pendientes.append(lbl_op)
                pendientes_data.append({
                    "label": lbl_op,
                    "eq_id": eq["id"],
                    "eq_nombre": eq["nombre"],
                    "f_prox": f_prox,
                    "es_vencido": es_v
                })
                
        ctk.CTkLabel(sf, text="Mantenimiento Pendiente (Cronograma):", font=ctk.CTkFont(weight="bold", size=12), text_color=C_BLUE).pack(anchor="w", padx=10, pady=(10,0))
        c_pendiente = ctk.CTkComboBox(sf, values=opciones_pendientes, width=600)
        c_pendiente.pack(pady=5, padx=10)
        c_pendiente.set("Ninguno / Mantenimiento Aparte")
        
        ctk.CTkLabel(sf, text="Seleccione el Equipo Médico:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10,0))
        
        noms = [f"{eq['id']} - {eq['nombre']}" for eq in self.datos.get("equipos", [])]
        c_eq = ctk.CTkComboBox(sf, values=noms if noms else ["Vacío"], width=600)
        c_eq.pack(pady=5, padx=10)
        habilitar_autocompletado(c_eq, noms)
        
        def al_seleccionar_pendiente(val_sel):
            if val_sel == "Ninguno / Mantenimiento Aparte":
                return
            match = next((p for p in pendientes_data if p["label"] == val_sel), None)
            if match:
                eq_str = f"{match['eq_id']} - {match['eq_nombre']}"
                if eq_str in noms:
                    c_eq.set(eq_str)
                    try:
                        actualizar_filtro_repuestos()
                    except:
                        pass
                var_tipo.set("Preventivo")
                
        c_pendiente.configure(command=al_seleccionar_pendiente)
        
        f_extra = ctk.CTkFrame(sf, fg_color="transparent")
        f_extra.pack(fill="x", pady=5)
        
        ctk.CTkLabel(f_extra, text="Tipo:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_tipo_eq = ctk.CTkComboBox(f_extra, values=["1", "2", "3"], width=150)
        e_tipo_eq.pack(side="left", padx=5)
        
        f_repuestos = ctk.CTkFrame(sf, fg_color="transparent")
        f_repuestos.pack(fill="x", pady=5)
        
        def toggle_repuestos():
            if var_usar_repuesto.get():
                c_repuestos.configure(state="normal")
                e_cant_repuesto.configure(state="normal")
            else:
                c_repuestos.configure(state="disabled")
                e_cant_repuesto.configure(state="disabled")

        chk_repuesto = ctk.CTkCheckBox(f_repuestos, text="¿Se usó repuesto?", variable=var_usar_repuesto, command=toggle_repuestos, font=ctk.CTkFont(weight="bold"))
        chk_repuesto.pack(side="left", padx=10)
        
        repuestos_opciones = [f"{r.get('nombre_repuesto')} - {r.get('tipo_equipo')} (Disp: {r.get('cantidad')})" for r in self.datos.get("repuestos", [])]
        c_repuestos = ctk.CTkComboBox(f_repuestos, values=repuestos_opciones if repuestos_opciones else ["No hay repuestos"], width=300)
        c_repuestos.pack(side="left", padx=5)
        c_repuestos.configure(state="disabled")
        habilitar_autocompletado(c_repuestos, repuestos_opciones)
        
        def actualizar_filtro_repuestos(val_sel=None):
            eq_sel = c_eq.get().strip()
            eq_id = eq_sel.split(" - ", 1)[0].strip() if " - " in eq_sel else eq_sel
            
            # Buscar el equipo seleccionado en los datos en memoria
            eq_match = next((eq for eq in self.datos.get("equipos", []) if str(eq.get("id")).strip() == eq_id), None)
            
            repuestos_compatibles = []
            if eq_match:
                eq_nom = str(eq_match.get("nombre", "")).strip().lower()
                cat_full = f"{eq_match.get('nombre', '')} - {eq_match.get('marca', '')} - {eq_match.get('modelo', '')}".strip().lower()

                for r in self.datos.get("repuestos", []):
                    # Solo repuestos en stock y con cantidad disponible
                    if str(r.get("estado_disponibilidad", "En Stock")).strip().lower() == "requerido" or int(r.get("cantidad", 0)) <= 0:
                        continue
                    r_tipo = str(r.get("tipo_equipo", "")).strip().lower()
                    if (r_tipo == cat_full or 
                        r_tipo == eq_nom or 
                        (eq_nom and eq_nom in r_tipo) or 
                        (r_tipo and r_tipo in cat_full)):
                        repuestos_compatibles.append(r)
            
            if repuestos_compatibles:
                ops_filtradas = [
                    f"{r.get('nombre_repuesto')} - {r.get('tipo_equipo')} (Disp: {r.get('cantidad')})"
                    for r in repuestos_compatibles
                ]
            else:
                todos_rep_stock = [r for r in self.datos.get("repuestos", []) if str(r.get("estado_disponibilidad", "En Stock")).strip().lower() != "requerido" and int(r.get("cantidad", 0)) > 0]
                if todos_rep_stock:
                    ops_filtradas = [
                        f"{r.get('nombre_repuesto')} - {r.get('tipo_equipo')} (Disp: {r.get('cantidad')})"
                        for r in todos_rep_stock
                    ]
                else:
                    ops_filtradas = ["No hay repuestos disponibles en stock"]
                
            c_repuestos.configure(values=ops_filtradas)
            c_repuestos.set(ops_filtradas[0])
            habilitar_autocompletado(c_repuestos, ops_filtradas)
            
        c_eq.configure(command=actualizar_filtro_repuestos)
        actualizar_filtro_repuestos()
        
        ctk.CTkLabel(f_repuestos, text="Cant:").pack(side="left", padx=5)
        e_cant_repuesto = ctk.CTkEntry(f_repuestos, width=50)
        e_cant_repuesto.insert(0, "1")
        e_cant_repuesto.pack(side="left", padx=5)
        e_cant_repuesto.configure(state="disabled")

        f_fechas = ctk.CTkFrame(sf, fg_color="transparent")
        f_fechas.pack(fill="x", pady=5)
        
        ctk.CTkLabel(f_fechas, text="Fecha Recepción:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_fecha_rec = ctk.CTkEntry(f_fechas, width=120)
        e_fecha_rec.insert(0, datetime.now().strftime("%d / %m / %Y"))
        e_fecha_rec.pack(side="left", padx=5)
        ctk.CTkButton(f_fechas, text="Usar Actual", width=80, fg_color=C_BLUE, command=lambda: [e_fecha_rec.delete(0, "end"), e_fecha_rec.insert(0, datetime.now().strftime("%d / %m / %Y"))]).pack(side="left", padx=5)
        
        ctk.CTkLabel(f_fechas, text="Fecha Entrega:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_fecha_ent = ctk.CTkEntry(f_fechas, width=120)
        e_fecha_ent.insert(0, datetime.now().strftime("%d / %m / %Y"))
        e_fecha_ent.pack(side="left", padx=5)
        ctk.CTkButton(f_fechas, text="Usar Actual", width=80, fg_color=C_BLUE, command=lambda: [e_fecha_ent.delete(0, "end"), e_fecha_ent.insert(0, datetime.now().strftime("%d / %m / %Y"))]).pack(side="left", padx=5)

        f_tiempos = ctk.CTkFrame(sf, fg_color="transparent")
        f_tiempos.pack(fill="x", pady=5)
        ctk.CTkLabel(f_tiempos, text="Tiempo de Reparación (Horas):", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        e_tiempo_rep = ctk.CTkEntry(f_tiempos, width=80)
        e_tiempo_rep.insert(0, "0")
        e_tiempo_rep.pack(side="left", padx=5)

        f_opc = ctk.CTkFrame(sf, fg_color="transparent")
        f_opc.pack(fill="x", pady=10)
        
        f_tipo = ctk.CTkFrame(f_opc, fg_color="transparent")
        f_tipo.pack(side="left", expand=True, fill="both", padx=10)
        ctk.CTkLabel(f_tipo, text="Tipo Mantenimiento:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        ctk.CTkRadioButton(f_tipo, text="Preventivo", variable=var_tipo, value="Preventivo").pack(anchor="w", pady=5)
        ctk.CTkRadioButton(f_tipo, text="Correctivo", variable=var_tipo, value="Correctivo").pack(anchor="w", pady=5)

        f_cond = ctk.CTkFrame(f_opc, fg_color="transparent")
        f_cond.pack(side="left", expand=True, fill="both", padx=10)
        ctk.CTkLabel(f_cond, text="Condición Encontrada:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        for op in ["Óptimo", "Aceptable", "Crítica", "Inoperante", "F/Servicio"]:
            ctk.CTkRadioButton(f_cond, text=op, variable=var_cond, value=op).pack(anchor="w", pady=2)

        f_est = ctk.CTkFrame(f_opc, fg_color="transparent")
        f_est.pack(side="left", expand=True, fill="both", padx=10)
        ctk.CTkLabel(f_est, text="Estado del Equipo:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        for op in ["Óptimo", "Bueno", "Regular", "Malo", "Obsoleto"]:
            ctk.CTkRadioButton(f_est, text=op, variable=var_est, value=op).pack(anchor="w", pady=2)

        ctk.CTkLabel(sf, text="Deficiencia Encontrada:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(15,0))
        e_def = ctk.CTkTextbox(sf, width=600, height=60, fg_color=C_BG, corner_radius=8)
        e_def.pack(padx=10, pady=5)
        
        ctk.CTkLabel(sf, text="D. Trabajo Realizado:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10,0))
        e_trab = ctk.CTkTextbox(sf, width=600, height=80, fg_color=C_BG, corner_radius=8)
        e_trab.pack(padx=10, pady=5)
        
        ctk.CTkLabel(sf, text="E. Observaciones y Recomendaciones:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10,0))
        e_obs = ctk.CTkTextbox(sf, width=600, height=80, fg_color=C_BG, corner_radius=8)
        e_obs.pack(padx=10, pady=5)

        f_acciones = ctk.CTkFrame(sf, fg_color="transparent")
        ruta_ht_excel_act = ctk.StringVar()

        def guardar():
            sel = c_eq.get()
            if not sel or "Vacío" in sel:
                messagebox.showwarning("Aviso", "Seleccione un equipo.")
                return
            if not var_tipo.get() or not var_cond.get() or not var_est.get():
                messagebox.showwarning("Dato Obligatorio", "Debe marcar todos los botones de opciones.")
                return
            
            id_e = sel.split(" - ")[0]
            eq_data = next((eq for eq in self.datos["equipos"] if eq["id"] == id_e), None)
            
            if not eq_data: 
                return

            # Validar si el mantenimiento es preventivo y está vencido (más del 5 del mes siguiente al programado)
            if var_tipo.get() == "Preventivo":
                crit_eq = str(eq_data.get("criticidad") or "Riesgo Medio")
                meses_eq = 3 if "Alto" in crit_eq else (4 if "Medio" in crit_eq else 6)
                f_reg_eq = eq_data.get("fecha_adquisicion") or eq_data.get("fecha_registro", self.hoy)
                if isinstance(f_reg_eq, datetime):
                    f_reg_eq = f_reg_eq.date()
                elif isinstance(f_reg_eq, str):
                    try:
                        f_reg_eq = datetime.strptime(f_reg_eq, "%Y-%m-%d").date()
                    except:
                        f_reg_eq = eq_data.get("fecha_registro", self.hoy)
                        if isinstance(f_reg_eq, str):
                            try:
                                f_reg_eq = datetime.strptime(f_reg_eq, "%Y-%m-%d").date()
                            except:
                                f_reg_eq = self.hoy
                        elif isinstance(f_reg_eq, datetime):
                            f_reg_eq = f_reg_eq.date()
                
                if eq_data.get("garantia") == "Con Garantía" and eq_data.get("fecha_vencimiento_garantia"):
                    f_venc_g = eq_data.get("fecha_vencimiento_garantia")
                    if isinstance(f_venc_g, str):
                        try: f_venc_g = datetime.strptime(f_venc_g, "%Y-%m-%d").date()
                        except: f_venc_g = None
                    elif isinstance(f_venc_g, datetime):
                        f_venc_g = f_venc_g.date()
                    if f_venc_g:
                        f_reg_eq = f_venc_g + relativedelta(days=+1)

                # Buscar el primer slot no completado como f_prox_eq
                f_prox_eq = None
                f_check_eq = f_reg_eq
                iter_count_eq = 0
                while iter_count_eq < 50:
                    iter_count_eq += 1
                    f_check_eq = f_check_eq + relativedelta(months=+meses_eq)
                    
                    slot_is_completed = False
                    for m in eq_data.get("historial_intervenciones", []):
                        if m["tipo"] == "Preventivo":
                            m_prog = m.get("fecha_programada")
                            if m_prog:
                                if isinstance(m_prog, str):
                                    try: m_prog_d = datetime.strptime(m_prog, "%Y-%m-%d").date()
                                    except: m_prog_d = None
                                else:
                                    m_prog_d = m_prog
                                if m_prog_d == f_check_eq:
                                    slot_is_completed = True
                                    break
                            else:
                                m_f = m.get("fecha")
                                if isinstance(m_f, str):
                                    try: m_f_d = datetime.strptime(m_f, "%Y-%m-%d").date()
                                    except: m_f_d = None
                                else:
                                    m_f_d = m_f
                                if m_f_d and m_f_d.year == f_check_eq.year and m_f_d.month == f_check_eq.month:
                                    slot_is_completed = True
                                    break
                    if not slot_is_completed:
                        f_prox_eq = f_check_eq
                        break
                
                if not f_prox_eq:
                    f_prox_eq = f_reg_eq + relativedelta(months=+meses_eq)
                
                limit_date_eq = date(f_prox_eq.year, f_prox_eq.month, 1) + relativedelta(months=+1, day=5)
                if self.hoy > limit_date_eq:
                    # Mantenimiento vencido detectado
                    if not self.es_jefe:
                        from customtkinter import CTkInputDialog
                        dialog = CTkInputDialog(text=f"El mantenimiento preventivo de este equipo venció el {limit_date_eq.strftime('%d/%m/%Y')}.\nSe requiere autorización del Jefe para registrarlo.\nIngrese la contraseña del Jefe/Administrador:", title="Autorización Requerida")
                        pwd_val = dialog.get_input()
                        if not pwd_val:
                            messagebox.showwarning("Acceso Denegado", "Operación cancelada. El registro de mantenimientos vencidos requiere autorización.")
                            return
                        
                        if not self.verificar_autorizacion_jefe(pwd_val):
                            messagebox.showerror("Error de Autorización", "Contraseña del Jefe incorrecta. No se autorizó el registro.")
                            return
                        else:
                            messagebox.showinfo("Autorizado", "Registro de mantenimiento vencido autorizado correctamente por el Jefe.")

            txt_def = e_def.get("1.0", "end-1c").strip()
            txt_trab = e_trab.get("1.0", "end-1c").strip()
            txt_obs = e_obs.get("1.0", "end-1c").strip()
            
            r_usado = var_usar_repuesto.get()
            r_nombre = ""
            r_cant = 0
            rep_rec = None
            if r_usado:
                r_sel = c_repuestos.get()
                if not r_sel or r_sel.startswith("No hay repuestos"):
                    messagebox.showwarning("Aviso", "Seleccione un repuesto válido de la lista.")
                    return
                rep_rec = None
                for r in self.datos.get("repuestos", []):
                    disp_str = f"{r.get('nombre_repuesto')} - {r.get('tipo_equipo')} (Disp: {r.get('cantidad')})"
                    if disp_str == r_sel:
                        rep_rec = r
                        break
                    elif r_sel.startswith(r.get("nombre_repuesto", "")):
                        rep_rec = r
                        break
                if not rep_rec:
                    messagebox.showerror("Error", "No se encontró el repuesto en el inventario.")
                    return
                r_nombre = rep_rec.get("nombre_repuesto", "")

                try:
                    r_cant = int(e_cant_repuesto.get().strip())
                    if r_cant <= 0:
                        raise ValueError()
                except:
                    messagebox.showwarning("Dato Inválido", "La cantidad de repuestos debe ser un número entero positivo.")
                    return
                
                if r_cant > rep_rec.get("cantidad", 0):
                    messagebox.showerror("Stock Insuficiente", f"No hay suficiente stock del repuesto. Disponible: {rep_rec.get('cantidad')}.")
                    return

            try: 
                f_recepcion_iso = datetime.strptime(e_fecha_rec.get().strip(), "%d / %m / %Y").strftime("%Y-%m-%d")
            except: 
                f_recepcion_iso = datetime.now().strftime("%Y-%m-%d")

            try: 
                f_entrega_iso = datetime.strptime(e_fecha_ent.get().strip(), "%d / %m / %Y").strftime("%Y-%m-%d")
            except: 
                f_entrega_iso = datetime.now().strftime("%Y-%m-%d")

            # Determinar fecha_programada para el preventivo
            f_prog_val = None
            if var_tipo.get() == "Preventivo":
                sel_p = c_pendiente.get()
                if sel_p != "Ninguno / Mantenimiento Aparte":
                    match = next((p for p in pendientes_data if p["label"] == sel_p), None)
                    if match:
                        f_prog_val = match["f_prox"].strftime("%Y-%m-%d")
                if not f_prog_val:
                    f_prog_val = f_prox_eq.strftime("%Y-%m-%d") if 'f_prox_eq' in locals() else None

            try:
                tiempo_rep_val = float(e_tiempo_rep.get().strip().replace(",", "."))
                if tiempo_rep_val < 0:
                    raise ValueError()
            except:
                messagebox.showwarning("Dato Inválido", "El tiempo de reparación debe ser un número positivo.")
                return

            datos_intervencion = {
                'equipo_id': id_e,
                'fecha': f_recepcion_iso,
                'tipo': var_tipo.get(),
                'detalle': txt_trab,
                'condicion': var_cond.get(),
                'estado_equipo': var_est.get(),
                'deficiencia': txt_def,
                'trabajo': txt_trab,
                'observaciones': txt_obs,
                'fecha_entrega': f_entrega_iso,
                'servicio_ht': eq_data.get('servicio', ''),
                'tipo_ht': e_tipo_eq.get(),
                'repuesto_usado': r_usado,
                'repuesto_nombre': r_nombre,
                'repuesto_cantidad': r_cant,
                'fecha_programada': f_prog_val,
                'realizado_por': self.usuario_actual.get("nombre_completo", "Administrador"),
                'hora_entrega': datetime.now().strftime("%H:%M"),
                'tiempo_reparacion': tiempo_rep_val
            }

            # 1. Actualizar memoria y caché de forma instantánea (0 ms)
            eq_data.setdefault("historial_intervenciones", []).insert(0, datos_intervencion)
            if r_usado and rep_rec:
                rep_rec["cantidad"] = max(0, int(rep_rec.get("cantidad", 0)) - r_cant)

            guardar_cache_local_datos(self.datos)
            self._calendario_sucio = True
            
            # Refrescar vistas en pantalla
            self.vistas["Historial"].refrescar_datos()
            self.vistas["Repuestos"].refrescar_datos()
            self.vistas["Cronograma"].refrescar_datos()
            self.vistas["Inventario"].refrescar_datos()

            # 2. Guardar en PostgreSQL en segundo plano sin congelar
            def _guardar_mantenimiento_db(datos, id_equipo, r_usado_flag, r_nom, r_cantidad, rep_tipo):
                conn = obtener_conexion()
                if conn:
                    try:
                        cur = conn.cursor()
                        cur.execute("""
                            INSERT INTO historial_intervenciones 
                            (equipo_id, fecha, tipo, detalle, condicion, estado_equipo, deficiencia, trabajo, observaciones, fecha_entrega, servicio_ht, tipo_ht, repuesto_usado, repuesto_nombre, repuesto_cantidad, fecha_programada, realizado_por, hora_entrega, tiempo_reparacion) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            id_equipo, datos['fecha'], datos['tipo'], datos['detalle'], datos['condicion'], datos['estado_equipo'],
                            datos['deficiencia'], datos['trabajo'], datos['observaciones'], datos['fecha_entrega'], datos['servicio_ht'],
                            datos['tipo_ht'], r_usado_flag, r_nom, r_cantidad, datos['fecha_programada'], datos['realizado_por'],
                            datos['hora_entrega'], datos['tiempo_reparacion']
                        ))
                        
                        if r_usado_flag and rep_tipo:
                            cur.execute("""
                                UPDATE repuestos 
                                SET cantidad = GREATEST(0, cantidad - %s) 
                                WHERE nombre_repuesto = %s AND tipo_equipo = %s
                            """, (r_cantidad, r_nom, rep_tipo))
                        
                        conn.commit()
                        cur.close()
                        conn.close()
                    except Exception as db_err:
                        print(f"[WARN] Error al insertar en PostgreSQL, guardando offline: {db_err}")
                        guardar_mantenimiento_offline_cola(datos)
                else:
                    guardar_mantenimiento_offline_cola(datos)

            rep_tipo_str = rep_rec.get("tipo_equipo") if (r_usado and rep_rec) else None
            ejecutar_en_segundo_plano(_guardar_mantenimiento_db, datos_intervencion, id_e, r_usado, r_nombre, r_cant, rep_tipo_str)

            # Re-calcular y refrescar los mantenimientos pendientes en el combobox c_pendiente
            opciones_pendientes.clear()
            opciones_pendientes.append("Ninguno / Mantenimiento Aparte")
            pendientes_data.clear()
            
            for eq in self.datos.get("equipos", []):
                if eq.get("estado") == "Baja":
                    continue
                
                proximos = calcular_proximos_mantenimientos(eq, cantidad=1, hoy=self.hoy)
                if not proximos:
                    continue
                
                f_prox = proximos[0]
                fecha_inicio_mes = date(f_prox.year, f_prox.month, 1)
                fecha_activacion = fecha_inicio_mes - timedelta(days=5)
                
                if self.hoy >= fecha_activacion:
                    limit_date = date(f_prox.year, f_prox.month, 1) + relativedelta(months=+1, day=5)
                    es_v = self.hoy > limit_date
                    if es_v:
                        lbl_op = f"⚠️ [VENCIDO] {eq['id']} - {eq['nombre']} (Venció: {f_prox.strftime('%Y-%m-%d')} - {nombres_meses[f_prox.month - 1]})"
                    else:
                        lbl_op = f"{eq['id']} - {eq['nombre']} (Vence: {f_prox.strftime('%Y-%m-%d')} - {nombres_meses[f_prox.month - 1]})"
                    opciones_pendientes.append(lbl_op)
                    pendientes_data.append({
                        "label": lbl_op,
                        "eq_id": eq["id"],
                        "eq_nombre": eq["nombre"],
                        "f_prox": f_prox,
                        "es_vencido": es_v
                    })
            
            c_pendiente.configure(values=opciones_pendientes)
            c_pendiente.set("Ninguno / Mantenimiento Aparte")

            messagebox.showinfo("Éxito", "✅ Intervención guardada exitosamente.")
            btn_guardar.configure(state="disabled")
            generar_excel_ht(eq_data)


        def generar_excel_ht(eq_data):
            ruta_plantilla_ht = obtener_ruta_plantilla("plantilla_trabajo.xlsx")
            
            if not os.path.exists(ruta_plantilla_ht):
                messagebox.showerror("Error", f"No se encontró la plantilla en:\n{ruta_plantilla_ht}")
                return
                
            try:
                wb = openpyxl.load_workbook(ruta_plantilla_ht)
                ws = wb.active
                
                def w(celda, valor, is_check=False):
                    try:
                        ws[celda] = valor
                        if is_check and valor == 'X':
                            ws[celda].font = Font(name='Calibri', bold=True, color='000000', size=14)
                            ws[celda].alignment = Alignment(horizontal='center', vertical='center')
                    except: 
                        pass
                        
                def w_texto(celda, valor):
                    try:
                        ws[celda] = valor
                        ws[celda].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    except: 
                        pass
                        
                w('F11', eq_data.get('area', ''))
                w('AA11', eq_data.get('servicio', ''))
                w('S21', e_tipo_eq.get())
                w('J15', eq_data.get('nombre', ''))
                w('AE15', eq_data.get('id', ''))
                w('E17', eq_data.get('procedencia', ''))
                w('AB17', eq_data.get('anio_fab', ''))
                w('E19', eq_data.get('marca', ''))
                w('AB19', eq_data.get('fabricante', ''))
                w('F21', eq_data.get('modelo', ''))
                w('AG21', eq_data.get('numero_serie', ''))
                w('M23', e_fecha_rec.get())
                w('I62', f"{e_fecha_ent.get()}  {datetime.now().strftime('%H:%M')}")
                
                # CORRECCIÓN DE "CONDICIÓN ENCONTRADA" VACÍA (Restaurado desde tu código original)
                if var_cond.get() == "Óptimo":
                    w('P26', 'X', True)
                elif var_cond.get() == "Aceptable":
                    w('W26', 'X', True)
                elif var_cond.get() == "Crítica":
                    w('AC26', 'X', True)
                elif var_cond.get() == "Inoperante":
                    w('AJ26', 'X', True)
                elif var_cond.get() == "F/Servicio":
                    w('AP26', 'X', True)

                # Restauración de Estados del Equipo
                if var_est.get() == "Óptimo":
                    w('O29', 'X', True)
                elif var_est.get() == "Bueno":
                    w('U29', 'X', True)
                elif var_est.get() == "Regular":
                    w('AB29', 'X', True)
                elif var_est.get() == "Malo":
                    w('AH29', 'X', True)
                elif var_est.get() == "Obsoleto":
                    w('AO29', 'X', True)

                # Restauración Tipo Mantenimiento
                if var_tipo.get() == "Preventivo":
                    w('Q43', 'X', True)
                else:
                    w('AL43', 'X', True)

                w_texto('B33', e_def.get("1.0", "end-1c").strip())
                w_texto('B47', e_trab.get("1.0", "end-1c").strip())
                w_texto('B53', e_obs.get("1.0", "end-1c").strip())
                
                timestamp_seguro = datetime.now().strftime('%H%M%S')
                nombre_salida = f"HT_{eq_data['id']}_{datetime.now().strftime('%Y%m%d')}_{timestamp_seguro}.xlsx"
                area_name = eq_data.get("area", "General")
                area_folder = "".join([c for c in area_name if c.isalnum() or c==' ']).strip()
                dir_mantenimiento = os.path.join(CARPETAS["areas"], area_folder, "mantenimientos")
                os.makedirs(dir_mantenimiento, exist_ok=True)
                ruta_salida = os.path.join(dir_mantenimiento, nombre_salida)
                
                wb.save(ruta_salida)
                ruta_ht_excel_act.set(os.path.abspath(ruta_salida))
                f_acciones.pack(fill="x", pady=10, padx=20)
                
            except Exception as e:
                messagebox.showerror("Error Excel", f"Fallo al generar hoja:\n{e}")

        def abrir_excel_ht():
            try: 
                os.startfile(ruta_ht_excel_act.get())
            except Exception as e: 
                messagebox.showerror("Error", str(e))

        def exportar_pdf_ht():
            ruta_pdf = filedialog.asksaveasfilename(initialdir=dir_mantenimiento, initialfile=f"Hoja_Trabajo.pdf", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
            if not ruta_pdf: 
                return
            
            messagebox.showinfo("Exportando", "Por favor espera, generando PDF...")
            if exportar_excel_a_pdf(ruta_ht_excel_act.get(), ruta_pdf, rango_impresion="$A$1:$AR$67"):
                os.startfile(os.path.abspath(ruta_pdf))
            else:
                messagebox.showerror("Error", "Fallo al generar el PDF.")

        btn_guardar = ctk.CTkButton(sf, text="Guardar Registro y Generar Hoja", height=45, fg_color=C_BLUE, hover_color=C_BLUE_HOVER, font=ctk.CTkFont(weight="bold"), command=guardar)
        btn_guardar.pack(pady=20, padx=20, fill="x")
        
        ctk.CTkButton(f_acciones, text="📄 Abrir Hoja de Trabajo (Excel)", font=ctk.CTkFont(weight="bold"), fg_color=C_BLUE, hover_color=C_BLUE_HOVER, command=abrir_excel_ht).pack(side="left", expand=True, padx=5)
        ctk.CTkButton(f_acciones, text="⬇ Exportar a PDF", font=ctk.CTkFont(weight="bold"), fg_color=C_PURPLE, hover_color="#963ECA", command=exportar_pdf_ht).pack(side="left", expand=True, padx=5)

# ========================================================
# ARRANQUE OFICIAL DE LA APLICACIÓN
# ========================================================
if __name__ == "__main__":
    inicializar_bd()
    inicializar_usuarios()
    
    login_win = VentanaLogin()
    login_win.mainloop()
    
    if getattr(login_win, "usuario_autenticado", None):
        app = SistemaMantenimiento(usuario=login_win.usuario_autenticado, contexto_sede=getattr(login_win, "contexto_sede", None))
        app.mainloop()