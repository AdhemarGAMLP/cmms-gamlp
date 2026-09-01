# generador_repuestos_excel.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from copy import copy
import io, base64, os, tempfile
from datetime import datetime

def obtener_ruta_plantilla_repuestos():
    posibles = [
        os.path.join(os.path.dirname(__file__), "plantillas", "plantilla_repuestos.xlsx"),
        os.path.join("plantillas", "plantilla_repuestos.xlsx"),
        "c:/Users/HP/Desktop/CMMS_GAMLP/plantillas/plantilla_repuestos.xlsx"
    ]
    for p in posibles:
        if os.path.exists(p):
            return p
    return "plantillas/plantilla_repuestos.xlsx"

def generar_excel_repuestos_wb(lista_repuestos, tipo="Stock"):
    """
    Genera el reporte de repuestos respetando 100% el diseño original de la plantilla.
    Escribe directamente sobre las filas de la plantilla y, si hay más filas que las
    pre-formateadas (fila 8 en adelante), clona el formato exacto de la fila 8.
    """
    ruta_plantilla = obtener_ruta_plantilla_repuestos()
    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active

    # 1. Título dinámico en E4 (E4:J5 combinado)
    anio_act = datetime.now().year
    tipo_str = str(tipo).strip().lower()
    if tipo_str in ["stock", "inventario", "en stock"]:
        titulo = f"Lista de Repuestos en Inventario - {anio_act}"
    else:
        titulo = f"Lista de Repuestos Requeridos - {anio_act}"
        
    ws["E4"].value = titulo

    temp_img_files = []
    start_row = 8

    # 2. Escribir datos sobre las filas de la plantilla
    for idx, r in enumerate(lista_repuestos):
        fila_idx = start_row + idx

        # Si la fila es mayor a las pre-existentes en la plantilla (fila 8),
        # clonamos el formato exacto de la fila 8 (estilos, bordes, fuentes, alineaciones)
        if fila_idx > 8:
            if 8 in ws.row_dimensions and ws.row_dimensions[8].height is not None:
                ws.row_dimensions[fila_idx].height = ws.row_dimensions[8].height
                
            # Combinar K{r}:N{r} (Costo Total) y O{r}:V{r} (Fotografía) si no están ya combinadas
            ws.merge_cells(start_row=fila_idx, start_column=11, end_row=fila_idx, end_column=14)
            ws.merge_cells(start_row=fila_idx, start_column=15, end_row=fila_idx, end_column=22)

            for c in range(2, 23):
                src_cell = ws.cell(row=8, column=c)
                dst_cell = ws.cell(row=fila_idx, column=c)
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = copy(src_cell.number_format)
                    dst_cell.protection = copy(src_cell.protection)
                    dst_cell.alignment = copy(src_cell.alignment)

        # Valores a escribir
        cant = int(r.get("cantidad", 0) or 0)
        costo_u = float(r.get("costo", 0) or 0)
        costo_tot = cant * costo_u

        # Escribir solo los valores manteniendo el formato intacto de la plantilla
        ws.cell(row=fila_idx, column=2, value=idx + 1)                                          # B: N°
        ws.cell(row=fila_idx, column=3, value=r.get("red_salud_nombre") or "")                 # C: RED
        ws.cell(row=fila_idx, column=4, value=r.get("centro_salud_nombre") or "")              # D: CENTRO DE SALUD
        ws.cell(row=fila_idx, column=5, value=r.get("area") or "")                             # E: AREA
        ws.cell(row=fila_idx, column=6, value=r.get("nombre_repuesto") or "")                  # F: NOMBRE DEL REPUESTO
        ws.cell(row=fila_idx, column=7, value=cant)                                            # G: CANTIDAD
        ws.cell(row=fila_idx, column=8, value=r.get("marca") or "")                            # H: MARCA
        ws.cell(row=fila_idx, column=9, value=r.get("modelo") or r.get("modelo_parte") or "") # I: MODELO
        ws.cell(row=fila_idx, column=10, value=costo_u)                                        # J: COSTO UNITARIO
        ws.cell(row=fila_idx, column=11, value=costo_tot)                                      # K (K:N): COSTO TOTAL

        # Formatos numéricos para costos
        ws.cell(row=fila_idx, column=10).number_format = '#,##0.00'
        ws.cell(row=fila_idx, column=11).number_format = '#,##0.00'

        # 3. Procesamiento de Fotografía en O{fila_idx}
        foto_data = r.get("foto")
        img_insertada = False

        if foto_data and str(foto_data).strip():
            try:
                pil_img = None
                foto_str = str(foto_data).strip()
                if foto_str.startswith("data:image") and "base64," in foto_str:
                    b64_clean = foto_str.split("base64,")[1]
                    img_bytes = base64.b64decode(b64_clean)
                    pil_img = PILImage.open(io.BytesIO(img_bytes))
                elif os.path.exists(foto_str):
                    pil_img = PILImage.open(foto_str)

                if pil_img:
                    if pil_img.mode != "RGB":
                        pil_img = pil_img.convert("RGB")
                    
                    # Escalar manteniendo proporción: max 130x55 px
                    max_w, max_h = 130, 55
                    orig_w, orig_h = pil_img.size
                    ratio = min(max_w / orig_w, max_h / orig_h)
                    new_w, new_h = max(1, int(orig_w * ratio)), max(1, int(orig_h * ratio))
                    
                    pil_img_resized = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                    
                    tf = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                    pil_img_resized.save(tf.name, format="PNG")
                    tf.close()
                    temp_img_files.append(tf.name)

                    openpyxl_img = OpenpyxlImage(tf.name)
                    openpyxl_img.anchor = f"O{fila_idx}"
                    ws.add_image(openpyxl_img)
                    
                    ws.row_dimensions[fila_idx].height = 55
                    img_insertada = True
            except Exception as ex:
                print(f"[WARN] Error insertando foto en fila {fila_idx}: {ex}")

        if not img_insertada:
            # Si no hay imagen, conservar la altura de la plantilla
            if fila_idx not in ws.row_dimensions or ws.row_dimensions[fila_idx].height is None or ws.row_dimensions[fila_idx].height < 20:
                ws.row_dimensions[fila_idx].height = 20

    return wb, temp_img_files

def guardar_excel_repuestos(lista_repuestos, tipo="Stock", ruta_salida="repuestos.xlsx"):
    wb, temp_files = generar_excel_repuestos_wb(lista_repuestos, tipo=tipo)
    wb.save(ruta_salida)
    for tf in temp_files:
        try: os.remove(tf)
        except: pass
    return ruta_salida
